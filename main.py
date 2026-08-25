"""
재고 관리 시스템 - main.py
FastAPI 기반 / SQLite ↔ Supabase(PostgreSQL) 겸용 / 지점별 로그인
"""
import os
import socket
import json
import httpx
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Optional

from fastapi import FastAPI, Request, Form, File, UploadFile, Cookie, HTTPException, Depends, Header, Query
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv
import openpyxl

load_dotenv("config/settings.env.example")
load_dotenv(".env")

from db import get_conn, pk_column  # noqa: E402

SERVER_PORT = int(os.getenv("SERVER_PORT", "28000"))
QR_DIR = "./qr_codes"

app = FastAPI(title="포포즈 발주_재고", version="1.2.0")

NOTIFICATION_TYPES = {
    "qr_raw_mismatch":     {"label": "재고 불일치 알림",      "desc": "QR-RAW 재고 불일치 시 (3시간마다 검사)"},
    "vendor_eval_missing": {"label": "거래처평가 미제출 알림", "desc": "거래처 미평가 지점 Teams 웹훅 연동"},
    "purchase_new":        {"label": "발주내역 알림",         "desc": "새 발주 등록 시"},
    "restock_missing":     {"label": "미입고 알림",           "desc": "(준비 중)"},
    "safety_stock_change": {"label": "안전재고 변경 알림",     "desc": "(준비 중)"},
}

TEAMS_ALERT_TYPES = {
    "unsubmitted_reminder": {
        "label": "거래처평가 미제출 알림",
        "toggle_key": "unsubmitted_reminder_enabled",
        "toggle_route": "/master/toggle-unsubmitted-reminder",
        "test_route": "/master/teams-webhook/test-unsubmitted-reminder",
        "test_js_func": "testUnsubmittedReminder",
        "on_desc": "매월 5,10,15,20,25,30일 자동발송",
        "off_desc": "자동발송 안 함",
        "default_enabled": True,
    },
    "qr_raw_mismatch_teams": {
        "label": "재고 불일치 Teams 알림",
        "toggle_key": "qr_raw_mismatch_teams_enabled",
        "toggle_route": "/master/toggle-qr-raw-mismatch-teams",
        "test_route": "/master/teams-webhook/test-qr-raw-mismatch",
        "test_js_func": "testQrRawMismatch",
        "on_desc": "3시간마다 Teams 발송",
        "off_desc": "Teams 발송 안 함 (웹푸시는 별개로 계속 동작)",
        "default_enabled": False,
    },
    "purchase_new_teams": {
        "label": "발주내역 Teams 알림",
        "toggle_key": "purchase_new_teams_enabled",
        "toggle_route": "/master/toggle-purchase-new-teams",
        "test_route": "/master/teams-webhook/test-purchase-new",
        "test_js_func": "testPurchaseNew",
        "on_desc": "10분마다 신규 발주 검사 후 Teams 발송",
        "off_desc": "Teams 발송 안 함 (웹푸시는 별개로 계속 동작)",
        "default_enabled": False,
    },
}

from auth.login import (  # noqa: E402
    init_auth_db, authenticate, create_session, get_session,
    delete_session, create_auto_login_token, get_auto_login_info,
    get_branches, add_branch, delete_branch, update_branch_account
)
init_auth_db()


# ── 공통 함수 ──────────────────────────────────────────

async def fetch_purchase_history(branch_name: str = None, limit: int = 200):
    """별도 Supabase 프로젝트(purchase_history)에서 발주내역 조회 (읽기 전용, anon 키 사용)."""
    supabase_url = os.environ.get("PURCHASE_SUPABASE_URL", "")
    supabase_key = os.environ.get("PURCHASE_SUPABASE_ANON_KEY", "")
    if not supabase_url or not supabase_key:
        return [], "발주내역 조회 설정이 되어있지 않습니다. 관리자에게 문의하세요."

    params = {
        "select": "*",
        "order": "registered_at.desc",
        "limit": str(limit)
    }
    if branch_name:
        params["branch"] = f"eq.{branch_name}"

    headers = {
        "apikey": supabase_key,
        "Authorization": f"Bearer {supabase_key}"
    }

    try:
        async with httpx.AsyncClient() as client:
            res = await client.get(
                f"{supabase_url}/rest/v1/purchase_history",
                params=params,
                headers=headers,
                timeout=15.0
            )
        if res.status_code != 200:
            return [], f"조회 실패 (status {res.status_code})"
        return res.json(), None
    except Exception as e:
        return [], f"조회 중 오류: {str(e)}"

async def delete_purchase_history_rows(order_ids: list):
    """service_role 키로 purchase_history 특정 행들을 삭제 (RLS 우회, 서버 전용)."""
    supabase_url = os.environ.get("PURCHASE_SUPABASE_URL", "")
    service_key = os.environ.get("PURCHASE_SUPABASE_SERVICE_ROLE_KEY", "")
    if not supabase_url or not service_key or not order_ids:
        return False, "설정 또는 대상이 없습니다."

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Prefer": "return=representation"
    }

    try:
        async with httpx.AsyncClient() as client:
            for oid in order_ids:
                params = {"order_id": f"eq.{oid}"}
                print(f"[DEBUG delete] params={params}")
                res = await client.delete(
                    f"{supabase_url}/rest/v1/purchase_history",
                    params=params,
                    headers=headers,
                    timeout=15.0
                )
                print(f"[DEBUG delete] status={res.status_code}")
                print(f"[DEBUG delete] body_response={res.text}")
                if res.status_code not in (200, 204):
                    return False, f"삭제 요청 실패 (order_id={oid}, status {res.status_code}, body: {res.text})"
        return True, None
    except Exception as e:
        print(f"[DEBUG delete] exception={str(e)}")
        return False, f"삭제 중 오류: {str(e)}"


async def update_purchase_history_quantity(order_id: str, quantity: float):
    """service_role 키로 특정 발주건의 수량만 수정 (RLS 우회, 서버 전용)."""
    supabase_url = os.environ.get("PURCHASE_SUPABASE_URL", "")
    service_key = os.environ.get("PURCHASE_SUPABASE_SERVICE_ROLE_KEY", "")
    if not supabase_url or not service_key:
        return False, "설정이 되어있지 않습니다."

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation"
    }
    params = {"order_id": f"eq.{order_id}"}
    body = {"quantity": quantity}

    print(f"[DEBUG update] url={supabase_url}/rest/v1/purchase_history")
    print(f"[DEBUG update] params={params}")
    print(f"[DEBUG update] body={body}")
    print(f"[DEBUG update] service_key_len={len(service_key)}")

    try:
        async with httpx.AsyncClient() as client:
            res = await client.patch(
                f"{supabase_url}/rest/v1/purchase_history",
                params=params,
                headers=headers,
                json=body,
                timeout=15.0
            )

        if res.status_code not in (200, 204):
            return False, f"수정 실패 (status {res.status_code}, body: {res.text})"
        return True, None
    except Exception as e:
        print(f"[DEBUG update] exception={str(e)}")
        return False, f"수정 중 오류: {str(e)}"
    

@app.get("/api/debug/check-service-key")
async def debug_check_service_key(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    import base64
    import json as json_lib

    def decode_jwt_role(token):
        try:
            parts = token.split(".")
            if len(parts) < 2:
                return "invalid_format"
            payload = parts[1]
            padding = "=" * (-len(payload) % 4)
            decoded = base64.urlsafe_b64decode(payload + padding)
            data = json_lib.loads(decoded)
            return data.get("role", "role_field_missing")
        except Exception as e:
            return f"decode_error: {str(e)}"

    key = os.environ.get("PURCHASE_SUPABASE_SERVICE_ROLE_KEY", "")
    anon_key = os.environ.get("PURCHASE_SUPABASE_ANON_KEY", "")

    return JSONResponse(content={
        "service_key_role": decode_jwt_role(key) if key else "(비어있음)",
        "service_key_length": len(key),
        "anon_key_role": decode_jwt_role(anon_key) if anon_key else "(비어있음)",
        "same_as_anon": key == anon_key
    })

async def delete_old_purchase_history(days: int = 21):
    """등록일 기준 N일 경과한 발주내역 전체 삭제 (cron 전용)."""
    supabase_url = os.environ.get("PURCHASE_SUPABASE_URL", "")
    service_key = os.environ.get("PURCHASE_SUPABASE_SERVICE_ROLE_KEY", "")
    if not supabase_url or not service_key:
        return 0, "설정이 되어있지 않습니다."

    from datetime import datetime, timedelta, timezone
    cutoff = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()

    headers = {
        "apikey": service_key,
        "Authorization": f"Bearer {service_key}",
        "Prefer": "return=representation"
    }
    params = {"registered_at": f"lt.{cutoff}"}

    try:
        async with httpx.AsyncClient() as client:
            res = await client.delete(
                f"{supabase_url}/rest/v1/purchase_history",
                params=params,
                headers=headers,
                timeout=30.0
            )
        if res.status_code not in (200, 204):
            return 0, f"삭제 실패 (status {res.status_code})"
        deleted = res.json() if res.text else []
        return len(deleted), None
    except Exception as e:
        return 0, f"삭제 중 오류: {str(e)}"

@app.post("/purchase-history/delete")
async def purchase_history_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    order_ids = data.get("order_ids", [])
    if not order_ids:
        return JSONResponse(status_code=400, content={"detail": "삭제할 항목이 없습니다."})

    success, err = await delete_purchase_history_rows(order_ids)
    if not success:
        return JSONResponse(status_code=400, content={"detail": err})
    return JSONResponse(content={"status": "ok", "deleted_count": len(order_ids)})


@app.post("/purchase-history/update-quantity")
async def purchase_history_update_quantity(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    order_id = data.get("order_id", "")
    quantity = data.get("quantity")
    if not order_id or quantity is None:
        return JSONResponse(status_code=400, content={"detail": "필수 항목이 누락되었습니다."})

    try:
        quantity = float(quantity)
    except (TypeError, ValueError):
        return JSONResponse(status_code=400, content={"detail": "수량은 숫자여야 합니다."})

    success, err = await update_purchase_history_quantity(order_id, quantity)
    if not success:
        return JSONResponse(status_code=400, content={"detail": err})
    return JSONResponse(content={"status": "ok"})


@app.get("/api/cron/cleanup-purchase-history")
async def cron_cleanup_purchase_history(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    deleted_count, err = await delete_old_purchase_history(days=21)
    if err:
        return JSONResponse(status_code=500, content={"detail": err})
    return JSONResponse(content={"deleted_count": deleted_count})

def send_push_notification(branch_code: str, title: str, body: str, event_type: str, url: str = "/"):
    """특정 지점의 등록된 기기 중, 해당 알림 종류를 켜놓은 기기에만 웹 푸시 발송 + notification_events 기록."""
    from pywebpush import webpush, WebPushException
    import json as json_lib

    conn = get_conn()
    conn.execute(
        "INSERT INTO notification_events (event_type, branch_code, title, body) VALUES (?, ?, ?, ?)",
        (event_type, branch_code, title, body)
    )
    conn.commit()

    subs = conn.execute(
        "SELECT * FROM push_subscriptions WHERE branch_code=?", (branch_code,)
    ).fetchall()

    vapid_private_key = os.environ.get("VAPID_PRIVATE_KEY", "")
    vapid_claim_email = os.environ.get("VAPID_CLAIM_EMAIL", "mailto:admin@example.com")

    sent, failed, skipped = 0, 0, 0
    for sub in subs:
        all_setting = conn.execute(
            "SELECT enabled FROM push_notification_settings WHERE subscription_id=? AND notification_type='all'",
            (sub["id"],)
        ).fetchone()
        type_setting = conn.execute(
            "SELECT enabled FROM push_notification_settings WHERE subscription_id=? AND notification_type=?",
            (sub["id"], event_type)
        ).fetchone()

        all_enabled = bool(all_setting["enabled"]) if all_setting else True
        type_enabled = bool(type_setting["enabled"]) if type_setting else True

        if not all_enabled or not type_enabled:
            skipped += 1
            continue

        try:
            webpush(
                subscription_info={
                    "endpoint": sub["endpoint"],
                    "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
                },
                data=json_lib.dumps({"title": title, "body": body, "url": url}, ensure_ascii=False),
                vapid_private_key=vapid_private_key,
                vapid_claims={"sub": vapid_claim_email}
            )
            sent += 1
        except WebPushException:
            failed += 1

    conn.close()
    return sent, failed


@app.post("/api/push/subscribe")
async def push_subscribe(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "로그인이 필요합니다."})

    data = await request.json()
    endpoint = data.get("endpoint", "")
    keys = data.get("keys", {})
    p256dh = keys.get("p256dh", "")
    auth = keys.get("auth", "")

    if not endpoint or not p256dh or not auth:
        return JSONResponse(status_code=400, content={"detail": "구독 정보가 올바르지 않습니다."})

    branch_code = user.get("branch_code") or "master"

    conn = get_conn()
    existing = conn.execute("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,)).fetchone()
    if existing:
        conn.execute(
            "UPDATE push_subscriptions SET branch_code=?, p256dh=?, auth=? WHERE endpoint=?",
            (branch_code, p256dh, auth, endpoint)
        )
        subscription_id = existing["id"]
    else:
        conn.execute(
            "INSERT INTO push_subscriptions (branch_code, endpoint, p256dh, auth) VALUES (?, ?, ?, ?)",
            (branch_code, endpoint, p256dh, auth)
        )
        new_row = conn.execute("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,)).fetchone()
        subscription_id = new_row["id"]

    # 신규/기존 구독 모두, 아직 설정이 없는 알림 종류는 기본값(켜짐)으로 초기화
    all_types = ["all"] + list(NOTIFICATION_TYPES.keys())
    for ntype in all_types:
        exists_setting = conn.execute(
            "SELECT id FROM push_notification_settings WHERE subscription_id=? AND notification_type=?",
            (subscription_id, ntype)
        ).fetchone()
        if not exists_setting:
            conn.execute(
                "INSERT INTO push_notification_settings (subscription_id, notification_type, enabled) VALUES (?, ?, TRUE)",
                (subscription_id, ntype)
            )

    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok", "subscription_id": subscription_id})


@app.get("/api/push/vapid-public-key")
async def push_vapid_public_key():
    return JSONResponse(content={"key": os.environ.get("VAPID_PUBLIC_KEY", "")})

@app.get("/api/push/settings")
async def push_get_settings(endpoint: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "로그인이 필요합니다."})

    conn = get_conn()
    sub = conn.execute("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,)).fetchone()
    if not sub:
        conn.close()
        return JSONResponse(content={"subscribed": False, "settings": {}})

    rows = conn.execute(
        "SELECT notification_type, enabled FROM push_notification_settings WHERE subscription_id=?",
        (sub["id"],)
    ).fetchall()
    conn.close()

    settings = {r["notification_type"]: bool(r["enabled"]) for r in rows}
    return JSONResponse(content={"subscribed": True, "settings": settings})


@app.post("/api/push/settings")
async def push_update_settings(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return JSONResponse(status_code=401, content={"detail": "로그인이 필요합니다."})

    data = await request.json()
    endpoint = data.get("endpoint", "")
    notification_type = data.get("notification_type", "")
    enabled = bool(data.get("enabled", True))

    if not endpoint or not notification_type:
        return JSONResponse(status_code=400, content={"detail": "필수 항목이 누락되었습니다."})

    conn = get_conn()
    sub = conn.execute("SELECT id FROM push_subscriptions WHERE endpoint=?", (endpoint,)).fetchone()
    if not sub:
        conn.close()
        return JSONResponse(status_code=404, content={"detail": "구독 정보를 찾을 수 없습니다."})

    conn.execute(
        "UPDATE push_notification_settings SET enabled=? WHERE subscription_id=? AND notification_type=?",
        (enabled, sub["id"], notification_type)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})

@app.get("/purchase-history", response_class=HTMLResponse)
async def purchase_history_page(
    session_token: str = Cookie(default=None),
    branch_filter: str = "",
    vendor_filter: str = "",
    product_filter: str = "",
    date_from: str = "",
    date_to: str = ""
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    is_master = user["role"] == "master"
    is_master = user["role"] == "master"
    if is_master:
        target_branch = branch_filter or None
    else:
        my_branch_code = user.get("branch_code")
        my_branches = get_branches(branch_type='branch')
        my_branch_info = next((b for b in my_branches if b["branch_code"] == my_branch_code), None)
        target_branch = my_branch_info["branch_name"] if my_branch_info else my_branch_code

    rows, err = await fetch_purchase_history(branch_name=target_branch)

    if not err and rows:
        if vendor_filter:
            rows = [r for r in rows if vendor_filter.lower() in (r.get("vendor") or "").lower()]
        if product_filter:
            rows = [r for r in rows if product_filter.lower() in (r.get("product_name") or "").lower()]
        if date_from:
            rows = [r for r in rows if (r.get("order_datetime") or r.get("registered_at") or "") >= date_from]
        if date_to:
            rows = [r for r in rows if (r.get("order_datetime") or r.get("registered_at") or "") <= date_to + "T23:59:59"]

    branch_filter_html = ""
    if is_master:
        branches = get_branches(branch_type='branch')
        options = '<option value="">전체 지점</option>'
        for b in branches:
            sel = 'selected' if b["branch_name"] == branch_filter else ''
            options += f'<option value="{b["branch_name"]}" {sel}>{b["branch_name"]}</option>'
        branch_filter_html = f"""
        <div>
          <label style="font-size:12px;color:#888;">지점</label>
          <select name="branch_filter">{options}</select>
        </div>
        """

    content_filter_html = f"""
    <div class="card">
      <form method="get" action="/purchase-history">
        <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(150px, 1fr));gap:10px;margin-bottom:10px;">
          {branch_filter_html}
          <div>
            <label style="font-size:12px;color:#888;">거래처</label>
            <input type="text" name="vendor_filter" value="{vendor_filter}" placeholder="거래처명 검색">
          </div>
          <div>
            <label style="font-size:12px;color:#888;">상품명</label>
            <input type="text" name="product_filter" value="{product_filter}" placeholder="상품명 검색">
          </div>
          <div>
            <label style="font-size:12px;color:#888;">시작일</label>
            <input type="date" name="date_from" value="{date_from}">
          </div>
          <div>
            <label style="font-size:12px;color:#888;">종료일</label>
            <input type="date" name="date_to" value="{date_to}">
          </div>
        </div>
        <div style="display:flex;gap:8px;">
          <button class="btn" type="submit">검색</button>
          <a href="/purchase-history" style="padding:10px 14px;background:#eee;
             border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
        </div>
      </form>
    </div>
    """

    master_toolbar_html = ""
    if is_master:
        master_toolbar_html = """
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;">
          <button type="button" class="btn" id="phSelectAllBtn" style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
          <button type="button" class="btn btn-red" id="phDeleteBtn" style="font-size:12px;padding:6px 12px;">선택 삭제</button>
        </div>
        <div id="phSelectedSummary" style="display:none;font-size:12px;color:#1E2761;background:#EFF6FF;padding:6px 10px;border-radius:6px;margin-bottom:10px;"></div>
        """

    total_count = len(rows) if not err else 0
    total_qty = 0
    if not err and rows:
        for r in rows:
            q = r.get("quantity")
            if isinstance(q, (int, float)):
                total_qty += q

    summary_bar_html = f"""
    <div class="card" style="display:flex;gap:24px;padding:12px 16px;">
      <div><span style="color:#888;font-size:12px;">건수</span> <b style="font-size:16px;">{total_count}건</b></div>
      <div><span style="color:#888;font-size:12px;">수량 합계</span> <b style="font-size:16px;">{total_qty:g}</b></div>
    </div>
    """

    if err:
        rows_html = f'<tr><td colspan="7" style="text-align:center;color:#888;padding:24px;">{err}</td></tr>'
    elif not rows:
        rows_html = '<tr><td colspan="7" style="text-align:center;color:#888;padding:24px;">발주 내역이 없습니다.</td></tr>'
    else:
        rows_html = ""
        for r in rows:
            order_dt_raw = r.get("order_datetime") or r.get("registered_at") or "-"
            if order_dt_raw != "-":
                try:
                    from datetime import timezone as _tz, timedelta as _td
                    _dt = datetime.fromisoformat(str(order_dt_raw).replace("Z", "+00:00"))
                    if _dt.tzinfo is None:
                        _dt = _dt.replace(tzinfo=_tz.utc)
                    order_dt = _dt.astimezone(_tz(_td(hours=9))).strftime("%Y-%m-%d")
                except Exception:
                    order_dt = order_dt_raw[:10]
            else:
                order_dt = "-"
            status = r.get("send_status") or ""
            status_badge = '<span class="badge-green">완료</span>' if status == "완료" else '<span class="badge-red">대기</span>'
            oid = r.get("order_id", "")
            qty = r.get("quantity", "-")
            if isinstance(qty, (int, float)):
                qty = int(qty) if qty == int(qty) else qty
            check_cell = f'<td style="text-align:center;"><input type="checkbox" class="ph-check" value="{oid}" style="width:16px;height:16px;"></td>' if is_master else ""
            qty_cell = (
                f'<td style="text-align:right;"><span id="qtyDisplay_{oid}">{qty}</span> '
                f'<button class="btn" style="font-size:10px;padding:2px 6px;margin-left:4px;" onclick="editQuantity(\'{oid}\', {qty})">수정</button></td>'
                if is_master else
                f'<td style="text-align:right;">{qty}</td>'
            )
            rows_html += f"""
            <tr>
                {check_cell}
                <td style="font-size:12px;">{order_dt}</td>
                <td>{r.get('branch', '-')}</td>
                <td>{r.get('vendor', '-')}</td>
                <td>{r.get('product_name', '-')}</td>
                {qty_cell}
                <td>{status_badge}</td>
            </tr>
            """

    check_header = '<th style="width:36px;text-align:center;"><input type="checkbox" id="phAllCheck" style="width:16px;height:16px;"></th>' if is_master else ""

    content = f"""
    <h2 style="margin-bottom:16px;">📦 발주내역</h2>
    {content_filter_html}
    {summary_bar_html}
    <div class="card">
      {master_toolbar_html}
      <table>
        <thead><tr>
          {check_header}
          <th>발주일시</th><th>지점</th><th>거래처</th><th>상품명</th><th>수량</th><th>상태</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    <script>
      (function() {{
        var allCheck = document.getElementById('phAllCheck');
        var selectBtn = document.getElementById('phSelectAllBtn');
        var deleteBtn = document.getElementById('phDeleteBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.ph-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        function updateSelectedSummary() {{
          var checkedBoxes = document.querySelectorAll('.ph-check:checked');
          var summaryEl = document.getElementById('phSelectedSummary');
          if (!summaryEl) return;
          if (checkedBoxes.length === 0) {{
            summaryEl.style.display = 'none';
            return;
          }}
          var qtySum = 0;
          checkedBoxes.forEach(function(c) {{
            var qtyEl = document.getElementById('qtyDisplay_' + c.value);
            if (qtyEl) {{
              var v = parseFloat(qtyEl.innerText);
              if (!isNaN(v)) qtySum += v;
            }}
          }});
          summaryEl.style.display = 'block';
          summaryEl.innerText = '선택 ' + checkedBoxes.length + '건 / 수량 합계 ' + qtySum;
        }}
        document.querySelectorAll('.ph-check').forEach(function(c) {{
          c.addEventListener('change', updateSelectedSummary);
        }});
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ setTimeout(updateSelectedSummary, 0); }}); }}
        if (selectBtn) {{ selectBtn.addEventListener('click', function() {{ setTimeout(updateSelectedSummary, 0); }}); }}

        if (deleteBtn) {{
          deleteBtn.addEventListener('click', async function() {{
            var checked = Array.from(document.querySelectorAll('.ph-check:checked')).map(c => c.value);
            if (checked.length === 0) {{ alert('삭제할 항목을 선택하세요.'); return; }}
            if (!confirm(checked.length + '건을 삭제합니다. 되돌릴 수 없습니다. 계속할까요?')) return;
            const res = await fetch('/purchase-history/delete', {{
              method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
              body: JSON.stringify({{ order_ids: checked }})
            }});
            if (res.ok) {{ location.reload(); }} else {{
              const err = await res.json();
              alert('오류: ' + (err.detail || '삭제 실패'));
            }}
          }});
        }}
      }})();

      async function editQuantity(orderId, currentQty) {{
        const newQty = prompt('새 수량을 입력하세요:', currentQty);
        if (newQty === null) return;
        if (isNaN(parseFloat(newQty))) {{ alert('숫자를 입력하세요.'); return; }}
        const res = await fetch('/purchase-history/update-quantity', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ order_id: orderId, quantity: parseFloat(newQty) }})
        }});
        if (res.ok) {{
          document.getElementById('qtyDisplay_' + orderId).innerText = newQty;
        }} else {{
          const err = await res.json();
          alert('오류: ' + (err.detail || '수정 실패'));
        }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "purchase-history"))

def init_db():
    conn = get_conn()
    pk = pk_column()
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS inventory (
            {pk},
            branch_code TEXT NOT NULL,
            item_name TEXT NOT NULL,
            item_code TEXT NOT NULL,
            quantity INTEGER DEFAULT 0,
            last_updated TEXT,
            UNIQUE(branch_code, item_code)
        )
    """)
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS adjustment_log (
            {pk},
            branch_code TEXT,
            item_name TEXT,
            item_code TEXT,
            delta INTEGER,
            result_quantity INTEGER,
            adjusted_at TEXT
        )
    """)
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS items (
            {pk},
            branch_code TEXT NOT NULL,
            branch_name TEXT NOT NULL,
            item_name TEXT NOT NULL,
            item_code TEXT NOT NULL,
            created_at TEXT,
            UNIQUE(branch_code, item_code)
        )
    """)
    
    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS raw_inventory (
            {pk},
            branch_code TEXT NOT NULL,
            branch_name TEXT NOT NULL,
            item_name TEXT NOT NULL,
            item_code TEXT NOT NULL,
            quantity INTEGER DEFAULT 0,
            source TEXT DEFAULT 'branch',
            uploaded_at TEXT,
            UNIQUE(branch_code, item_code, source)
        )
    """)

    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS qr_init_log (
            {pk},
            branch_code TEXT,
            item_code TEXT,
            init_quantity INTEGER,
            initialized_at TEXT
        )
    """)

    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS hq_bonus_log (
            {pk},
            branch_code TEXT NOT NULL,
            item_code TEXT NOT NULL,
            last_hq_total INTEGER DEFAULT 0,
            updated_at TEXT,
            UNIQUE(branch_code, item_code)
        )
    """)

    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS scan_log (
            {pk},
            branch_code TEXT NOT NULL,
            branch_name TEXT,
            item_name TEXT,
            item_code TEXT NOT NULL,
            scan_type TEXT,
            result_quantity INTEGER,
            scanned_at TEXT,
            device_info TEXT,
            client_ip TEXT
        )
    """)

    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS scan_log_delete_history (
            {pk},
            deleted_by TEXT,
            branch_code TEXT,
            item_name TEXT,
            item_code TEXT,
            scan_type TEXT,
            original_scanned_at TEXT,
            deleted_at TEXT
        )
    """)

    conn.execute(f"""
        CREATE TABLE IF NOT EXISTS login_history (
            {pk},
            login_id TEXT NOT NULL,
            role TEXT,
            branch_code TEXT,
            device_info TEXT,
            client_ip TEXT,
            logged_in_at TEXT
        )
    """)

    conn.commit()
    conn.close()


init_db()


# ── 공통 UI 컴포넌트 ────────────────────────────────────

def render_page(content: str, user: Optional[Dict] = None, active: str = "") -> str:
    """공통 레이아웃 — 상단 타이틀 + 하단 메뉴 포함"""
    # ⚠️ 모바일 전용 테이블 글씨 축소 — f-string 중괄호 충돌 방지를 위해 별도 문자열로 조립
    mobile_table_css = (
        "@media (max-width: 480px) {"
        "table { table-layout: auto; min-width: 560px; }"
        "th { font-size: 8px; padding: 3px 2px; resize: none; min-width: 48px; }"
        "td { font-size: 8px; padding: 3px 2px; }"
        "}"
    )
    branch_name = user["branch_code"] if user and user["role"] == "branch" else ("마스터" if user else "")    
    role_badge = f'<span style="background:#4FC3F7;color:white;padding:2px 10px;border-radius:12px;font-size:12px;margin-left:8px;">{branch_name}</span>' if user else ""

    is_master = user and user.get("role") == "master"
    raw_menu_href = "/master/raw-upload" if is_master else "/raw-branch"
    vendor_eval_href = "/master/vendor-eval" if is_master else "/vendor-eval"
    menus = [
        ("dashboard", "/", "⚠️", "대시보드"),
        ("inventory", "/inventory", "📦", "재고현황"),
        ("qr", "/qr", "📷", "QR생성"),
        ("adjust", "/adjust", "✏️", "수기조정"),
        ("raw-branch", raw_menu_href, "📤", "유비플러스 재고"),
        ("scanlog", "/scan-log", "📜", "스캔이력"),
        ("vendor-eval", vendor_eval_href, "🤝", "거래처평가"),
        ("purchase-history", "/purchase-history", "📦", "발주내역"),
    ]
    if is_master:
        menus.append(("teams-webhook", "/master/teams-webhook", "🔔", "팀즈웹훅"))
        menus.append(("master", "/master", "⚙️", "마스터"))
    menu_html = ""
    for key, href, icon, label in menus:
        is_active = "background:#1E2761;color:white;" if active == key else "color:#555;"
        menu_html += f"""
        <a href="{href}" style="flex:1;text-align:center;padding:8px 0;
           text-decoration:none;font-size:12px;{is_active}border-radius:8px;">
          <div style="font-size:20px;">{icon}</div>
          <div>{label}</div>
        </a>
        """

    return f"""
    <html><head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>재고 관리 시스템</title>
      <link rel="manifest" href="/manifest.json">
      <meta name="theme-color" content="#1E2761">
      <link rel="apple-touch-icon" href="/icon-192.png">
      <meta name="apple-mobile-web-app-capable" content="yes">
      <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
      <script>
        if ('serviceWorker' in navigator) {{
          window.addEventListener('load', function() {{
            navigator.serviceWorker.register('/sw.js').then(function(reg) {{
              window.__swRegistration = reg;
            }}).catch(function() {{}});
          }});
        }}

        function urlBase64ToUint8Array(base64String) {{
          var padding = '='.repeat((4 - base64String.length % 4) % 4);
          var base64 = (base64String + padding).replace(/-/g, '+').replace(/_/g, '/');
          var rawData = window.atob(base64);
          var outputArray = new Uint8Array(rawData.length);
          for (var i = 0; i < rawData.length; ++i) {{
            outputArray[i] = rawData.charCodeAt(i);
          }}
          return outputArray;
        }}

        async function enablePushNotification() {{
          if (!('serviceWorker' in navigator) || !('PushManager' in window)) {{
            alert('이 브라우저는 알림 기능을 지원하지 않습니다.');
            return;
          }}
          const permission = await Notification.requestPermission();
          if (permission !== 'granted') {{
            alert('알림 권한이 거부되었습니다. 브라우저 설정에서 허용해주세요.');
            return;
          }}
          const reg = window.__swRegistration || await navigator.serviceWorker.ready;
          const keyRes = await fetch('/api/push/vapid-public-key');
          const keyData = await keyRes.json();
          const sub = await reg.pushManager.subscribe({{
            userVisibleOnly: true,
            applicationServerKey: urlBase64ToUint8Array(keyData.key)
          }});
          const res = await fetch('/api/push/subscribe', {{
            method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify(sub)
          }});
          if (res.ok) {{ alert('알림이 활성화되었습니다.'); }} else {{ alert('알림 등록에 실패했습니다.'); }}
        }}
      </script>
      <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
               background: #f5f7fa; padding-bottom: 80px; }}
        .topbar {{ background: #1E2761; color: white; padding: 10px 14px;
                  display: flex; justify-content: space-between; align-items: center; }}
        .content {{ max-width: 960px; margin: 0 auto; padding: 10px 14px; }}
        .card {{ background: white; border-radius: 12px; padding: 16px;
                box-shadow: 0 1px 4px rgba(0,0,0,0.08); margin-bottom: 16px; }}
        .btn {{ background: #1E2761; color: white; padding: 8px 14px;
               border: none; border-radius: 8px; cursor: pointer; font-size: 14px; }}
        .btn:hover {{ background: #2a3580; }}
        .btn-red {{ background: #EF4444; }}
        input, select {{ width: 100%; padding: 8px; border: 1px solid #ddd;
                        border-radius: 8px; font-size: 14px; margin-top: 4px; }}
        table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
        .table-scroll-wrap {{ overflow-x: auto; -webkit-overflow-scrolling: touch; }}
        th {{ background: #1E2761; color: white; padding: 6px 5px; text-align: left; font-size: 12px;
             resize: horizontal; overflow: auto; position: relative;
             min-width: 60px; white-space: nowrap; border-right: 1px solid rgba(255,255,255,0.2); }}
        td {{ padding: 6px 5px; border-bottom: 1px solid #eee; font-size: 12px;
             overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
        {mobile_table_css}
        .badge-green {{ background:#D1FAE5;color:#065F46;padding:2px 8px;
                       border-radius:10px;font-size:12px; }}
        .badge-red {{ background:#FEE2E2;color:#991B1B;padding:2px 8px;
                     border-radius:10px;font-size:12px; }}
        .bottomnav {{ position:fixed;bottom:0;left:0;right:0;
                     background:white;border-top:1px solid #eee;
                     display:flex;padding:6px 8px;z-index:100; }}
      </style>
    </head>
    <body>
      <div class="topbar">
        <span style="font-weight:bold;">📦 재고 관리 시스템{role_badge}</span>
        <a href="/logout" style="color:#aaa;font-size:13px;text-decoration:none;">로그아웃</a>
      </div>
      <div class="content">{content}</div>
      <nav class="bottomnav">{menu_html}</nav>
      <script>
      (function() {{
        // 모바일 가로스크롤을 위해 모든 table을 스크롤 컨테이너로 자동 래핑
        document.querySelectorAll('table').forEach(function(table) {{
          if (table.parentElement && table.parentElement.classList.contains('table-scroll-wrap')) return;
          var wrap = document.createElement('div');
          wrap.className = 'table-scroll-wrap';
          table.parentNode.insertBefore(wrap, table);
          wrap.appendChild(table);
        }});
      }})();
      (function() {{
        function saveColumnWidths() {{
          document.querySelectorAll('table').forEach(function(table, tIdx) {{
            var widths = [];
            table.querySelectorAll('th').forEach(function(th) {{
              widths.push(th.offsetWidth);
            }});
            localStorage.setItem('colWidths_' + window.location.pathname + '_' + tIdx, JSON.stringify(widths));
          }});
        }}
        function restoreColumnWidths() {{
          document.querySelectorAll('table').forEach(function(table, tIdx) {{
            var saved = localStorage.getItem('colWidths_' + window.location.pathname + '_' + tIdx);
            if (!saved) return;
            try {{
              var widths = JSON.parse(saved);
              var ths = table.querySelectorAll('th');
              ths.forEach(function(th, i) {{
                if (widths[i]) th.style.width = widths[i] + 'px';
              }});
            }} catch (e) {{}}
          }});
        }}
        restoreColumnWidths();
        document.querySelectorAll('th').forEach(function(th) {{
          var observer = new ResizeObserver(function() {{ saveColumnWidths(); }});
          observer.observe(th);
        }});
      }})();
      (function() {{
        var PAGE_KEY = 'filter_branch_' + window.location.pathname;
        var selects = document.querySelectorAll('select[name="filter_branch"]');
        if (selects.length === 0) return;
        var saved = localStorage.getItem(PAGE_KEY);
        var urlParams = new URLSearchParams(window.location.search);
        var urlHasFilter = urlParams.has('filter_branch');
        selects.forEach(function(sel) {{
          if (!urlHasFilter && saved) {{
            var optionExists = Array.from(sel.options).some(function(o) {{ return o.value === saved; }});
            if (optionExists && sel.value !== saved) {{
              sel.value = saved;
              var form = sel.closest('form');
              if (form) {{
                form.submit();
                return;
              }}
            }}
          }}
          sel.addEventListener('change', function() {{
            localStorage.setItem(PAGE_KEY, sel.value);
          }});
        }});
        if (urlHasFilter) {{
          var urlVal = urlParams.get('filter_branch') || '';
          localStorage.setItem(PAGE_KEY, urlVal);
        }}
      }})();
      </script>
    </body></html>
    """


# ── 로그인 ──────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page():
    branches = get_branches()
    branch_options = "".join(
        f'<option value="{b["login_id"]}">{b["branch_name"]}</option>' for b in branches
    )
    return HTMLResponse(content=f"""
    <html><head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>로그인</title>
      <style>
        * {{ box-sizing:border-box; }}
        body {{ font-family:-apple-system,sans-serif;background:#f5f7fa;
               display:flex;justify-content:center;align-items:center;
               height:100vh;margin:0; }}
        .card {{ background:white;padding:32px;border-radius:16px;
                box-shadow:0 2px 12px rgba(0,0,0,0.1);width:320px; }}
        input, select {{ width:100%;padding:12px;border:1px solid #ddd;
                border-radius:8px;font-size:14px;margin-top:4px;margin-bottom:14px; }}
        .btn {{ width:100%;background:#1E2761;color:white;padding:13px;
               border:none;border-radius:8px;cursor:pointer;font-size:15px; }}
      </style>
    </head>
    <body>
      <div class="card">
        <h2 style="color:#1E2761;text-align:center;margin-bottom:24px;">📦 재고 관리 시스템</h2>
        <form method="post" action="/login">
          <label style="font-size:13px;color:#555;">지점 선택</label>
          <select name="login_id" required onchange="document.getElementById('pwInput').focus()">
            <option value="" disabled selected>지점을 선택하세요</option>
            {branch_options}
            <option value="__manual__">직접 입력 (마스터 등)</option>
          </select>
          <div id="manualIdWrap" style="display:none;">
            <label style="font-size:13px;color:#555;">아이디 직접 입력</label>
            <input name="login_id_manual" id="manualIdInput" placeholder="아이디 입력">
          </div>
          <label style="font-size:13px;color:#555;">비밀번호</label>
          <input name="password" type="password" id="pwInput" required placeholder="비밀번호 입력">
          <button class="btn" type="submit">로그인</button>
        </form>
      </div>
      <script>
        const selectEl = document.querySelector('select[name="login_id"]');
        const manualWrap = document.getElementById('manualIdWrap');
        const manualInput = document.getElementById('manualIdInput');
        selectEl.addEventListener('change', function() {{
          if (selectEl.value === '__manual__') {{
            manualWrap.style.display = 'block';
            selectEl.removeAttribute('name');
            manualInput.setAttribute('name', 'login_id');
            manualInput.focus();
          }} else {{
            manualWrap.style.display = 'none';
            selectEl.setAttribute('name', 'login_id');
            manualInput.removeAttribute('name');
          }}
        }});
      </script>
    </body></html>
    """)


@app.post("/login")
async def login_submit(request: Request, login_id: str = Form(...), password: str = Form(...)):
    account = authenticate(login_id, password)
    if not account:
        return HTMLResponse(content="""
        <html><head><meta charset="utf-8"></head>
        <body style="font-family:sans-serif;text-align:center;padding-top:80px;">
          <h3>❌ 아이디 또는 비밀번호가 틀렸습니다.</h3>
          <a href="/login">다시 시도</a>
        </body></html>""", status_code=401)
    device_info = request.headers.get("user-agent", "")[:255]
    client_ip = request.client.host if request.client else ""
    token = create_session(account["login_id"], account["role"], account["branch_code"], device_info, client_ip)
    resp = RedirectResponse(url="/", status_code=303)
    resp.set_cookie(key="session_token", value=token, max_age=7 * 24 * 3600, httponly=True)
    return resp


@app.get("/logout")
async def logout(session_token: str = Cookie(default=None)):
    if session_token:
        delete_session(session_token)
    resp = RedirectResponse(url="/login", status_code=303)
    resp.delete_cookie("session_token")
    return resp


@app.get("/auto-login/{token}")
async def auto_login(token: str):
    info = get_auto_login_info(token)
    if not info:
        return HTMLResponse(content="<h3>❌ 유효하지 않은 자동로그인 링크입니다.</h3>", status_code=404)
    session_token = create_session(info["login_id"], "branch", info["branch_code"])
    resp = RedirectResponse(url="/", status_code=303)
    resp.set_cookie(key="session_token", value=session_token, max_age=7 * 24 * 3600, httponly=True)
    return resp

def send_teams_notification_to_url(webhook_url: str, title: str, message: str, link_url: str = "", link_text: str = "", sent_by: str = "", device_info: str = "", target_label: str = "custom"):
    """webhook_url을 직접 지정해서 발송 (자유 채널용). teams_send_log에도 기록."""
    import httpx
    body_blocks = [
        {"type": "TextBlock", "text": title, "weight": "Bolder", "size": "Medium", "wrap": True},
        {"type": "TextBlock", "text": message, "wrap": True}
    ]
    if link_url:
        body_blocks.append({
            "type": "ActionSet",
            "actions": [{"type": "Action.OpenUrl", "title": link_text or "바로가기", "url": link_url}]
        })
    payload = {
        "type": "message",
        "attachments": [{
            "contentType": "application/vnd.microsoft.card.adaptive",
            "contentUrl": None,
            "content": {
                "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                "type": "AdaptiveCard", "version": "1.4", "body": body_blocks
            }
        }]
    }
    success, detail = False, ""
    try:
        with httpx.Client(timeout=10) as client:
            resp = client.post(webhook_url, json=payload)
            if resp.status_code >= 300:
                detail = f"status={resp.status_code} body={resp.text[:300]}"
            else:
                detail, success = "성공", True
    except Exception as e:
        detail = f"exception={str(e)[:300]}"

    conn = get_conn()
    conn.execute(
        "INSERT INTO teams_send_log (branch_code, title, message, success, detail, sent_by, device_info) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (target_label, title, message, success, detail, sent_by, device_info)
    )
    conn.commit()
    conn.close()
    return success, detail

def send_teams_notification(branch_code: str, title: str, message: str, link_url: str = "", link_text: str = "", sent_by: str = "", device_info: str = ""):
    """지정된 branch_code(또는 'master')의 Teams 웹훅으로 알림 발송 (Adaptive Card 형식, 링크 지원).
    웹훅 미등록 시 조용히 무시. 모든 시도를 teams_send_log에 기록.
    반환값: (성공여부: bool, 상세정보: str)"""
    import httpx
    conn = get_conn()
    row = conn.execute(
        "SELECT webhook_url FROM teams_webhook WHERE branch_code=?", (branch_code,)
    ).fetchone()

    if not row or not row["webhook_url"]:
        conn.execute(
            "INSERT INTO teams_send_log (branch_code, title, message, success, detail, sent_by, device_info) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (branch_code, title, message, False, "웹훅 미등록", sent_by, device_info)
        )
        conn.commit()
        conn.close()
        return False, "웹훅 미등록"

    body_blocks = [
        {"type": "TextBlock", "text": title, "weight": "Bolder", "size": "Medium", "wrap": True},
        {"type": "TextBlock", "text": message, "wrap": True}
    ]
    if link_url:
        body_blocks.append({
            "type": "ActionSet",
            "actions": [
                {"type": "Action.OpenUrl", "title": link_text or "바로가기", "url": link_url}
            ]
        })

    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "contentUrl": None,
                "content": {
                    "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
                    "type": "AdaptiveCard",
                    "version": "1.4",
                    "body": body_blocks
                }
            }
        ]
    }

    success = False
    detail = ""
    try:
        with httpx.Client(timeout=10) as client:
            resp = client.post(row["webhook_url"], json=payload)
            if resp.status_code >= 300:
                detail = f"status={resp.status_code} body={resp.text[:300]}"
                success = False
            else:
                detail = "성공"
                success = True
    except Exception as e:
        detail = f"exception={str(e)[:300]}"
        success = False

    conn.execute(
        "INSERT INTO teams_send_log (branch_code, title, message, success, detail, sent_by, device_info) VALUES (?, ?, ?, ?, ?, ?, ?)",
        (branch_code, title, message, success, detail, sent_by, device_info)
    )
    conn.commit()
    conn.close()
    return success, detail

    payload = {
        "@type": "MessageCard",
        "@context": "http://schema.org/extensions",
        "themeColor": color,
        "title": title,
        "text": message
    }

    try:
        with httpx.Client(timeout=10) as client:
            resp = client.post(row["webhook_url"], json=payload)
            return resp.status_code < 300
    except Exception:
        return False

# ── RAW 목 데이터 (추후 MSSQL 교체) ─────────────────────

def fetch_raw_inventory() -> List[Dict[str, Any]]:
    """RAW 재고 — master 소스 우선, 없으면 branch 소스 사용"""
    conn = get_conn()
    rows = conn.execute("SELECT * FROM raw_inventory ORDER BY source DESC").fetchall()
    conn.close()
    merged = {}
    for r in rows:
        key = f"{r['branch_code']}|{r['item_code']}"
        if key not in merged:
            merged[key] = dict(r)
    return list(merged.values())


# ── 대시보드 ────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def dashboard(
    session_token: str = Cookie(default=None),
    filter_branch: str = ""
):
    """대시보드 — 불일치 품목만 표시"""
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    if user["role"] == "master":
        if filter_branch:
            rows = conn.execute(
                "SELECT * FROM inventory WHERE branch_code=? ORDER BY item_code",
                (filter_branch,)
            ).fetchall()
            scan_last = conn.execute(
                "SELECT MAX(scanned_at) as t FROM scan_log WHERE branch_code=?",
                (filter_branch,)
            ).fetchone()
            raw_last = conn.execute(
                "SELECT MAX(uploaded_at) as t FROM raw_inventory WHERE branch_code=?",
                (filter_branch,)
            ).fetchone()
        else:
            rows = conn.execute(
                "SELECT * FROM inventory ORDER BY branch_code, item_code"
            ).fetchall()
            scan_last = conn.execute("SELECT MAX(scanned_at) as t FROM scan_log").fetchone()
            raw_last = conn.execute("SELECT MAX(uploaded_at) as t FROM raw_inventory").fetchone()
    else:
        rows = conn.execute(
            "SELECT * FROM inventory WHERE branch_code=? ORDER BY item_code",
            (user["branch_code"],)
        ).fetchall()
        scan_last = conn.execute(
            "SELECT MAX(scanned_at) as t FROM scan_log WHERE branch_code=?",
            (user["branch_code"],)
        ).fetchone()
        raw_last = conn.execute(
            "SELECT MAX(uploaded_at) as t FROM raw_inventory WHERE branch_code=?",
            (user["branch_code"],)
        ).fetchone()
    conn.close()

    def format_kst(raw_value):
        if not raw_value:
            return "기록 없음"
        try:
            from datetime import timezone, timedelta
            dt = datetime.fromisoformat(str(raw_value).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            kst = dt.astimezone(timezone(timedelta(hours=9)))
            return kst.strftime('%m/%d %H:%M')
        except Exception:
            return str(raw_value)[:16]

    scan_last_display = format_kst(scan_last["t"] if scan_last else None)
    raw_last_display = format_kst(raw_last["t"] if raw_last else None)

    raw_mock = fetch_raw_inventory()
    raw_map = {f"{r['branch_code']}|{r['item_code']}": r["quantity"] for r in raw_mock}

    disc_rows = ""
    disc_count = 0
    for r in rows:
        key = f"{r['branch_code']}|{r['item_code']}"
        raw_qty = raw_map.get(key, 0)
        diff = r["quantity"] - raw_qty
        if diff != 0:
            disc_count += 1
            color = "#EF4444" if diff < 0 else "#F59E0B"
            disc_rows += f"""
            <tr>
              <td>{r['branch_code']}</td>
              <td>{r['item_name']}</td>
              <td>{r['item_code']}</td>
              <td>{r['quantity']}</td>
              <td>{raw_qty}</td>
              <td style="color:{color};font-weight:bold;">{diff:+d}</td>
              <td>
                <a href="/adjust?preset_branch={r['branch_code']}&preset_code={r['item_code']}"
                   style="background:#1E2761;color:white;padding:4px 10px;
                          border-radius:6px;font-size:12px;text-decoration:none;">
                  수정
                </a>
              </td>
            </tr>"""

    if not disc_rows:
        disc_rows = '<tr><td colspan="7" style="text-align:center;padding:24px;color:#22C55E;">✅ 모든 재고가 일치합니다</td></tr>'

    branch_filter_html = ""
    if user["role"] == "master":
        branch_options = '<option value="">전체 지점</option>'
        branches = get_branches()
        for b in branches:
            sel = "selected" if filter_branch == b["branch_code"] else ""
            branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'
        branch_filter_html = f"""
        <form method="get" action="/" style="margin-bottom:16px;">
          <div style="display:flex;gap:8px;align-items:flex-end;">
            <div style="flex:1;max-width:220px;">
              <label style="font-size:12px;color:#888;">지점 필터</label>
              <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
            </div>
            <button class="btn" type="submit">선택</button>
            <a href="/" style="padding:10px 14px;background:#eee;
               border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
          </div>
        </form>"""

    notif_type_labels_js = json.dumps(
        {k: {"label": v["label"], "desc": v["desc"]} for k, v in NOTIFICATION_TYPES.items()},
        ensure_ascii=False
    )

    content = f"""
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
      <h2 style="margin:0;">⚠️ 대시보드</h2>
      <button class="btn" type="button" style="font-size:12px;padding:6px 12px;" onclick="openNotifSettings()">🔔 알림 설정</button>
    </div>
    <p style="color:#888;font-size:13px;margin-bottom:16px;">불일치 품목만 표시됩니다</p>
    {branch_filter_html}

    <div id="notifSettingsModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:420px;width:90%;max-height:85vh;overflow-y:auto;">
        <h3 style="margin-bottom:12px;">🔔 알림 설정</h3>
        <div id="notifSettingsBody">
          <p style="font-size:13px;color:#888;">불러오는 중...</p>
        </div>
        <div style="display:flex;gap:8px;margin-top:16px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeNotifSettings()">닫기</button>
        </div>
      </div>
    </div>

    <script>
      const NOTIF_TYPE_LABELS = {notif_type_labels_js};

      function urlBase64ToUint8Array(base64String) {{
        var padding = '='.repeat((4 - base64String.length % 4) % 4);
        var base64 = (base64String + padding).replace(/-/g, '+').replace(/_/g, '/');
        var rawData = window.atob(base64);
        var outputArray = new Uint8Array(rawData.length);
        for (var i = 0; i < rawData.length; ++i) {{
          outputArray[i] = rawData.charCodeAt(i);
        }}
        return outputArray;
      }}

      async function getCurrentEndpoint() {{
        if (!('serviceWorker' in navigator)) return null;
        const reg = await navigator.serviceWorker.ready;
        const sub = await reg.pushManager.getSubscription();
        return sub ? sub.endpoint : null;
      }}

      async function enablePushAndSubscribe() {{
        if (!('serviceWorker' in navigator) || !('PushManager' in window)) {{
          alert('이 브라우저는 알림 기능을 지원하지 않습니다.');
          return;
        }}
        const permission = await Notification.requestPermission();
        if (permission !== 'granted') {{
          alert('알림 권한이 거부되었습니다. 브라우저 설정에서 허용해주세요.');
          return;
        }}
        const reg = await navigator.serviceWorker.ready;
        let sub = await reg.pushManager.getSubscription();
        if (!sub) {{
          const keyRes = await fetch('/api/push/vapid-public-key');
          const keyData = await keyRes.json();
          sub = await reg.pushManager.subscribe({{
            userVisibleOnly: true,
            applicationServerKey: urlBase64ToUint8Array(keyData.key)
          }});
        }}
        await fetch('/api/push/subscribe', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify(sub)
        }});
        await loadNotifSettings();
      }}

      async function openNotifSettings() {{
        document.getElementById('notifSettingsModal').style.display = 'flex';
        await loadNotifSettings();
      }}

      function closeNotifSettings() {{
        document.getElementById('notifSettingsModal').style.display = 'none';
      }}

      async function loadNotifSettings() {{
        const body = document.getElementById('notifSettingsBody');
        const endpoint = await getCurrentEndpoint();

        if (!endpoint) {{
          body.innerHTML = '<p style="font-size:13px;color:#555;margin-bottom:12px;">이 기기에서 아직 알림을 켜지 않았습니다.</p>' +
            '<button class="btn" style="width:100%;" onclick="enablePushAndSubscribe()">이 기기에서 알림 켜기</button>';
          return;
        }}

        const res = await fetch('/api/push/settings?endpoint=' + encodeURIComponent(endpoint));
        const data = await res.json();

        if (!data.subscribed) {{
          body.innerHTML = '<p style="font-size:13px;color:#555;margin-bottom:12px;">이 기기에서 아직 알림을 켜지 않았습니다.</p>' +
            '<button class="btn" style="width:100%;" onclick="enablePushAndSubscribe()">이 기기에서 알림 켜기</button>';
          return;
        }}

        const settings = data.settings || {{}};
        let html = '';
        html += renderToggleRow('all', '전체 알림', '이 기기의 모든 알림을 켜고 끕니다', settings['all'] !== false);
        html += '<hr style="margin:12px 0;border:none;border-top:1px solid #eee;">';
        for (const key in NOTIF_TYPE_LABELS) {{
          const info = NOTIF_TYPE_LABELS[key];
          html += renderToggleRow(key, info.label, info.desc, settings[key] !== false);
        }}
        body.innerHTML = html;

        body.querySelectorAll('.notif-toggle-label').forEach(function(labelEl) {{
          labelEl.addEventListener('click', function() {{
            toggleSwitch(labelEl, labelEl.getAttribute('data-notif-type'));
          }});
        }});
      }}

      function renderToggleRow(type, label, desc, checked) {{
        return '<div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;">' +
          '<div><div style="font-size:14px;font-weight:' + (type === 'all' ? 'bold' : 'normal') + ';">' + label + '</div>' +
          '<div style="font-size:11px;color:#888;">' + desc + '</div></div>' +
          '<label class="notif-toggle-label" data-notif-type="' + type + '" style="position:relative;display:inline-block;width:44px;height:24px;flex-shrink:0;margin-left:8px;cursor:pointer;">' +
          '<input type="checkbox" class="notif-toggle" data-type="' + type + '" ' + (checked ? 'checked' : '') + ' style="opacity:0;width:0;height:0;position:absolute;">' +
          '<span class="notif-toggle-track" style="position:absolute;pointer-events:none;top:0;left:0;right:0;bottom:0;background:' + (checked ? '#1E2761' : '#ccc') + ';border-radius:24px;transition:0.2s;">' +
          '<span class="notif-toggle-knob" style="position:absolute;height:18px;width:18px;left:' + (checked ? '23px' : '3px') + ';top:3px;background:white;border-radius:50%;transition:0.2s;"></span>' +
          '</span></label></div>';
      }}

      async function toggleSwitch(labelEl, type) {{
        const input = labelEl.querySelector('input.notif-toggle');
        const newChecked = !input.checked;
        input.checked = newChecked;

        const track = labelEl.querySelector('.notif-toggle-track');
        const knob = labelEl.querySelector('.notif-toggle-knob');
        track.style.background = newChecked ? '#1E2761' : '#ccc';
        knob.style.left = newChecked ? '23px' : '3px';

        const endpoint = await getCurrentEndpoint();
        if (!endpoint) return;

        await fetch('/api/push/settings', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{
            endpoint: endpoint,
            notification_type: type,
            enabled: newChecked
          }})
        }});
      }}

    </script>
    <div style="display:flex;gap:12px;margin-bottom:16px;">
      <div class="card" style="flex:1;text-align:center;padding:16px;">
        <div style="color:#888;font-size:12px;">불일치 품목</div>
        <div style="font-size:28px;font-weight:bold;color:#EF4444;">{disc_count}</div>
      </div>
      <div class="card" style="flex:1;text-align:center;padding:16px;">
        <div style="color:#888;font-size:12px;">QR 마지막 스캔</div>
        <div style="font-size:13px;font-weight:bold;color:#1E2761;">{scan_last_display}</div>
      </div>
      <div class="card" style="flex:1;text-align:center;padding:16px;">
        <div style="color:#888;font-size:12px;">RAW 마지막 확인</div>
        <div style="font-size:13px;font-weight:bold;color:#1E2761;">{raw_last_display}</div>
      </div>
    </div>
    <div class="card">
      <table>
        <thead><tr>
          <th>지점</th><th>상품명</th><th>품번</th>
          <th>QR재고</th><th>RAW재고</th><th>차이</th><th>수정</th>
        </tr></thead>
        <tbody>{disc_rows}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "dashboard"))


# ── 재고현황 ────────────────────────────────────────────

@app.get("/inventory", response_class=HTMLResponse)
async def inventory_page(
    session_token: str = Cookie(default=None),
    search: str = "",
    filter_branch: str = "",
    filter_item: str = ""
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    query = "SELECT * FROM inventory WHERE 1=1"
    params: list = []

    if user["role"] == "branch":
        query += " AND branch_code=?"
        params.append(user["branch_code"])
    elif filter_branch:
        query += " AND branch_code=?"
        params.append(filter_branch)

    if filter_item:
        query += " AND item_name=?"
        params.append(filter_item)

    if search:
        query += " AND (item_name LIKE ? OR item_code LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]

    query += " ORDER BY branch_code, item_code"
    rows = conn.execute(query, params).fetchall()

    all_branches = conn.execute(
        "SELECT DISTINCT branch_code FROM inventory ORDER BY branch_code"
    ).fetchall() if user["role"] == "master" else []
    all_items = conn.execute(
        "SELECT DISTINCT item_name FROM inventory ORDER BY item_name"
    ).fetchall()
    conn.close()

    raw_mock = fetch_raw_inventory()
    raw_map = {f"{r['branch_code']}|{r['item_code']}": r["quantity"] for r in raw_mock}

    default_branch_sel = "selected" if not filter_branch else ""
    branch_options = f'<option value="" {default_branch_sel}>전체 지점</option>'
    for b in all_branches:
      sel = "selected" if filter_branch == b["branch_code"] else ""
      branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_code"]}</option>'

    default_sel = "selected" if not filter_item else ""
    item_options = f'<option value="" {default_sel}>전체 품목</option>'
    for it in all_items:
      sel = "selected" if filter_item == it["item_name"] else ""
      item_options += f'<option value="{it["item_name"]}" {sel}>{it["item_name"]}</option>'

    rows_html = ""
    if not rows:
        rows_html = '<tr><td colspan="8" style="text-align:center;padding:24px;color:#888;">데이터 없음</td></tr>'
    else:
        for r in rows:
            key = f"{r['branch_code']}|{r['item_code']}"
            raw_qty = raw_map.get(key, "-")
            diff = (r["quantity"] - raw_qty) if isinstance(raw_qty, int) else "-"
            if diff == "-":
                badge = '<span class="badge-red">RAW없음</span>'
            elif diff == 0:
                badge = '<span class="badge-green">일치</span>'
            else:
                badge = f'<span class="badge-red">불일치 ({diff:+d})</span>'
            rows_html += f"""
            <tr>
              <td style="text-align:center;">
                <input type="checkbox" name="selected_ids" value="{r['id']}"
                       class="inv-check" style="width:16px;height:16px;">
              </td>
              <td>{r['branch_code']}</td>
              <td>{r['item_name']}</td>
              <td>{r['item_code']}</td>
              <td>{r['quantity']}</td>
              <td>{raw_qty}</td>
              <td>{diff if diff != '-' else '-'}</td>
              <td>{badge}</td>
            </tr>"""

    branch_filter_html = ""
    if user["role"] == "master":
        branch_filter_html = f"""
        <div style="flex:1;min-width:120px;">
          <label style="font-size:12px;color:#888;">지점 필터</label>
          <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
        </div>"""

    content = f"""
    <h2 style="margin-bottom:8px;">📦 재고현황</h2>
    <p style="color:#888;font-size:13px;margin-bottom:16px;">전체 재고를 표시합니다</p>
    <div class="card">
      <form method="get" action="/inventory"
            style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        {branch_filter_html}
        <div style="flex:1;min-width:120px;">
          <label style="font-size:12px;color:#888;">품목 필터</label>
          <select name="filter_item" style="margin-top:4px;">{item_options}</select>
        </div>
        <div style="flex:2;min-width:160px;">
          <label style="font-size:12px;color:#888;">검색 (상품명/품번)</label>
          <input name="search" value="{search}" placeholder="검색어 입력"
                 style="margin-top:4px;">
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/inventory" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
    </div>
    <div class="card">
      <form method="post" action="/inventory/delete-selected" id="invForm">
        <div style="display:flex;justify-content:space-between;
                    align-items:center;margin-bottom:12px;">
          <span style="font-size:13px;color:#888;">{len(rows)}개 항목</span>
          <div style="display:flex;gap:8px;">
            <button type="button" class="btn" id="invSelectAllBtn"
                    style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
            <button type="submit" class="btn btn-red"
                    style="font-size:12px;padding:6px 12px;"
                    onclick="return confirm('선택한 재고를 삭제할까요?')">선택삭제</button>
          </div>
        </div>
        <table>
          <thead><tr>
            <th style="width:40px;text-align:center;">
              <input type="checkbox" id="invAllCheck" style="width:16px;height:16px;">
            </th>
            <th>지점</th><th>상품명</th><th>품번</th>
            <th>QR재고</th><th>RAW재고</th><th>차이</th><th>상태</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </form>
    </div>
    <script>
      (function() {{
        var allCheck = document.getElementById('invAllCheck');
        var selectBtn = document.getElementById('invSelectAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.inv-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{
          allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }});
        }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
      }})();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "inventory"))


@app.post("/inventory/delete-selected")
async def inventory_delete_selected(
    request: Request,
    session_token: str = Cookie(default=None)
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("selected_ids")
    if ids:
        conn = get_conn()
        if user["role"] == "master":
            conn.execute(
                f"DELETE FROM inventory WHERE id IN ({','.join('?' for _ in ids)})",
                [int(i) for i in ids]
            )
        else:
            conn.execute(
                f"DELETE FROM inventory WHERE id IN ({','.join('?' for _ in ids)}) AND branch_code=?",
                [int(i) for i in ids] + [user["branch_code"]]
            )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/inventory", status_code=303)


# ── 데이터 관리 (마스터 전용 - 구 경로, 호환용 리다이렉트) ─

@app.get("/data")
async def data_page_redirect():
    return RedirectResponse(url="/master/data", status_code=303)


# ── QR 생성 ─────────────────────────────────────────────

def generate_qr_bytes(server_url, branch_code, item_code, scan_type, item_name="") -> bytes:
    """QR 코드를 메모리에서 생성 + 하단에 상품명(최대 3줄 자동 줄바꿈)/입출고 라벨 삽입"""
    import qrcode
    import io
    from PIL import Image, ImageDraw, ImageFont

    url = (f"{server_url}/scan"
           f"?branch_code={branch_code}"
           f"&item_code={item_code}"
           f"&scan_type={scan_type}")
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    def wrap_text(draw_obj, text, font_obj, max_width, max_lines=3):
        if not text:
            return [""]
        lines = []
        current = ""
        for ch in text:
            test = current + ch
            bbox = draw_obj.textbbox((0, 0), test, font=font_obj)
            w = bbox[2] - bbox[0]
            if w > max_width and current:
                lines.append(current)
                current = ch
                if len(lines) >= max_lines - 1:
                    break
            else:
                current = test
        if current:
            lines.append(current)
        if len(lines) > max_lines:
            lines = lines[:max_lines]
        if len(lines) == max_lines:
            last = lines[-1]
            while True:
                bbox = draw_obj.textbbox((0, 0), last + "…", font=font_obj)
                if bbox[2] - bbox[0] <= max_width or len(last) <= 1:
                    break
                last = last[:-1]
            consumed = sum(len(l) for l in lines[:-1]) + len(last)
            if consumed < len(text):
                lines[-1] = last + "…"
        return lines

    draw_probe = ImageDraw.Draw(qr_img)
    try:
        font_path = os.path.join(os.path.dirname(__file__), "fonts", "NanumGothic-Bold.ttf")
        font_size = int(qr_img.height * 0.065)
        font = ImageFont.truetype(font_path, font_size)
        type_font_size = int(qr_img.height * 0.075)
        type_font = ImageFont.truetype(font_path, type_font_size)
    except Exception:
        font = ImageFont.load_default()
        type_font = font
        font_size = 14
        type_font_size = 16

    max_text_width = int(qr_img.width * 0.92)
    name_lines = wrap_text(draw_probe, item_name if item_name else item_code, font, max_text_width, max_lines=3)

    type_label = "입고 IN" if scan_type == "IN" else "출고 OUT"

    line_height = int(font_size * 1.25)
    name_block_height = line_height * len(name_lines)
    type_block_height = int(type_font_size * 1.3)
    label_height = int(name_block_height + type_block_height + qr_img.height * 0.08)

    canvas = Image.new("RGB", (qr_img.width, qr_img.height + label_height), "white")
    canvas.paste(qr_img, (0, 0))

    draw = ImageDraw.Draw(canvas)

    y = qr_img.height + int(qr_img.height * 0.04)
    for line in name_lines:
        bbox = draw.textbbox((0, 0), line, font=font)
        w = bbox[2] - bbox[0]
        x = (canvas.width - w) / 2
        draw.text((x, y), line, fill="black", font=font)
        y += line_height

    y += int(qr_img.height * 0.02)
    bbox = draw.textbbox((0, 0), type_label, font=type_font)
    w = bbox[2] - bbox[0]
    x = (canvas.width - w) / 2
    draw.text((x, y), type_label, fill="black", font=type_font)

    buf = io.BytesIO()
    canvas.save(buf, format="PNG")
    buf.seek(0)
    return buf.getvalue()


@app.get("/qr/download/{filename}")
async def qr_download(filename: str):
    """QR 이미지를 즉석에서 생성해 반환 (파일 저장 없이 동작, Vercel 호환)"""
    from fastapi.responses import Response
    name = filename.replace(".png", "")
    parts = name.rsplit("_", 1)
    if len(parts) != 2:
        return Response(content=b"Invalid filename format", status_code=404)
    prefix, scan_type = parts
    branch_part = prefix.split("_", 1)
    if len(branch_part) != 2:
        return Response(content=b"Invalid filename format", status_code=404)
    branch_code, item_code = branch_part

    hostname_env = os.getenv("PUBLIC_SERVER_URL")
    if hostname_env:
        server_url = hostname_env
    else:
        hostname = socket.gethostbyname(socket.gethostname())
        server_url = f"http://{hostname}:{SERVER_PORT}"

    item_name = ""
    try:
        conn = get_conn()
        item = conn.execute(
            "SELECT item_name FROM items WHERE branch_code=? AND item_code=?",
            (branch_code, item_code)
        ).fetchone()
        conn.close()
        item_name = item["item_name"] if item else ""
    except Exception:
        item_name = ""

    try:
        img_bytes = generate_qr_bytes(server_url, branch_code, item_code, scan_type, item_name)
    except Exception as e:
        return Response(content=f"QR generation failed: {str(e)}".encode(), status_code=500)

    return Response(content=img_bytes, media_type="image/png")

def generate_qr_image(server_url, branch_code, item_name, item_code, scan_type, output_dir=QR_DIR):
    """QR 코드 생성 — 로컬 환경에서는 파일로도 저장 (호환용), Vercel에서는 파일 저장 실패해도 무시"""
    filename = f"{branch_code}_{item_code}_{scan_type}.png"
    conn = get_conn()
    item = conn.execute(
        "SELECT item_name FROM items WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    conn.close()
    item_name = item["item_name"] if item else ""
    img_bytes = generate_qr_bytes(server_url, branch_code, item_code, scan_type, item_name)
    try:
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        file_path = Path(output_dir) / filename
        with open(file_path, "wb") as f:
            f.write(img_bytes)
        return str(file_path), filename
    except Exception:
        # Vercel 등 읽기전용 파일시스템에서는 파일 저장 생략, 바이트만 사용
        return None, filename


@app.get("/qr", response_class=HTMLResponse)
async def qr_page(
    session_token: str = Cookie(default=None),
    filter_branch: str = ""
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    if user["role"] == "branch":
        filter_branch = user["branch_code"]

    conn = get_conn()
    if filter_branch:
        items = conn.execute(
            "SELECT * FROM items WHERE branch_code=? ORDER BY item_name",
            (filter_branch,)
        ).fetchall()
    else:
        items = conn.execute(
            "SELECT * FROM items ORDER BY branch_code, item_name"
        ).fetchall()
    conn.close()

    options_html = '<option value="">-- 품목 선택 --</option>'
    for it in items:
        options_html += f'<option value="{it["branch_code"]}|{it["item_name"]}|{it["item_code"]}">{it["branch_name"]} / {it["item_name"]} ({it["item_code"]})</option>'

    branch_filter_html = ""
    bulk_html = ""

    if user["role"] == "master":
        branch_options = '<option value="">전체 지점</option>'
        branches = get_branches()
        for b in branches:
            sel = "selected" if filter_branch == b["branch_code"] else ""
            branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'

        branch_filter_html = f"""
        <form method="get" action="/qr" style="margin-bottom:12px;">
          <div style="display:flex;gap:8px;align-items:flex-end;">
            <div style="flex:1;">
              <label style="font-size:12px;color:#888;">지점 선택</label>
              <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
            </div>
            <button class="btn" type="submit">선택</button>
            <a href="/qr" style="padding:10px 14px;background:#eee;
               border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
          </div>
        </form>"""

        if filter_branch:
            bulk_html = f"""
            <div class="card" style="border:1px solid #1E2761;">
              <h3 style="margin-bottom:8px;">📦 일괄 생성</h3>
              <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <form method="post" action="/master/qr/generate-bulk" style="flex:1;" onsubmit="showZipLoading();">
                  <input type="hidden" name="branch_code" value="{filter_branch}">
                  <button class="btn" type="submit" style="width:100%;">
                    📦 {filter_branch} ZIP 다운로드
                  </button>
                </form>
                <form method="post" action="/master/qr/generate-bulk" style="flex:1;" onsubmit="showZipLoading();">
                  <input type="hidden" name="branch_code" value="ALL">
                  <button class="btn" type="submit"
                          style="width:100%;background:#64748B;">
                    🌐 전체 지점 ZIP 다운로드
                  </button>
                </form>
              </div>
            </div>"""
        else:
            bulk_html = """
            <div class="card" style="border:1px solid #1E2761;">
              <h3 style="margin-bottom:8px;">📦 일괄 생성</h3>
              <div style="display:flex;gap:8px;flex-wrap:wrap;">
                <form method="post" action="/master/qr/generate-bulk" style="flex:1;" onsubmit="showZipLoading();">
                  <input type="hidden" name="branch_code" value="ALL">
                  <button class="btn" type="submit" style="width:100%;">
                    🌐 전체 지점 ZIP 다운로드
                  </button>
                </form>
              </div>
              <p style="color:#aaa;font-size:12px;margin-top:8px;">
                특정 지점만 생성하려면 위에서 지점을 선택하세요.
              </p>
            </div>"""
    else:
        # 지점 계정 — 본인 지점 일괄생성 버튼 항상 표시
        bulk_html = f"""
        <div class="card" style="border:1px solid #1E2761;">
          <h3 style="margin-bottom:8px;">📦 일괄 생성</h3>
          <p style="color:#666;font-size:12px;margin-bottom:12px;">
            우리 지점({user['branch_code']})의 전체 품목 QR을 ZIP으로 한 번에 받을 수 있어요.
          </p>
          <form method="post" action="/master/qr/generate-bulk" onsubmit="showZipLoading();">
            <input type="hidden" name="branch_code" value="{user['branch_code']}">
            <button class="btn" type="submit" style="width:100%;">
              📦 우리 지점 전체 QR ZIP 다운로드
            </button>
          </form>
        </div>"""

    content = f"""
    <h2 style="margin-bottom:16px;">📷 QR / 바코드 생성</h2>
    {branch_filter_html}
    <div class="card">
      <h3 style="margin-bottom:12px;">개별 생성</h3>
      <form method="post" action="/qr/generate" id="qrGenForm"
            onsubmit="startQrGenTimer();">
        <label style="font-size:13px;color:#555;">품목 선택 ({len(items)}개)</label>
        <select name="item_key" required style="margin-bottom:14px;">
          {options_html}
        </select>
        <button class="btn" type="submit" id="qrGenBtn" style="width:100%;">입고/출고 생성</button>
      </form>
    </div>
    {bulk_html}

    <div id="zipLoadingOverlay" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;
         background:rgba(0,0,0,0.5);z-index:999;justify-content:center;align-items:center;">
      <div style="background:white;padding:24px 32px;border-radius:12px;text-align:center;">
        <div style="font-size:32px;margin-bottom:8px;">📦</div>
        <div style="font-weight:bold;color:#1E2761;">ZIP 생성 중입니다...</div>
        <div id="zipTimerText" style="color:#888;font-size:13px;margin-top:6px;">잠시만 기다려주세요</div>
      </div>
    </div>
    <script>
    function startQrGenTimer() {{
      const btn = document.getElementById('qrGenBtn');
      let seconds = 0;
      btn.disabled = true;
      const interval = setInterval(() => {{
        seconds++;
        btn.innerHTML = `⏳ 생성 중... (${{seconds}}초째)`;
      }}, 1000);
      btn.innerHTML = '⏳ 생성 중... (0초째)';
    }}
    function showZipLoading() {{
      document.getElementById('zipLoadingOverlay').style.display = 'flex';
      let seconds = 0;
      const label = document.getElementById('zipTimerText');
      const interval = setInterval(() => {{
        seconds++;
        label.textContent = `${{seconds}}초째 생성 중...`;
      }}, 1000);
      setTimeout(() => {{
        document.getElementById('zipLoadingOverlay').style.display = 'none';
        clearInterval(interval);
      }}, 60000);
    }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "qr"))


@app.post("/qr/generate", response_class=HTMLResponse)
async def qr_generate(
    session_token: str = Cookie(default=None),
    item_key: str = Form(...)
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    parts = item_key.split("|")
    if len(parts) != 3:
        return HTMLResponse(content=render_page('<div class="card"><p>❌ 품목 선택 오류</p></div>', user, "qr"))

    branch_code, item_name, item_code = parts
    hostname_env = os.getenv("PUBLIC_SERVER_URL")
    if hostname_env:
        server_url = hostname_env
    else:
        hostname = socket.gethostbyname(socket.gethostname())
        server_url = f"http://{hostname}:{SERVER_PORT}"

    _, in_file = generate_qr_image(server_url, branch_code, item_name, item_code, "IN")
    _, out_file = generate_qr_image(server_url, branch_code, item_name, item_code, "OUT")

    content = f"""
    <h2 style="margin-bottom:16px;">✅ QR 생성 완료</h2>
    <div class="card">
      <p><b>지점:</b> {branch_code}</p>
      <p><b>상품명:</b> {item_name}</p>
      <p style="margin-bottom:16px;"><b>품번:</b> {item_code}</p>
      <div style="display:flex;gap:16px;flex-wrap:wrap;">
        <div style="text-align:center;flex:1;">
          <p style="font-weight:bold;color:#22C55E;margin-bottom:8px;">📥 입고 QR</p>
          <img src="/qr/download/{in_file}" style="width:160px;height:160px;border:1px solid #eee;border-radius:8px;"><br>
          <a href="/qr/download/{in_file}" download
             style="display:inline-block;margin-top:8px;background:#1E2761;
                    color:white;padding:8px 16px;border-radius:8px;
                    text-decoration:none;font-size:13px;">다운로드</a>
        </div>
        <div style="text-align:center;flex:1;">
          <p style="font-weight:bold;color:#EF4444;margin-bottom:8px;">📤 출고 QR</p>
          <img src="/qr/download/{out_file}" style="width:160px;height:160px;border:1px solid #eee;border-radius:8px;"><br>
          <a href="/qr/download/{out_file}" download
             style="display:inline-block;margin-top:8px;background:#1E2761;
                    color:white;padding:8px 16px;border-radius:8px;
                    text-decoration:none;font-size:13px;">다운로드</a>
        </div>
      </div>
      <div style="margin-top:16px;">
        <a href="/qr" style="color:#1E2761;font-size:13px;">← QR 생성으로 돌아가기</a>
      </div>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "qr"))


# ── QR 스캔 / 재고 조정 로직 ─────────────────────────────

def adjust_quantity(branch_code: str, item_code: str, delta: int, absolute: bool = False) -> int:
    conn = get_conn()
    item = conn.execute(
        "SELECT * FROM items WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    item_name = item["item_name"] if item else item_code
    row = conn.execute(
        "SELECT quantity FROM inventory WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    now = datetime.now().isoformat()

    if absolute:
        new_qty = max(0, delta)
    else:
        new_qty = max(0, (row["quantity"] if row else 0) + delta)

    if row is None:
        conn.execute(
            "INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated) VALUES (?, ?, ?, ?, ?)",
            (branch_code, item_name, item_code, new_qty, now)
        )
    else:
        conn.execute(
            "UPDATE inventory SET quantity=?, item_name=?, last_updated=? WHERE branch_code=? AND item_code=?",
            (new_qty, item_name, now, branch_code, item_code)
        )
    conn.commit()
    conn.close()
    return new_qty


@app.get("/scan", response_class=HTMLResponse)
async def scan_get(request: Request, branch_code: str, item_code: str, scan_type: str):
    delta = 1 if scan_type == "IN" else -1
    new_qty = adjust_quantity(branch_code, item_code, delta)

    device_info = request.headers.get("user-agent", "")[:255]
    client_ip = request.client.host if request.client else ""

    conn = get_conn()
    item = conn.execute(
        "SELECT item_name, branch_name FROM items WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    item_name = item["item_name"] if item else item_code
    branch_name = item["branch_name"] if item else branch_code
    now = datetime.now().isoformat()
    conn.execute(
        "INSERT INTO scan_log (branch_code, branch_name, item_name, item_code, scan_type, result_quantity, scanned_at, device_info, client_ip) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (branch_code, branch_name, item_name, item_code, scan_type, new_qty, now, device_info, client_ip)
    )
    conn.commit()
    conn.close()

    action_label = "입고 ✅" if scan_type == "IN" else "출고 📤"
    bg_color = "#D1FAE5" if scan_type == "IN" else "#FEE2E2"
    text_color = "#065F46" if scan_type == "IN" else "#991B1B"

    return HTMLResponse(content=f"""
    <html><head>
      <meta charset="utf-8">
      <meta name="viewport" content="width=device-width, initial-scale=1">
      <title>스캔 완료</title>
    </head>
    <body style="font-family:-apple-system,sans-serif;background:#f5f7fa;
                 display:flex;justify-content:center;align-items:center;
                 min-height:100vh;margin:0;">
      <div style="background:white;max-width:340px;width:90%;padding:32px;
                  border-radius:20px;box-shadow:0 4px 20px rgba(0,0,0,0.1);
                  text-align:center;">
        <div style="background:{bg_color};border-radius:12px;padding:16px;
                    margin-bottom:20px;">
          <div style="font-size:36px;margin-bottom:4px;">{action_label}</div>
          <div style="font-size:20px;font-weight:bold;color:{text_color};">
            {'입고' if scan_type == 'IN' else '출고'} 처리 완료
          </div>
        </div>
        <div style="text-align:left;background:#f8fafc;border-radius:10px;
                    padding:16px;margin-bottom:16px;">
          <div style="margin-bottom:10px;">
            <div style="font-size:11px;color:#888;margin-bottom:2px;">지점</div>
            <div style="font-size:15px;font-weight:bold;">{branch_name}</div>
          </div>
          <div style="margin-bottom:10px;">
            <div style="font-size:11px;color:#888;margin-bottom:2px;">상품명</div>
            <div style="font-size:15px;font-weight:bold;">{item_name}</div>
          </div>
          <div style="margin-bottom:10px;">
            <div style="font-size:11px;color:#888;margin-bottom:2px;">품번</div>
            <div style="font-size:14px;color:#64748B;">{item_code}</div>
          </div>
          <div>
            <div style="font-size:11px;color:#888;margin-bottom:2px;">현재 재고</div>
            <div style="font-size:28px;font-weight:bold;color:#1E2761;">{new_qty}개</div>
          </div>
        </div>
        <div style="font-size:12px;color:#aaa;">
          {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        </div>
      </div>
    </body></html>
    """)


@app.post("/scan")
async def scan_post(data: Dict[str, str]):
    delta = 1 if data.get("scan_type") == "IN" else -1
    new_qty = adjust_quantity(data["branch_code"], data["item_code"], delta)
    return {"status": "scanned", "scan_type": data.get("scan_type"), "new_quantity": new_qty}


# ── 수기 조정 ──────────────────────────────────────────

@app.get("/adjust", response_class=HTMLResponse)
async def adjust_get(
    session_token: str = Cookie(default=None),
    preset_branch: str = "",
    preset_code: str = "",
    filter_branch: str = ""
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    effective_branch = preset_branch or filter_branch

    conn = get_conn()
    if user["role"] == "master":
        if effective_branch:
            items = conn.execute(
                "SELECT * FROM items WHERE branch_code=? ORDER BY item_name",
                (effective_branch,)
            ).fetchall()
            logs = conn.execute(
                "SELECT * FROM adjustment_log WHERE branch_code=? ORDER BY id DESC LIMIT 20",
                (effective_branch,)
            ).fetchall()
        else:
            items = conn.execute(
                "SELECT * FROM items ORDER BY branch_code, item_name"
            ).fetchall()
            logs = conn.execute(
                "SELECT * FROM adjustment_log ORDER BY id DESC LIMIT 20"
            ).fetchall()
    else:
        items = conn.execute(
            "SELECT * FROM items WHERE branch_code=? ORDER BY item_name",
            (user["branch_code"],)
        ).fetchall()
        logs = conn.execute(
            "SELECT * FROM adjustment_log WHERE branch_code=? ORDER BY id DESC LIMIT 20",
            (user["branch_code"],)
        ).fetchall()
    conn.close()

    branch_filter_html = ""
    if user["role"] == "master":
        branch_options = '<option value="">전체 지점</option>'
        branches = get_branches()
        for b in branches:
            sel = "selected" if filter_branch == b["branch_code"] else ""
            branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'
        branch_filter_html = f"""
        <div class="card">
          <form method="get" action="/adjust" style="display:flex;gap:8px;align-items:flex-end;">
            <input type="hidden" name="preset_branch" value="{preset_branch}">
            <input type="hidden" name="preset_code" value="{preset_code}">
            <div style="flex:1;max-width:220px;">
              <label style="font-size:12px;color:#888;">지점 필터</label>
              <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
            </div>
            <button class="btn" type="submit">선택</button>
            <a href="/adjust" style="padding:10px 14px;background:#eee;
               border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
          </form>
        </div>"""

    options_html = ""
    for it in items:
        sel = "selected" if preset_code == it["item_code"] and preset_branch == it["branch_code"] else ""
        options_html += f'<option value="{it["branch_code"]}|{it["item_code"]}" {sel}>{it["branch_name"]} / {it["item_name"]} ({it["item_code"]})</option>'

    log_rows = ""
    if not logs:
        log_rows = '<tr><td colspan="7" style="text-align:center;padding:16px;color:#888;">이력 없음</td></tr>'
    else:
        from datetime import timezone as _tz, timedelta as _td
        for lg in logs:
            delta_str = f"+{lg['delta']}" if lg['delta'] > 0 else str(lg['delta'])
            adj_display = "-"
            if lg['adjusted_at']:
                try:
                    _dt = datetime.fromisoformat(str(lg['adjusted_at']).replace("Z", "+00:00"))
                    if _dt.tzinfo is None:
                        _dt = _dt.replace(tzinfo=_tz.utc)
                    adj_display = _dt.astimezone(_tz(_td(hours=9))).strftime("%Y-%m-%d %H:%M")
                except Exception:
                    adj_display = lg['adjusted_at'][:16]
            log_rows += f"""
            <tr>
              <td style="text-align:center;">
                <input type="checkbox" name="log_ids" value="{lg['id']}" class="log-check" style="width:16px;height:16px;">
              </td>
              <td>{adj_display}</td>
              <td>{lg['branch_code']}</td>
              <td>{lg['item_name'] or '-'}</td>
              <td>{lg['item_code']}</td>
              <td>{delta_str}</td>
              <td>{lg['result_quantity']}</td>
            </tr>"""

    log_controls = """
        <div style="display:flex;gap:8px;">
          <button type="button" class="btn" id="logSelectAllBtn"
                  style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
          <button type="submit" class="btn btn-red"
                  style="font-size:12px;padding:6px 12px;"
                  onclick="return confirm('선택한 이력을 삭제할까요?')">선택삭제</button>
          <button type="button" class="btn btn-red" id="logDeleteAllBtn"
                  style="font-size:12px;padding:6px 12px;">전체삭제</button>
        </div>"""

    log_header_check = """<th style="width:40px;text-align:center;">
              <input type="checkbox" id="logAllCheck" style="width:16px;height:16px;">
            </th>"""

    log_section = f"""
    <div class="card">
      <form method="post" action="/adjust/delete-logs" id="logForm">
        <div style="display:flex;justify-content:space-between;
                    align-items:center;margin-bottom:12px;">
          <h3>최근 조정 이력</h3>
          {log_controls}
        </div>
        <table>
          <thead><tr>
            {log_header_check}
            <th>시각</th><th>지점</th><th>상품명</th><th>품번</th><th>조정</th><th>결과</th>
          </tr></thead>
          <tbody>{log_rows}</tbody>
        </table>
      </form>
    </div>
    <form method="post" action="/adjust/delete-all-logs" id="logDeleteAllForm"></form>
    <script>
      (function() {{
        var allCheck = document.getElementById('logAllCheck');
        var selectBtn = document.getElementById('logSelectAllBtn');
        var deleteAllBtn = document.getElementById('logDeleteAllBtn');

        function applyAll(checked) {{
          document.querySelectorAll('.log-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}

        if (allCheck) {{
          allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }});
        }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        if (deleteAllBtn) {{
          deleteAllBtn.addEventListener('click', function() {{
            if (confirm('전체 이력을 삭제합니다.')) {{
              document.getElementById('logDeleteAllForm').submit();
            }}
          }});
        }}
      }})();
    </script>
    """

    content = f"""
    <h2 style="margin-bottom:16px;">✏️ 수기 조정</h2>
    {branch_filter_html}
    <div class="card">
      <div style="margin-bottom:12px;">
        <label style="font-size:13px;color:#555;">품목 검색</label>
        <input id="itemSearch" placeholder="상품명 또는 품번 입력"
               oninput="filterItems()"
               style="margin-top:4px;">
      </div>
      <form method="post" action="/adjust">
        <label style="font-size:13px;color:#555;">품목 선택</label>
        <select name="item_key" id="itemSelect" required
                style="margin-bottom:14px;" onchange="loadQty(this)">
          <option value="">-- 품목 선택 --</option>
          {options_html}
        </select>

        <div id="qtyInfo" style="display:none;background:#f0f4ff;border-radius:8px;
             padding:12px;margin-bottom:14px;">
          <div style="display:flex;gap:16px;">
            <div>
              <div style="font-size:11px;color:#888;">QR 재고</div>
              <div id="qrQty" style="font-size:20px;font-weight:bold;color:#1E2761;">-</div>
            </div>
            <div>
              <div style="font-size:11px;color:#888;">RAW 재고</div>
              <div id="rawQty" style="font-size:20px;font-weight:bold;color:#64748B;">-</div>
            </div>
            <div>
              <div style="font-size:11px;color:#888;">차이</div>
              <div id="diffQty" style="font-size:20px;font-weight:bold;color:#EF4444;">-</div>
            </div>
          </div>
        </div>

        <label style="font-size:13px;color:#555;">현재 재고 수량으로 설정 (입력값이 곧 현재 재고)</label>
        <input name="delta" type="number" required placeholder="예: 5 또는 100"
               style="margin-bottom:14px;">
        <button class="btn" type="submit" style="width:100%;">조정 적용</button>
      </form>
    </div>
    {log_section}

    <script>
    const allOptions = Array.from(document.querySelectorAll('#itemSelect option'));

    function filterItems() {{
      const kw = document.getElementById('itemSearch').value.toLowerCase();
      const sel = document.getElementById('itemSelect');
      sel.innerHTML = '';
      allOptions.forEach(opt => {{
        if (!opt.value || opt.text.toLowerCase().includes(kw)) {{
          sel.appendChild(opt.cloneNode(true));
        }}
      }});
    }}

    function loadQty(sel) {{
      const val = sel.value;
      if (!val) {{
        document.getElementById('qtyInfo').style.display = 'none';
        return;
      }}
      const [branch, code] = val.split('|');
      fetch(`/api/qty?branch_code=${{branch}}&item_code=${{code}}`)
        .then(r => r.json())
        .then(d => {{
          document.getElementById('qrQty').textContent = d.qr_qty;
          document.getElementById('rawQty').textContent = d.raw_qty ?? '-';
          const diff = (d.raw_qty !== null) ? d.qr_qty - d.raw_qty : null;
          document.getElementById('diffQty').textContent = diff !== null
            ? (diff >= 0 ? '+' : '') + diff : '-';
          document.getElementById('qtyInfo').style.display = 'block';
        }});
    }}

    window.onload = function() {{
      const sel = document.getElementById('itemSelect');
      if (sel.value) loadQty(sel);
    }};
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "adjust"))


@app.post("/adjust")
async def adjust_post(
    session_token: str = Cookie(default=None),
    item_key: str = Form(...),
    delta: int = Form(...)
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    parts = item_key.split("|")
    if len(parts) != 2:
        return RedirectResponse(url="/adjust", status_code=303)
    branch_code, item_code = parts
    new_qty = adjust_quantity(branch_code, item_code, delta, absolute=True)
    conn = get_conn()
    item = conn.execute(
        "SELECT item_name FROM items WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    item_name = item["item_name"] if item else item_code
    conn.execute(
        "INSERT INTO adjustment_log (branch_code, item_name, item_code, delta, result_quantity, adjusted_at) VALUES (?, ?, ?, ?, ?, ?)",
        (branch_code, item_name, item_code, delta, new_qty, datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/adjust", status_code=303)


@app.post("/adjust/delete-logs")
async def adjust_delete_logs(
    request: Request,
    session_token: str = Cookie(default=None)
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("log_ids")
    if ids:
        conn = get_conn()
        if user["role"] == "master":
            conn.execute(
                f"DELETE FROM adjustment_log WHERE id IN ({','.join('?' for _ in ids)})",
                [int(i) for i in ids]
            )
        else:
            conn.execute(
                f"DELETE FROM adjustment_log WHERE id IN ({','.join('?' for _ in ids)}) AND branch_code=?",
                [int(i) for i in ids] + [user["branch_code"]]
            )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/adjust", status_code=303)


@app.post("/adjust/delete-all-logs")
async def adjust_delete_all_logs(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/adjust", status_code=303)
    conn = get_conn()
    if user["role"] == "master":
        conn.execute("DELETE FROM adjustment_log")
    else:
        conn.execute("DELETE FROM adjustment_log WHERE branch_code=?", (user["branch_code"],))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/adjust", status_code=303)


@app.get("/api/qty")
async def api_qty(branch_code: str, item_code: str,
                   session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return {"qr_qty": 0, "raw_qty": None}
    conn = get_conn()
    row = conn.execute(
        "SELECT quantity FROM inventory WHERE branch_code=? AND item_code=?",
        (branch_code, item_code)
    ).fetchone()
    conn.close()
    qr_qty = row["quantity"] if row else 0
    raw_data = fetch_raw_inventory()
    raw_map = {f"{r['branch_code']}|{r['item_code']}": r["quantity"] for r in raw_data}
    raw_qty = raw_map.get(f"{branch_code}|{item_code}", None)
    return {"qr_qty": qr_qty, "raw_qty": raw_qty}

# ── 스캔 이력 조회 ──────────────────────────────────────

@app.get("/scan-log", response_class=HTMLResponse)
async def scan_log_page(
    session_token: str = Cookie(default=None),
    search: str = "",
    date_filter: str = "",
    filter_branch: str = ""
):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    query = "SELECT * FROM scan_log WHERE 1=1"
    params: list = []
    if user["role"] == "branch":
        query += " AND branch_code=?"
        params.append(user["branch_code"])
    elif filter_branch:
        query += " AND branch_code=?"
        params.append(filter_branch)
    if search:
        query += " AND (item_name LIKE ? OR item_code LIKE ?)"
        params += [f"%{search}%", f"%{search}%"]
    if date_filter:
        query += " AND scanned_at LIKE ?"
        params.append(f"{date_filter}%")
    query += " ORDER BY id DESC LIMIT 200"
    logs = conn.execute(query, params).fetchall()
    conn.close()

    from datetime import timedelta, timezone
    KST = timezone(timedelta(hours=9))

    def to_kst_str(iso_str):
        if not iso_str:
            return "-"
        try:
            dt = datetime.fromisoformat(iso_str)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            dt_kst = dt.astimezone(KST)
            return dt_kst.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return iso_str[:16] if len(iso_str) >= 16 else iso_str

    rows_html = ""
    if not logs:
        rows_html = '<tr><td colspan="7" style="text-align:center;padding:20px;color:#888;">스캔 이력 없음</td></tr>'
    else:
        for lg in logs:
            type_label = "입고" if lg["scan_type"] == "IN" else "출고"
            type_color = "#22C55E" if lg["scan_type"] == "IN" else "#EF4444"
            check_cell = ""
            device_cell = ""
            if user["role"] == "master":
                check_cell = f'<td style="text-align:center;"><input type="checkbox" name="log_ids" value="{lg["id"]}" class="scanlog-check" style="width:16px;height:16px;"></td>'
                device_raw = (lg["device_info"] if "device_info" in lg.keys() else "") or "-"
                ip_raw = (lg["client_ip"] if "client_ip" in lg.keys() else "") or "-"
                device_short = device_raw[:30] + ("..." if len(device_raw) > 30 else "")
                device_cell = f'<td style="font-size:9px;color:#888;" title="{device_raw}">{device_short}<br>{ip_raw}</td>'
            rows_html += f"""
            <tr>
              {check_cell}
              <td>{to_kst_str(lg['scanned_at'])}</td>
              <td>{lg['branch_name'] or lg['branch_code']}</td>
              <td>{lg['item_name'] or '-'}</td>
              <td>{lg['item_code']}</td>
              <td style="color:{type_color};font-weight:bold;">{type_label}</td>
              <td>{lg['result_quantity']}</td>
              {device_cell}
            </tr>"""

    branch_filter_html = ""
    if user["role"] == "master":
        branch_options = '<option value="">전체 지점</option>'
        branches = get_branches()
        for b in branches:
            sel = "selected" if filter_branch == b["branch_code"] else ""
            branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'
        branch_filter_html = f"""
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">지점 필터</label>
          <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
        </div>"""

    delete_controls = ""
    header_check = ""
    if user["role"] == "master":
        header_check = '<th style="width:40px;text-align:center;"><input type="checkbox" id="scanlogAllCheck" style="width:16px;height:16px;"></th>'
        delete_controls = """
        <div style="display:flex;gap:8px;margin-bottom:12px;">
          <button type="button" class="btn" id="scanlogSelectAllBtn"
                  style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
          <button type="submit" class="btn btn-red"
                  style="font-size:12px;padding:6px 12px;"
                  onclick="return confirm('선택한 스캔 이력을 삭제할까요?')">선택삭제</button>
          <button type="button" class="btn btn-red" id="scanlogDeleteAllBtn"
                  style="font-size:12px;padding:6px 12px;">전체삭제</button>
        </div>"""

    table_open = '<form method="post" action="/scan-log/delete">' if user["role"] == "master" else '<div>'
    table_close = '</form>' if user["role"] == "master" else '</div>'

    content = f"""WLS
    <h2 style="margin-bottom:16px;">📜 스캔 이력</h2>
    <div class="card">
      <form method="get" action="/scan-log" style="display:flex;gap:8px;flex-wrap:wrap;">
        {branch_filter_html}
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">검색</label>
          <input name="search" value="{search}" placeholder="상품명/품번 검색" style="margin-top:4px;">
        </div>
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">날짜</label>
          <input name="date_filter" type="date" value="{date_filter}" style="margin-top:4px;">
        </div>
        <div style="display:flex;align-items:flex-end;gap:8px;">
          <button class="btn" type="submit">검색</button>
          <a href="/scan-log" style="padding:10px 14px;background:#eee;
             border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
        </div>
      </form>
    </div>
    <div class="card">
      {table_open}
        {delete_controls}
        <table>
          <thead><tr>
            {header_check}
            <th>시각</th><th>지점</th><th>상품명</th><th>품번</th><th>구분</th><th>처리후 재고</th>{'<th>기기/IP</th>' if user["role"] == "master" else ''}
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      {table_close}
    </div>
    <form method="post" action="/scan-log/delete-all" id="scanlogDeleteAllForm"></form>
    <script>
      (function() {{
        var allCheck = document.getElementById('scanlogAllCheck');
        var selectBtn = document.getElementById('scanlogSelectAllBtn');
        var deleteAllBtn = document.getElementById('scanlogDeleteAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.scanlog-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        if (deleteAllBtn) {{
          deleteAllBtn.addEventListener('click', function() {{
            if (confirm('전체 스캔 이력을 삭제합니다. 되돌릴 수 없습니다.')) {{
              document.getElementById('scanlogDeleteAllForm').submit();
            }}
          }});
        }}
      }})();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "scanlog"))

@app.get("/vendor-eval", response_class=HTMLResponse)
async def vendor_eval_page(session_token: str = Cookie(default=None), eval_month: str = ""):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user["role"] == "master":
        return RedirectResponse(url="/master/vendor-eval", status_code=303)

    branch_code = user["branch_code"]

    from datetime import date
    today = date.today()
    prev_month_num = today.month - 1 if today.month > 1 else 12
    prev_month_year = today.year if today.month > 1 else today.year - 1
    default_month = f"{prev_month_year}-{prev_month_num:02d}"

    conn = get_conn()
    criteria_list = conn.execute(
        "SELECT * FROM eval_criteria WHERE active = TRUE ORDER BY display_order"
    ).fetchall()

    if not criteria_list:
        conn.close()
        content = """
        <div class="card" style="text-align:center;padding:40px;">
          <div style="font-size:32px;">📝</div>
          <p style="color:#888;margin-top:12px;">등록된 평가 문항이 없습니다. 마스터에게 문항 등록을 요청해주세요.</p>
        </div>
        """
        return HTMLResponse(content=render_page(content, user, "vendor-eval"))

    criteria_data = []
    for c in criteria_list:
        options = conn.execute(
            "SELECT * FROM eval_criteria_option WHERE criteria_id=? ORDER BY score", (c["id"],)
        ).fetchall()
        criteria_data.append({
            "id": c["id"], "key": c["criteria_key"], "label": c["label"], "max_score": c["max_score"],
            "description": c["description"] or "",
            "options": [{"score": o["score"], "label": o["label"], "desc": o["description"] or "",
                         "requires_comment": bool(o["requires_comment"])} for o in options]
        })

    all_vendors = conn.execute("SELECT vendor_name FROM vendor_master ORDER BY vendor_name").fetchall()

    existing_round = conn.execute("""
        SELECT eval_month FROM vendor_evaluation_v2
        WHERE branch_code = ? AND TO_CHAR(created_at, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')
        ORDER BY created_at DESC LIMIT 1
    """, (branch_code,)).fetchone()

    if not eval_month:
        eval_month = existing_round["eval_month"] if existing_round else default_month

    done_vendors = conn.execute("""
        SELECT DISTINCT vendor_name FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
    """, (branch_code, eval_month)).fetchall()

    resubmit_requests = conn.execute("""
        SELECT vendor_name FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ? AND status = 'resubmit_requested'
    """, (branch_code, eval_month)).fetchall()
    conn.close()

    resubmit_names = [r["vendor_name"] for r in resubmit_requests]

    done_set = {v["vendor_name"] for v in done_vendors}
    remaining = [v["vendor_name"] for v in all_vendors if v["vendor_name"] not in done_set]

    total_count = len(all_vendors)
    done_count = total_count - len(remaining)

    if not all_vendors:
        content = """
        <div class="card" style="text-align:center;padding:40px;">
          <div style="font-size:32px;">🏢</div>
          <p style="color:#888;margin-top:12px;">등록된 거래처가 없습니다. 마스터에게 거래처 등록을 요청해주세요.</p>
        </div>
        """
        return HTMLResponse(content=render_page(content, user, "vendor-eval"))

    if not remaining:
        content = f"""
        <div class="card" style="text-align:center;padding:40px;">
          <div style="font-size:32px;">✅</div>
          <p style="font-weight:bold;margin-top:12px;">{eval_month} 거래처 평가가 완료되었습니다.</p>
          <p style="color:#888;font-size:13px;margin-top:4px;">{total_count}개 거래처 평가 완료</p>
          <div style="display:flex;gap:8px;justify-content:center;margin-top:16px;flex-wrap:wrap;">
            <a href="/vendor-eval/history?eval_month={eval_month}" class="btn" style="text-decoration:none;">📋 제출 내역 확인</a>
            <a href="/vendor-eval/history?eval_month={eval_month}&edit=1" class="btn" style="text-decoration:none;background:#F59E0B;">✏️ 거래처별 수정</a>
            <a href="/vendor-eval/reeval?eval_month={eval_month}" class="btn" style="text-decoration:none;background:#8B5CF6;">🔄 전체 재평가</a>
          </div>
        </div>
        """
        return HTMLResponse(content=render_page(content, user, "vendor-eval"))

    current_vendor = remaining[0]

    criteria_js = json.dumps(criteria_data, ensure_ascii=False)
    current_vendor_js = json.dumps(current_vendor, ensure_ascii=False)

    eval_month = default_month  # 항상 전월로 고정, 선택 불가
    month_selector_html = f'<p style="color:#888;font-size:12px;margin-bottom:12px;">평가월: {eval_month} (전월 고정)</p>'
    eval_month_js = json.dumps(eval_month, ensure_ascii=False)

    # 전체 문항을 한 화면에 렌더링 (순차진행 폐지)
    criteria_blocks = ""
    for idx, c in enumerate(criteria_data):
        num = idx + 1
        desc_html = f'<p style="color:#888;font-size:12px;margin-bottom:8px;">{c["description"]}</p>' if c["description"] else ""
        criteria_blocks += f"""
        <div class="ve-field">
            <label>{num}. {c['label']}</label>
            {desc_html}
            <div id="options_{c['key']}"></div>
            <textarea id="comment_{c['key']}" placeholder="사유를 입력하세요 (필수)"></textarea>
        </div>
        """

    resubmit_banner = ""
    if resubmit_names:
        names_str = ", ".join(resubmit_names)
        resubmit_banner = f"""
        <div class="card" style="background:#FEF3C7;border:1px solid #F59E0B;max-width:520px;margin:0 auto 12px;">
          <p style="font-size:13px;color:#92400E;">⚠️ 재평가를 요청한 거래처가 있습니다: <b>{names_str}</b></p>
        </div>
        """

    content = f"""
    {resubmit_banner}
    <style>
        .ve-card {{ background:#fff; border-radius:12px; padding:20px; max-width:520px; margin:0 auto; box-shadow:0 2px 8px rgba(0,0,0,0.08); }}
        .ve-card h2 {{ font-size:18px; margin-bottom:4px; }}
        .ve-progress {{ color:#2563eb; font-size:13px; font-weight:bold; margin-bottom:16px; }}
        .ve-field {{ margin-bottom:20px; padding-bottom:16px; border-bottom:1px solid #eee; }}
        .ve-field label {{ display:block; font-weight:bold; margin-bottom:6px; font-size:14px; }}
        .ve-option {{ border:1px solid #ddd; border-radius:8px; padding:10px; margin-bottom:8px; cursor:pointer; }}
        .ve-option.selected {{ border-color:#2563eb; background:#eff6ff; }}
        .ve-option .ve-label {{ font-weight:bold; font-size:14px; }}
        .ve-option .ve-desc {{ font-size:12px; color:#888; margin-top:2px; }}
        .ve-card textarea {{ width:100%; padding:8px; border:1px solid #ccc; border-radius:6px; box-sizing:border-box; font-size:14px; min-height:70px; margin-top:8px; display:none; }}
        .ve-card button {{ padding:12px 20px; border:none; border-radius:6px; font-size:15px; cursor:pointer; width:100%; }}
        .ve-btn-submit {{ background:#2563eb; color:#fff; }}
        .ve-btn-submit:disabled {{ background:#ccc; cursor:not-allowed; }}
    </style>
    <div class="ve-card">
        <div style="display:flex;justify-content:space-between;align-items:center;">
            <div class="ve-progress">진행 상황: {done_count + 1} / {total_count}</div>
            {'<a href="/vendor-eval/history?eval_month=' + eval_month + '&edit=1" style="font-size:12px;color:#2563eb;text-decoration:none;">← 이전 거래처 수정</a>' if done_count > 0 else ''}
        </div>
        <h2>거래처 평가 — {current_vendor}</h2>
        <p style="color:#888;font-size:12px;margin-bottom:16px;">평가월: {eval_month} (전월 고정)</p>
        {criteria_blocks}
        <button class="ve-btn-submit" id="submitBtn" onclick="submitEval()" disabled>제출</button>
    </div>

    <script>
        const criteriaData = {criteria_js};
        const currentVendor = {current_vendor_js};
        let currentEvalMonth = {eval_month_js};

        let selected = {{}};
        criteriaData.forEach(c => selected[c.key] = null);

        function renderOptions(criteria) {{
            const container = document.getElementById('options_' + criteria.key);
            container.innerHTML = '';
            criteria.options.forEach(opt => {{
                const div = document.createElement('div');
                div.className = 've-option';
                div.innerHTML = '<div class="ve-label">' + opt.label + '</div>' +
                    (opt.desc ? '<div class="ve-desc">' + opt.desc + '</div>' : '');
                div.onclick = () => {{
                    selected[criteria.key] = opt.score;
                    document.querySelectorAll('#options_' + criteria.key + ' .ve-option').forEach(o => o.classList.remove('selected'));
                    div.classList.add('selected');
                    const commentEl = document.getElementById('comment_' + criteria.key);
                    if (opt.requires_comment) {{
                        commentEl.style.display = 'block';
                    }} else {{
                        commentEl.style.display = 'none';
                        commentEl.value = '';
                    }}
                    checkAllValid();
                }};
                container.appendChild(div);
            }});
        }}

        function checkAllValid() {{
            const allValid = criteriaData.every(c => {{
                const score = selected[c.key];
                if (!score) return false;
                const opt = c.options.find(o => o.score === score);
                const needsComment = opt && opt.requires_comment;
                if (!needsComment) return true;
                const comment = document.getElementById('comment_' + c.key).value.trim();
                return comment.length > 0;
            }});
            document.getElementById('submitBtn').disabled = !allValid;
        }}

        criteriaData.forEach(c => {{
            renderOptions(c);
            document.getElementById('comment_' + c.key).addEventListener('input', checkAllValid);
        }});

        async function submitEval() {{
            const answers = criteriaData.map(c => ({{
                criteria_id: c.id,
                score: selected[c.key],
                comment: document.getElementById('comment_' + c.key).value.trim(),
                max_score: c.max_score
            }}));
            const payload = {{
                vendor_name: currentVendor,
                eval_month: currentEvalMonth,
                answers: answers
            }};
            const res = await fetch('/vendor-eval/submit', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(payload)
            }});
            if (res.ok) {{
                window.location.href = '/vendor-eval';
            }} else {{
                const err = await res.json();
                alert('오류: ' + (err.detail || '등록 실패'));
            }}
        }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))


@app.post("/vendor-eval/submit")
async def vendor_eval_submit(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return JSONResponse(status_code=403, content={"detail": "지점 계정만 등록 가능합니다."})

    branch_code = user["branch_code"]
    branch_name = next((b["branch_name"] for b in get_branches() if b["branch_code"] == branch_code), branch_code)

    data = await request.json()
    vendor_name = data.get("vendor_name", "").strip()
    eval_month = data.get("eval_month", "").strip()
    answers = data.get("answers", [])

    if not vendor_name or not eval_month or not answers:
        return JSONResponse(status_code=400, content={"detail": "필수 항목이 누락되었습니다."})

    conn = get_conn()

    normalized_scores = []
    for a in answers:
        score = a.get("score")
        max_score = a.get("max_score", 5)
        criteria_id = a.get("criteria_id")
        comment = a.get("comment", "").strip()

        if not score or not criteria_id:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "일부 문항이 답변되지 않았습니다."})

        opt = conn.execute(
            "SELECT requires_comment FROM eval_criteria_option WHERE criteria_id=? AND score=?",
            (criteria_id, score)
        ).fetchone()
        if opt and opt["requires_comment"] and not comment:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "사유가 필요한 문항에 사유가 누락되었습니다."})

        normalized_scores.append((score / max_score) * 5)

    total_score = round(sum(normalized_scores) / len(normalized_scores), 1) if normalized_scores else 0

    # 재제출 요청 상태였던 기존 레코드가 있으면 삭제 후 재생성 (같은 거래처+월 중복 방지)
    existing_resubmit = conn.execute(
        "SELECT id FROM vendor_evaluation_v2 WHERE branch_code=? AND vendor_name=? AND eval_month=? AND status='resubmit_requested'",
        (branch_code, vendor_name, eval_month)
    ).fetchone()
    if existing_resubmit:
        conn.execute("DELETE FROM vendor_evaluation_v2 WHERE id=?", (existing_resubmit["id"],))

    cur_result = conn.execute("""
        INSERT INTO vendor_evaluation_v2
        (branch_code, branch_name, vendor_name, eval_month, total_score, evaluated_by, status)
        VALUES (?, ?, ?, ?, ?, ?, 'completed')
    """, (branch_code, branch_name, vendor_name, eval_month, total_score, user["login_id"]))

    new_id_row = conn.execute("SELECT id FROM vendor_evaluation_v2 WHERE branch_code=? AND vendor_name=? AND eval_month=? ORDER BY created_at DESC LIMIT 1",
                               (branch_code, vendor_name, eval_month)).fetchone()
    evaluation_id = new_id_row["id"]

    for a in answers:
        conn.execute(
            "INSERT INTO vendor_evaluation_answer (evaluation_id, criteria_id, score, comment) VALUES (?, ?, ?, ?)",
            (evaluation_id, a.get("criteria_id"), a.get("score"), a.get("comment", "").strip())
        )

    conn.commit()
    conn.close()

    return JSONResponse(content={"status": "ok"})

@app.get("/vendor-eval/history", response_class=HTMLResponse)
async def vendor_eval_history(session_token: str = Cookie(default=None), eval_month: str = "", edit: str = ""):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return RedirectResponse(url="/login", status_code=303)

    branch_code = user["branch_code"]
    if not eval_month:
        return RedirectResponse(url="/vendor-eval", status_code=303)

    conn = get_conn()
    evaluations = conn.execute("""
        SELECT * FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ?
        ORDER BY vendor_name
    """, (branch_code, eval_month)).fetchall()

    rows_html = ""
    for ev in evaluations:
        answers = conn.execute("""
            SELECT a.score, a.comment, c.label
            FROM vendor_evaluation_answer a
            JOIN eval_criteria c ON c.id = a.criteria_id
            WHERE a.evaluation_id = ?
            ORDER BY c.display_order
        """, (ev["id"],)).fetchall()
        answers_html = "<br>".join(
            f"<b>{a['label']}</b>: {a['score']}점" + (f" ({a['comment']})" if a['comment'] else "")
            for a in answers
        )
        edit_btn = f'<a href="/vendor-eval/edit/{ev["id"]}" class="btn" style="text-decoration:none;font-size:12px;padding:6px 10px;">수정</a>' if edit else ""
        rows_html += f"""
        <tr>
            <td>{ev['vendor_name']}</td>
            <td style="font-size:12px;">{answers_html}</td>
            <td><b>{ev['total_score']}</b></td>
            <td>{edit_btn}</td>
        </tr>
        """
    conn.close()

    if not evaluations:
        rows_html = '<tr><td colspan="4" style="text-align:center;padding:20px;color:#888;">제출된 평가 없음</td></tr>'

    content = f"""
    <h2 style="margin-bottom:8px;">📋 {eval_month} 제출 내역</h2>
    <div class="card">
      <table>
        <thead><tr><th>거래처</th><th>답변 내역</th><th>총점</th><th></th></tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    <a href="/vendor-eval" style="display:inline-block;margin-top:8px;color:#2563eb;font-size:13px;text-decoration:none;">← 돌아가기</a>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))

@app.get("/vendor-eval/reeval")
async def vendor_eval_reeval_start(eval_month: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    first_ev = conn.execute("""
        SELECT id FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ?
        ORDER BY vendor_name LIMIT 1
    """, (user["branch_code"], eval_month)).fetchone()
    conn.close()

    if not first_ev:
        return RedirectResponse(url="/vendor-eval", status_code=303)

    return RedirectResponse(url=f"/vendor-eval/edit/{first_ev['id']}?reeval=1", status_code=303)

@app.get("/vendor-eval/edit/{evaluation_id}", response_class=HTMLResponse)
async def vendor_eval_edit_page(evaluation_id: int, session_token: str = Cookie(default=None), reeval: str = ""):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    ev = conn.execute("SELECT * FROM vendor_evaluation_v2 WHERE id=?", (evaluation_id,)).fetchone()
    if not ev or ev["branch_code"] != user["branch_code"]:
        conn.close()
        return RedirectResponse(url="/vendor-eval", status_code=303)

    criteria_list = conn.execute("SELECT * FROM eval_criteria WHERE active = TRUE ORDER BY display_order").fetchall()
    criteria_data = []
    for c in criteria_list:
        options = conn.execute(
            "SELECT * FROM eval_criteria_option WHERE criteria_id=? ORDER BY score", (c["id"],)
        ).fetchall()
        existing_answer = conn.execute(
            "SELECT score, comment FROM vendor_evaluation_answer WHERE evaluation_id=? AND criteria_id=?",
            (evaluation_id, c["id"])
        ).fetchone()
        criteria_data.append({
            "id": c["id"], "key": c["criteria_key"], "label": c["label"], "max_score": c["max_score"],
            "description": c["description"] or "",
            "options": [{"score": o["score"], "label": o["label"], "desc": o["description"] or "",
                         "requires_comment": bool(o["requires_comment"])} for o in options],
            "existing_score": existing_answer["score"] if existing_answer else None,
            "existing_comment": existing_answer["comment"] if existing_answer else ""
        })
    conn.close()

    criteria_js = json.dumps(criteria_data, ensure_ascii=False)
    vendor_name_js = json.dumps(ev["vendor_name"], ensure_ascii=False)
    eval_month_js = json.dumps(ev["eval_month"], ensure_ascii=False)

    # 전체 문항을 한 화면에 렌더링 (순차진행 폐지, 등록 화면과 동일한 방식)
    criteria_blocks = ""
    for idx, c in enumerate(criteria_data):
        num = idx + 1
        desc_html = f'<p style="color:#888;font-size:12px;margin-bottom:8px;">{c["description"]}</p>' if c["description"] else ""
        criteria_blocks += f"""
        <div class="ve-field">
            <label>{num}. {c['label']}</label>
            {desc_html}
            <div id="options_{c['key']}"></div>
            <textarea id="comment_{c['key']}" placeholder="사유를 입력하세요 (필수)"></textarea>
        </div>
        """

    content = f"""
    <style>
        .ve-card {{ background:#fff; border-radius:12px; padding:20px; max-width:520px; margin:0 auto; box-shadow:0 2px 8px rgba(0,0,0,0.08); }}
        .ve-card h2 {{ font-size:18px; margin-bottom:4px; }}
        .ve-field {{ margin-bottom:20px; padding-bottom:16px; border-bottom:1px solid #eee; }}
        .ve-field label {{ display:block; font-weight:bold; margin-bottom:6px; font-size:14px; }}
        .ve-option {{ border:1px solid #ddd; border-radius:8px; padding:10px; margin-bottom:8px; cursor:pointer; }}
        .ve-option.selected {{ border-color:#2563eb; background:#eff6ff; }}
        .ve-option .ve-label {{ font-weight:bold; font-size:14px; }}
        .ve-option .ve-desc {{ font-size:12px; color:#888; margin-top:2px; }}
        .ve-card textarea {{ width:100%; padding:8px; border:1px solid #ccc; border-radius:6px; box-sizing:border-box; font-size:14px; min-height:70px; margin-top:8px; display:none; }}
        .ve-card button {{ padding:12px 20px; border:none; border-radius:6px; font-size:15px; cursor:pointer; width:100%; }}
        .ve-btn-submit {{ background:#2563eb; color:#fff; }}
    </style>
    <div class="ve-card">
        <h2>거래처 평가 수정 — {ev['vendor_name']}</h2>
        <p style="color:#888;font-size:12px;margin-bottom:16px;">평가월: {ev['eval_month']}</p>
        {criteria_blocks}
        <button class="ve-btn-submit" id="submitBtn" onclick="submitEdit()">저장</button>
    </div>

    <script>
        const criteriaData = {criteria_js};
        const vendorName = {vendor_name_js};
        const evalMonth = {eval_month_js};
        const evaluationId = {evaluation_id};
        const reevalMode = {json.dumps(bool(reeval))};
        const currentEvalMonth = {json.dumps(ev["eval_month"], ensure_ascii=False)};

        let selected = {{}};
        criteriaData.forEach(c => selected[c.key] = c.existing_score);

        function renderOptions(criteria) {{
            const container = document.getElementById('options_' + criteria.key);
            container.innerHTML = '';
            criteria.options.forEach(opt => {{
                const div = document.createElement('div');
                div.className = 've-option' + (opt.score === criteria.existing_score ? ' selected' : '');
                div.innerHTML = '<div class="ve-label">' + opt.label + '</div>' +
                    (opt.desc ? '<div class="ve-desc">' + opt.desc + '</div>' : '');
                div.onclick = () => {{
                    selected[criteria.key] = opt.score;
                    document.querySelectorAll('#options_' + criteria.key + ' .ve-option').forEach(o => o.classList.remove('selected'));
                    div.classList.add('selected');
                    const commentEl = document.getElementById('comment_' + criteria.key);
                    if (opt.requires_comment) {{
                        commentEl.style.display = 'block';
                    }} else {{
                        commentEl.style.display = 'none';
                    }}
                }};
                container.appendChild(div);
            }});
        }}

        criteriaData.forEach(c => {{
            renderOptions(c);
            const commentEl = document.getElementById('comment_' + c.key);
            commentEl.value = c.existing_comment || '';
            const curOpt = c.options.find(o => o.score === c.existing_score);
            if (curOpt && curOpt.requires_comment) {{ commentEl.style.display = 'block'; }}
        }});

        async function submitEdit() {{
            const answers = criteriaData.map(c => ({{
                criteria_id: c.id,
                score: selected[c.key],
                comment: document.getElementById('comment_' + c.key).value.trim(),
                max_score: c.max_score
            }}));
            const payload = {{ vendor_name: vendorName, eval_month: evalMonth, answers: answers }};
            const res = await fetch('/vendor-eval/edit/' + evaluationId + '/submit', {{
                method: 'POST',
                headers: {{ 'Content-Type': 'application/json' }},
                body: JSON.stringify(payload)
            }});
            if (res.ok) {{
                alert('수정되었습니다.');
                if (reevalMode) {{
                    const nextRes = await fetch('/vendor-eval/reeval/next?evaluation_id=' + evaluationId + '&eval_month=' + encodeURIComponent(currentEvalMonth));
                    const nextData = await nextRes.json();
                    window.location.href = nextData.next_url;
                }} else {{
                    const result = await res.json();
                    window.location.href = result.next_url || '/vendor-eval';
                }}
            }} else {{
                const err = await res.json();
                alert('오류: ' + (err.detail || '수정 실패'));
            }}
        }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))


@app.post("/vendor-eval/edit/{evaluation_id}/submit")
async def vendor_eval_edit_submit(evaluation_id: int, request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    conn = get_conn()
    ev = conn.execute("SELECT * FROM vendor_evaluation_v2 WHERE id=?", (evaluation_id,)).fetchone()
    if not ev or ev["branch_code"] != user["branch_code"]:
        conn.close()
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    answers = data.get("answers", [])

    normalized_scores = []
    for a in answers:
        score = a.get("score")
        max_score = a.get("max_score", 5)
        criteria_id = a.get("criteria_id")
        comment = a.get("comment", "").strip()

        if not score or not criteria_id:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "일부 문항이 답변되지 않았습니다."})

        opt = conn.execute(
            "SELECT requires_comment FROM eval_criteria_option WHERE criteria_id=? AND score=?",
            (criteria_id, score)
        ).fetchone()
        if opt and opt["requires_comment"] and not comment:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "사유가 필요한 문항에 사유가 누락되었습니다."})

        normalized_scores.append((score / max_score) * 5)

    total_score = round(sum(normalized_scores) / len(normalized_scores), 1) if normalized_scores else 0

    conn.execute("DELETE FROM vendor_evaluation_answer WHERE evaluation_id=?", (evaluation_id,))
    for a in answers:
        conn.execute(
            "INSERT INTO vendor_evaluation_answer (evaluation_id, criteria_id, score, comment) VALUES (?, ?, ?, ?)",
            (evaluation_id, a.get("criteria_id"), a.get("score"), a.get("comment", "").strip())
        )
    conn.execute("UPDATE vendor_evaluation_v2 SET total_score=? WHERE id=?", (total_score, evaluation_id))

    branch_code = ev["branch_code"]
    eval_month = ev["eval_month"]
    all_vendors = conn.execute("SELECT vendor_name FROM vendor_master ORDER BY vendor_name").fetchall()
    done_vendors = conn.execute("""
        SELECT DISTINCT vendor_name FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
    """, (branch_code, eval_month)).fetchall()
    conn.commit()
    conn.close()

    done_set = {v["vendor_name"] for v in done_vendors}
    remaining = [v["vendor_name"] for v in all_vendors if v["vendor_name"] not in done_set]

    # 아직 미완료 거래처가 남아있으면 등록 화면으로, 없으면 제출 내역으로 이동
    next_url = "/vendor-eval" if remaining else f"/vendor-eval/history?eval_month={eval_month}&edit=1"

    return JSONResponse(content={"status": "ok", "next_url": next_url})

@app.get("/vendor-eval/reeval/next")
async def vendor_eval_reeval_next(evaluation_id: int, eval_month: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] == "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    conn = get_conn()
    current_ev = conn.execute("SELECT vendor_name FROM vendor_evaluation_v2 WHERE id=?", (evaluation_id,)).fetchone()
    if not current_ev:
        conn.close()
        return JSONResponse(content={"next_url": "/vendor-eval"})

    next_ev = conn.execute("""
        SELECT id FROM vendor_evaluation_v2
        WHERE branch_code = ? AND eval_month = ? AND vendor_name > ?
        ORDER BY vendor_name LIMIT 1
    """, (user["branch_code"], eval_month, current_ev["vendor_name"])).fetchone()
    conn.close()

    if next_ev:
        return JSONResponse(content={"next_url": f"/vendor-eval/edit/{next_ev['id']}?reeval=1"})
    else:
        return JSONResponse(content={"next_url": f"/vendor-eval/history?eval_month={eval_month}&edit=1"})

@app.get("/master/vendor-eval/status", response_class=HTMLResponse)
async def master_vendor_eval_status(session_token: str = Cookie(default=None), month: str = ""):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    from datetime import date
    today = date.today()
    if not month:
        prev_month_num = today.month - 1 if today.month > 1 else 12
        prev_month_year = today.year if today.month > 1 else today.year - 1
        month = f"{prev_month_year}-{prev_month_num:02d}"

    conn = get_conn()
    total_vendors = conn.execute("SELECT COUNT(*) as cnt FROM vendor_master").fetchone()["cnt"]

    rows_html = ""
    submitted_count = 0
    branches = get_branches()
    for b in branches:
        done_cnt = conn.execute("""
            SELECT COUNT(DISTINCT vendor_name) as cnt FROM vendor_evaluation_v2
            WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
        """, (b["branch_code"], month)).fetchone()["cnt"]

        is_complete = (done_cnt >= total_vendors and total_vendors > 0)
        if is_complete:
            submitted_count += 1
            status_badge = '<span class="badge-green">제출완료</span>'
        elif done_cnt > 0:
            status_badge = '<span class="badge-red">일부제출</span>'
        else:
            status_badge = '<span class="badge-red">미제출</span>'

        rows_html += f"""
        <tr>
            <td>{b['branch_name']}</td>
            <td>{done_cnt} / {total_vendors}</td>
            <td>{status_badge}</td>
        </tr>
        """
    conn.close()

    total_branches = len(get_branches())
    month_options = ""
    for i in range(6):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        ym = f"{y}-{m:02d}"
        sel = 'selected' if ym == month else ''
        month_options += f'<option value="{ym}" {sel}>{ym}</option>'

    content = f"""
    <h2 style="margin-bottom:16px;">📊 거래처평가 제출 현황</h2>
    <div class="card">
      <form method="get" action="/master/vendor-eval/status" style="display:flex;gap:8px;align-items:flex-end;">
        <div>
          <label style="font-size:12px;color:#888;">평가월</label>
          <select name="month" onchange="this.form.submit()">{month_options}</select>
        </div>
      </form>
    </div>
    <div class="card">
      <p style="font-size:14px;margin-bottom:12px;"><b>{month}</b> 기준 — 전체 {total_branches}개 지점 중 <b style="color:#22C55E;">{submitted_count}개 제출완료</b>, <b style="color:#EF4444;">{total_branches - submitted_count}개 미완료</b></p>
      <table>
        <thead><tr><th>지점</th><th>제출 거래처 수</th><th>상태</th></tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))

@app.get("/master/vendor-eval", response_class=HTMLResponse)
async def master_vendor_eval_page(session_token: str = Cookie(default=None), branch: str = "", month: str = ""):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()

    branches = conn.execute("SELECT DISTINCT branch_code, branch_name FROM vendor_evaluation_v2 ORDER BY branch_name").fetchall()
    months = conn.execute("""
        SELECT DISTINCT eval_month as ym FROM vendor_evaluation_v2
        WHERE eval_month IS NOT NULL ORDER BY ym DESC
    """).fetchall()

    query = "SELECT * FROM vendor_evaluation_v2 WHERE 1=1"
    params = []
    if branch:
        query += " AND branch_code = ?"
        params.append(branch)
    if month:
        query += " AND eval_month = ?"
        params.append(month)
    query += " ORDER BY created_at DESC"

    evaluations = conn.execute(query, params).fetchall()

    branch_options_html = '<option value="">전체 지점</option>'
    branches = get_branches()
    for b in branches:
        sel = 'selected' if branch == b["branch_code"] else ''
        branch_options_html += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'

    month_options_html = '<option value="">전체 기간</option>'
    for m in months:
        ym = m["ym"]
        sel = 'selected' if month == ym else ''
        month_options_html += f'<option value="{ym}" {sel}>{ym}</option>'

    rows_html = ""
    if not evaluations:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:20px;color:#888;">평가 데이터 없음</td></tr>'
    for ev in evaluations:
        answers = conn.execute("""
            SELECT a.score, a.comment, c.label
            FROM vendor_evaluation_answer a
            JOIN eval_criteria c ON c.id = a.criteria_id
            WHERE a.evaluation_id = ?
            ORDER BY c.display_order
        """, (ev["id"],)).fetchall()
        answers_html = "<br>".join(
            f"<b>{a['label']}</b>: {a['score']}점" + (f" ({a['comment']})" if a['comment'] else "")
            for a in answers
        )
        rows_html += f"""
        <tr>
            <td style="text-align:center;">
              <input type="checkbox" name="selected_ids" value="{ev['id']}" class="ve-check" style="width:16px;height:16px;">
            </td>
            <td>{ev['branch_name']}</td><td>{ev['vendor_name']}</td><td>{ev['eval_month']}</td>
            <td style="font-size:12px;">{answers_html}</td>
            <td><b>{ev['total_score']}</b></td>
            <td>{ev['evaluated_by']}<br><span style="font-size:11px;color:#888;">{str(ev['created_at'])[:16]}</span></td>
        </tr>
        """
    conn.close()

    query_string = f"branch={branch}&month={month}"

    from datetime import date as _date
    _today = _date.today()
    _prev_month_num = _today.month - 1 if _today.month > 1 else 12
    _prev_month_year = _today.year if _today.month > 1 else _today.year - 1
    summary_month = f"{_prev_month_year}-{_prev_month_num:02d}"

    conn3 = get_conn()
    summary_total_vendors = conn3.execute("SELECT COUNT(*) as cnt FROM vendor_master").fetchone()["cnt"]

    summary_chips_html = ""
    all_branches = get_branches(branch_type='branch')
    for b in all_branches:
        done_cnt = conn3.execute("""
            SELECT COUNT(DISTINCT vendor_name) as cnt FROM vendor_evaluation_v2
            WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
        """, (b["branch_code"], summary_month)).fetchone()["cnt"]

        is_complete = (done_cnt >= summary_total_vendors and summary_total_vendors > 0)
        if is_complete:
            chip_bg, chip_color, chip_icon = "#D1FAE5", "#065F46", "✅"
        elif done_cnt > 0:
            chip_bg, chip_color, chip_icon = "#FEF3C7", "#92400E", "⚠️"
        else:
            chip_bg, chip_color, chip_icon = "#FEE2E2", "#991B1B", "❌"

        summary_chips_html += f"""
        <div style="background:{chip_bg};border-radius:10px;padding:10px 8px;text-align:center;">
          <div style="font-size:12px;font-weight:bold;color:#333;margin-bottom:4px;">{b['branch_name']}</div>
          <div style="font-size:13px;font-weight:bold;color:{chip_color};">{chip_icon} {done_cnt}/{summary_total_vendors}</div>
        </div>
        """
    conn3.close()

    content = f"""
    <h2 style="margin-bottom:16px;">🤝 거래처 평가 (마스터 조회)</h2>
    <div class="card" style="margin-bottom:16px;">
      <h3 style="margin-bottom:10px;font-size:14px;">📊 {summary_month} 지점별 제출 현황 (전월 기준 자동)</h3>
      <div style="display:grid;grid-template-columns:repeat(auto-fill, minmax(110px, 1fr));gap:8px;">
        {summary_chips_html}
      </div>
    </div>
    <div class="card">
      <form method="get" action="/master/vendor-eval" style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        <div>
          <label style="font-size:12px;color:#888;">지점 필터</label>
          <select name="branch">{branch_options_html}</select>
        </div>
        <div>
          <label style="font-size:12px;color:#888;">월 필터</label>
          <select name="month">{month_options_html}</select>
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/master/vendor-eval" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
        <a href="/master/vendor-eval/summary?{query_string}" class="btn" style="text-decoration:none;background:#64748B;">📊 종합분석</a>
        <a href="/master/vendor-eval/export?{query_string}" class="btn" style="text-decoration:none;background:#22C55E;">⬇️ 엑셀 다운로드</a>
      </form>
    </div>
    <div class="card">
      <form method="post" action="/master/vendor-eval/delete" id="veDeleteForm">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
          <span style="font-size:13px;color:#888;">{len(evaluations)}개 항목</span>
          <div style="display:flex;gap:8px;">
            <button type="button" class="btn" id="veSelectAllBtn" style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
            <button type="button" class="btn" id="veResubmitBtn" style="background:#F59E0B;font-size:12px;padding:6px 12px;">재제출 요청</button>
            <button type="submit" class="btn btn-red" style="font-size:12px;padding:6px 12px;"
                    onclick="return confirm('선택한 평가를 완전 삭제합니다. 되돌릴 수 없습니다. 계속할까요?')">완전삭제</button>
            <button type="button" class="btn btn-red" id="veDeleteAllBtn" style="font-size:12px;padding:6px 12px;">전체삭제</button>
          </div>
        </div>
        <table>
            <thead><tr>
                <th style="width:40px;text-align:center;"><input type="checkbox" id="veAllCheck" style="width:16px;height:16px;"></th>
                <th>지점</th><th>거래처</th><th>평가월</th>
                <th>답변 내역</th><th>총점</th><th>등록자/일시</th>
            </tr></thead>
            <tbody>{rows_html}</tbody>
        </table>
      </form>
    </div>
    <form method="post" action="/master/vendor-eval/delete-all" id="veDeleteAllForm">
      <input type="hidden" name="branch" value="{branch}">
      <input type="hidden" name="month" value="{month}">
    </form>
    <script>
      (function() {{
        var allCheck = document.getElementById('veAllCheck');
        var selectBtn = document.getElementById('veSelectAllBtn');
        var deleteAllBtn = document.getElementById('veDeleteAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.ve-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        if (deleteAllBtn) {{
          deleteAllBtn.addEventListener('click', function() {{
            if (confirm('현재 필터 기준 평가 데이터를 전체 삭제합니다. 되돌릴 수 없습니다.')) {{
              document.getElementById('veDeleteAllForm').submit();
            }}
          }});
        }}
        var resubmitBtn = document.getElementById('veResubmitBtn');
        if (resubmitBtn) {{
          resubmitBtn.addEventListener('click', function() {{
            var checked = document.querySelectorAll('.ve-check:checked');
            if (checked.length === 0) {{ alert('재제출을 요청할 항목을 선택하세요.'); return; }}
            if (confirm(checked.length + '건에 대해 재제출을 요청합니다. 지점에 재평가 알림이 표시됩니다. 계속할까요?')) {{
              var form = document.getElementById('veDeleteForm');
              form.action = '/master/vendor-eval/request-resubmit';
              form.submit();
            }}
          }});
        }}
      }})();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))

@app.post("/master/vendor-eval/request-resubmit")
async def master_vendor_eval_request_resubmit(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("selected_ids")
    if ids:
        conn = get_conn()
        conn.execute(
            f"UPDATE vendor_evaluation_v2 SET status='resubmit_requested' WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/master/vendor-eval", status_code=303)

@app.post("/master/vendor-eval/delete")
async def master_vendor_eval_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("selected_ids")
    if ids:
        conn = get_conn()
        conn.execute(
            f"DELETE FROM vendor_evaluation_v2 WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/master/vendor-eval", status_code=303)


@app.post("/master/vendor-eval/delete-all")
async def master_vendor_eval_delete_all(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    branch = form.get("branch", "")
    month = form.get("month", "")

    query = "DELETE FROM vendor_evaluation_v2 WHERE 1=1"
    params = []
    if branch:
        query += " AND branch_code = ?"
        params.append(branch)
    if month:
        query += " AND eval_month = ?"
        params.append(month)

    conn = get_conn()
    conn.execute(query, params)
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/vendor-eval", status_code=303)


@app.get("/master/vendor-eval/export")
async def master_vendor_eval_export(session_token: str = Cookie(default=None), branch: str = "", month: str = ""):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    query = "SELECT * FROM vendor_evaluation_v2 WHERE 1=1"
    params = []
    if branch:
        query += " AND branch_code = ?"
        params.append(branch)
    if month:
        query += " AND eval_month = ?"
        params.append(month)
    query += " ORDER BY created_at DESC"

    conn = get_conn()
    evaluations = conn.execute(query, params).fetchall()

    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "거래처평가"
    headers = ["지점", "거래처", "평가월", "문항", "점수", "사유", "총점", "등록자", "등록일시"]
    ws.append(headers)

    for ev in evaluations:
        answers = conn.execute("""
            SELECT a.score, a.comment, c.label
            FROM vendor_evaluation_answer a
            JOIN eval_criteria c ON c.id = a.criteria_id
            WHERE a.evaluation_id = ?
            ORDER BY c.display_order
        """, (ev["id"],)).fetchall()
        if not answers:
            ws.append([ev["branch_name"], ev["vendor_name"], ev["eval_month"], "-", "-", "-",
                       float(ev["total_score"]) if ev["total_score"] is not None else None,
                       ev["evaluated_by"], str(ev["created_at"]) if ev["created_at"] else ""])
        for a in answers:
            ws.append([ev["branch_name"], ev["vendor_name"], ev["eval_month"], a["label"], a["score"], a["comment"],
                       float(ev["total_score"]) if ev["total_score"] is not None else None,
                       ev["evaluated_by"], str(ev["created_at"]) if ev["created_at"] else ""])
    conn.close()

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"거래처평가_{branch or '전체'}_{month or '전체'}.xlsx"
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@app.get("/master/vendor-eval/summary", response_class=HTMLResponse)
async def master_vendor_eval_summary(session_token: str = Cookie(default=None), branch: str = "", month: str = ""):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()

    where_clause = "WHERE 1=1"
    params = []
    if branch:
        where_clause += " AND branch_code = ?"
        params.append(branch)
    if month:
        where_clause += " AND eval_month = ?"
        params.append(month)

    vendor_rows = conn.execute(f"""
        SELECT vendor_name,
               COUNT(*) as cnt,
               ROUND(AVG(total_score), 1) as avg_total
        FROM vendor_evaluation_v2
        {where_clause}
        GROUP BY vendor_name
        ORDER BY avg_total DESC
    """, params).fetchall()

    branch_rows = []
    if not branch:
        month_where = "WHERE 1=1"
        month_params = []
        if month:
            month_where += " AND eval_month = ?"
            month_params.append(month)
        branch_rows = conn.execute(f"""
            SELECT branch_name,
                   COUNT(*) as cnt,
                   ROUND(AVG(total_score), 1) as avg_total
            FROM vendor_evaluation_v2
            {month_where}
            GROUP BY branch_name
            ORDER BY avg_total DESC
        """, month_params).fetchall()
    conn.close()

    vendor_rows_html = ""
    if not vendor_rows:
        vendor_rows_html = '<tr><td colspan="4" style="text-align:center;padding:20px;color:#888;">데이터 없음</td></tr>'
    for i, r in enumerate(vendor_rows, start=1):
        vendor_rows_html += f"""
        <tr>
            <td>{i}</td><td>{r['vendor_name']}</td><td>{r['cnt']}건</td>
            <td><b>{r['avg_total']}</b></td>
        </tr>
        """

    branch_section = ""
    if branch:
        branch_name_disp = next((b["branch_name"] for b in get_branches() if b["branch_code"] == branch), branch)
        branch_section = f"""
        <div class="card">
          <p style="font-size:13px;color:#888;">지점 필터가 적용되어 있어 지점별 비교는 생략됩니다 (현재: {branch_name_disp}).</p>
        </div>
        """
    else:
        branch_rows_html = ""
        if not branch_rows:
            branch_rows_html = '<tr><td colspan="3" style="text-align:center;padding:20px;color:#888;">데이터 없음</td></tr>'
        for r in branch_rows:
            branch_rows_html += f"""
            <tr>
                <td>{r['branch_name']}</td><td>{r['cnt']}건</td>
                <td><b>{r['avg_total']}</b></td>
            </tr>
            """
        branch_section = f"""
        <div class="card">
          <h3 style="font-size:15px;margin-bottom:12px;">지점별 평가 성향 비교</h3>
          <table>
              <thead><tr><th>지점</th><th>평가건수</th><th>총점평균</th></tr></thead>
              <tbody>{branch_rows_html}</tbody>
          </table>
          <p style="font-size:12px;color:#888;margin-top:8px;">
            ※ 총점평균이 낮을수록 그 지점이 거래처를 상대적으로 엄격하게 평가하는 경향입니다.
          </p>
        </div>
        """

    filter_label = []
    if branch:
        branch_name_disp = next((b["branch_name"] for b in get_branches() if b["branch_code"] == branch), branch)
        filter_label.append(f"지점: {branch_name_disp}")
    if month:
        filter_label.append(f"기간: {month}")
    filter_display = f'<p style="color:#888;font-size:13px;margin-bottom:12px;">현재 필터 — {" / ".join(filter_label)}</p>' if filter_label else ""

    content = f"""
    <h2 style="margin-bottom:8px;">📊 거래처 평가 종합분석</h2>
    {filter_display}

    <div class="card">
      <h3 style="font-size:15px;margin-bottom:12px;">거래처별 랭킹 (총점 높은 순)</h3>
      <table>
          <thead><tr><th>순위</th><th>거래처</th><th>평가건수</th><th>총점평균</th></tr></thead>
          <tbody>{vendor_rows_html}</tbody>
      </table>
    </div>

    {branch_section}

    <a href="/master/vendor-eval?branch={branch}&month={month}" style="display:inline-block;margin-top:8px;color:#2563eb;font-size:13px;text-decoration:none;">← 상세 목록으로 돌아가기</a>
    """
    return HTMLResponse(content=render_page(content, user, "vendor-eval"))


@app.get("/master/vendor-master", response_class=HTMLResponse)
async def vendor_master_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    vendors = conn.execute("SELECT * FROM vendor_master ORDER BY vendor_name").fetchall()
    conn.close()

    rows_html = ""
    if not vendors:
        rows_html = '<tr><td colspan="3" style="text-align:center;padding:20px;color:#888;">등록된 거래처 없음</td></tr>'
    for v in vendors:
        rows_html += f"""
        <tr>
          <td style="text-align:center;">
            <input type="checkbox" name="selected_ids" value="{v['id']}" class="vm-check" style="width:16px;height:16px;">
          </td>
          <td>{v['vendor_name']}</td>
          <td>{str(v['created_at'])[:10] if v['created_at'] else '-'}</td>
        </tr>
        """

    content = f"""
    <h2 style="margin-bottom:16px;">🏢 거래처 관리</h2>
    <div class="card">
      <h3 style="margin-bottom:8px;">거래처 추가</h3>
      <div style="display:flex;gap:8px;">
        <input type="text" id="newVendorName" placeholder="거래처명 입력" style="flex:1;">
        <button class="btn" type="button" onclick="addVendor()">추가</button>
      </div>
      <div id="addResult" style="margin-top:8px;font-size:13px;"></div>
    </div>
    <div class="card">
      <form method="post" action="/master/vendor-master/delete">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
          <span style="font-size:13px;color:#888;">{len(vendors)}개 거래처</span>
          <div style="display:flex;gap:8px;">
            <button type="button" class="btn" id="vmSelectAllBtn" style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
            <button type="submit" class="btn btn-red" style="font-size:12px;padding:6px 12px;"
                    onclick="return confirm('선택한 거래처를 삭제할까요? 기존 평가 기록은 유지됩니다.')">선택삭제</button>
          </div>
        </div>
        <table>
          <thead><tr>
            <th style="width:40px;text-align:center;"><input type="checkbox" id="vmAllCheck" style="width:16px;height:16px;"></th>
            <th>거래처명</th><th>등록일</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </form>
    </div>
    <script>
      async function addVendor() {{
        const input = document.getElementById('newVendorName');
        const name = input.value.trim();
        if (!name) {{ alert('거래처명을 입력하세요.'); return; }}
        const res = await fetch('/master/vendor-master/add', {{
          method: 'POST',
          headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ vendor_name: name }})
        }});
        const box = document.getElementById('addResult');
        if (res.ok) {{
          box.style.color = '#22C55E';
          box.innerText = '추가되었습니다.';
          setTimeout(() => location.reload(), 800);
        }} else {{
          const err = await res.json();
          box.style.color = '#EF4444';
          box.innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}
      (function() {{
        var allCheck = document.getElementById('vmAllCheck');
        var selectBtn = document.getElementById('vmSelectAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.vm-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
      }})();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/vendor-master/add")
async def vendor_master_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    vendor_name = data.get("vendor_name", "").strip()
    if not vendor_name:
        return JSONResponse(status_code=400, content={"detail": "거래처명을 입력하세요."})

    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO vendor_master (vendor_name) VALUES (?) ON CONFLICT (vendor_name) DO NOTHING",
            (vendor_name,)
        )
        conn.commit()
    finally:
        conn.close()

    return JSONResponse(content={"status": "ok"})


@app.post("/master/vendor-master/delete")
async def vendor_master_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("selected_ids")
    if ids:
        conn = get_conn()
        conn.execute(
            f"DELETE FROM vendor_master WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/master/vendor-master", status_code=303)

@app.get("/master/webhook-send-log", response_class=HTMLResponse)
async def master_webhook_send_log_page(
    session_token: str = Cookie(default=None),
    filter_branch: str = "",
    filter_sender: str = "",
    date_filter: str = ""
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    query = "SELECT * FROM teams_send_log WHERE (sent_by IS NULL OR sent_by NOT LIKE ?)"
    params: list = ["system_cron%"]

    if filter_branch:
        query += " AND branch_code=?"
        params.append(filter_branch)
    if filter_sender:
        query += " AND sent_by LIKE ?"
        params.append(f"%{filter_sender}%")
    if date_filter:
        query += " AND CAST(sent_at AS TEXT) LIKE ?"
        params.append(f"{date_filter}%")

    query += " ORDER BY sent_at DESC LIMIT 300"
    logs = conn.execute(query, params).fetchall()
    conn.close()

    from datetime import timedelta, timezone
    KST = timezone(timedelta(hours=9))

    def to_kst_str(raw_value):
        if not raw_value:
            return "-"
        try:
            dt = datetime.fromisoformat(str(raw_value).replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(KST).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(raw_value)[:16]

    rows_html = ""
    if not logs:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:20px;color:#888;">발송 이력 없음</td></tr>'
    else:
        for lg in logs:
            status_badge = '<span class="badge-green">성공</span>' if lg["success"] else '<span class="badge-red">실패</span>'
            msg_preview = (lg["message"] or "")[:60] + ("..." if len(lg["message"] or "") > 60 else "")
            rows_html += f"""
            <tr>
              <td style="font-size:12px;">{to_kst_str(lg['sent_at'])}</td>
              <td>{lg['sent_by'] or '-'}</td>
              <td>{lg['branch_code']}</td>
              <td>{lg['title'] or '-'}</td>
              <td style="font-size:12px;color:#555;">{msg_preview}</td>
              <td>{status_badge}</td>
            </tr>"""

    branch_options = '<option value="">전체 지점/채널</option>'
    for b in get_branches():
        sel = "selected" if filter_branch == b["branch_code"] else ""
        branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>📨 웹훅 발송 이력</h2>
    </div>
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">마스터/지점 계정이 직접 발송한 메시지만 표시됩니다 (자동 cron 발송 제외).</p>
    </div>
    <div class="card">
      <form method="get" action="/master/webhook-send-log" style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">대상 지점/채널</label>
          <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
        </div>
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">발송자 검색</label>
          <input name="filter_sender" value="{filter_sender}" placeholder="예: admin_hq" style="margin-top:4px;">
        </div>
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">날짜</label>
          <input name="date_filter" type="date" value="{date_filter}" style="margin-top:4px;">
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/master/webhook-send-log" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
    </div>
    <div class="card">
      <p style="font-size:13px;color:#888;margin-bottom:12px;">{len(logs)}건 (최근 300건까지 표시)</p>
      <table>
        <thead><tr>
          <th>발송시각</th><th>발송자</th><th>대상</th><th>제목</th><th>내용 미리보기</th><th>결과</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))

@app.get("/master/login-history", response_class=HTMLResponse)
async def master_login_history_page(
    session_token: str = Cookie(default=None),
    filter_login_id: str = "",
    date_filter: str = ""
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    query = "SELECT * FROM login_history WHERE 1=1"
    params: list = []

    if filter_login_id:
        query += " AND login_id=?"
        params.append(filter_login_id)
    if date_filter:
        query += " AND logged_in_at LIKE ?"
        params.append(f"{date_filter}%")

    query += " ORDER BY logged_in_at DESC LIMIT 300"
    logs = conn.execute(query, params).fetchall()

    all_login_ids = conn.execute(
        "SELECT DISTINCT login_id FROM login_history ORDER BY login_id"
    ).fetchall()
    conn.close()

    from datetime import timedelta, timezone
    KST = timezone(timedelta(hours=9))

    def to_kst_str(raw_value):
        if not raw_value:
            return "-"
        try:
            dt = datetime.fromisoformat(str(raw_value))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(KST).strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(raw_value)[:16]

    rows_html = ""
    if not logs:
        rows_html = '<tr><td colspan="5" style="text-align:center;padding:20px;color:#888;">접속 이력 없음</td></tr>'
    else:
        for lg in logs:
            device_raw = lg["device_info"] or "-"
            device_short = device_raw[:40] + ("..." if len(device_raw) > 40 else "")
            rows_html += f"""
            <tr>
              <td style="font-size:12px;">{to_kst_str(lg['logged_in_at'])}</td>
              <td>{lg['login_id']}</td>
              <td>{lg['role'] or '-'}</td>
              <td style="font-size:11px;color:#888;" title="{device_raw}">{device_short}</td>
              <td style="font-size:12px;">{lg['client_ip'] or '-'}</td>
            </tr>"""

    login_id_options = '<option value="">전체 계정</option>'
    for lid in all_login_ids:
        sel = "selected" if filter_login_id == lid["login_id"] else ""
        login_id_options += f'<option value="{lid["login_id"]}" {sel}>{lid["login_id"]}</option>'

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>🔐 계정별 접속 이력</h2>
    </div>
    <div class="card">
      <form method="get" action="/master/login-history" style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">계정</label>
          <select name="filter_login_id" style="margin-top:4px;">{login_id_options}</select>
        </div>
        <div style="flex:1;min-width:140px;">
          <label style="font-size:12px;color:#888;">날짜</label>
          <input name="date_filter" type="date" value="{date_filter}" style="margin-top:4px;">
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/master/login-history" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
    </div>
    <div class="card">
      <p style="font-size:13px;color:#888;margin-bottom:12px;">{len(logs)}건 (최근 300건까지 표시)</p>
      <table>
        <thead><tr>
          <th>접속시각</th><th>계정</th><th>역할</th><th>기기정보</th><th>IP</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))

@app.get("/master/branch-manage", response_class=HTMLResponse)
async def master_branch_manage_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    branches = get_branches()
    rows_html = ""
    for b in branches:
        b_type = b.get("branch_type") or "branch"
        type_badge = '<span class="badge-green">지점</span>' if b_type == "branch" else '<span class="badge-red">본사</span>'
        safe_name = b["branch_name"].replace("'", "")
        safe_login = b["login_id"].replace("'", "")
        rows_html += f"""
        <tr>
            <td>{b['branch_name']}</td>
            <td>{b['branch_code']}</td>
            <td>{b['login_id']}</td>
            <td>{b.get('password', '-')}</td>
            <td>{type_badge}</td>
            <td style="display:flex;gap:4px;flex-wrap:wrap;">
              <button class="btn" style="font-size:12px;padding:6px 10px;" onclick="editBranch('{b['branch_code']}', '{safe_name}', '{safe_login}', '{b.get('password','')}', '{b_type}')">수정</button>
              <button class="btn btn-red" style="font-size:12px;padding:6px 10px;" onclick="deleteBranch('{b['branch_code']}', '{safe_name}')">삭제</button>
            </td>
        </tr>
        """

    content = f"""
    <h2 style="margin-bottom:16px;">🏬 지점 관리</h2>
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">새 지점을 추가하면 로그인 계정이 함께 생성됩니다. 삭제 시 계정도 함께 삭제되며, 되돌릴 수 없습니다. "본사"로 등록하면 로그인 드롭다운과 미제출 알림 대상에서 제외됩니다.</p>
    </div>
    <div class="card">
      <h3 style="margin-bottom:8px;">새 지점 추가</h3>
      <div style="display:flex;flex-direction:column;gap:8px;max-width:400px;">
        <input type="text" id="newBranchName" placeholder="지점명 (예: 경기 파주점)">
        <input type="text" id="newBranchCode" placeholder="지점코드/로그인ID (예: 경기파주점, 공백없이)">
        <input type="password" id="newBranchPassword" placeholder="초기 비밀번호 (미입력시 1234)">
        <div style="display:flex;gap:16px;font-size:14px;">
          <label><input type="radio" name="newBranchType" value="branch" checked> 지점</label>
          <label><input type="radio" name="newBranchType" value="hq"> 본사</label>
        </div>
        <button class="btn" type="button" onclick="addBranch()">지점 추가</button>
      </div>
      <div id="addBranchResult" style="margin-top:8px;font-size:13px;"></div>
    </div>
    <div class="card">
      <table>
        <thead><tr><th>지점명</th><th>지점코드</th><th>로그인ID</th><th>비밀번호</th><th>구분</th><th></th></tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>

    <div id="editBranchModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:400px;width:90%;">
        <h3 style="margin-bottom:12px;">계정 정보 수정</h3>
        <input type="hidden" id="editBranchCode">
        <label style="font-size:13px;color:#555;">지점명</label>
        <input type="text" id="editBranchName" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <label style="font-size:13px;color:#555;">로그인 ID</label>
        <input type="text" id="editLoginId" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <label style="font-size:13px;color:#555;">비밀번호</label>
        <input type="text" id="editPassword" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <div style="display:flex;gap:16px;font-size:14px;margin-bottom:16px;">
          <label><input type="radio" name="editBranchType" value="branch"> 지점</label>
          <label><input type="radio" name="editBranchType" value="hq"> 본사</label>
        </div>
        <div style="display:flex;gap:8px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeEditBranch()">취소</button>
          <button class="btn" style="flex:1;" onclick="saveEditBranch()">저장</button>
        </div>
      </div>
    </div>

    <script>
      async function addBranch() {{
        const name = document.getElementById('newBranchName').value.trim();
        const code = document.getElementById('newBranchCode').value.trim();
        const pw = document.getElementById('newBranchPassword').value.trim() || '1234';
        const type = document.querySelector('input[name="newBranchType"]:checked').value;
        if (!name || !code) {{ alert('지점명과 지점코드를 입력하세요.'); return; }}
        if (code.includes(' ')) {{ alert('지점코드에는 공백을 사용할 수 없습니다.'); return; }}
        const res = await fetch('/master/branch-manage/add', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_name: name, branch_code: code, login_id: code, password: pw, branch_type: type }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('addBranchResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}
      async function deleteBranch(branchCode, branchName) {{
        if (!confirm(branchName + ' 지점을 삭제합니다. 계정도 함께 삭제되며 되돌릴 수 없습니다. 계속할까요?')) return;
        const res = await fetch('/master/branch-manage/delete', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: branchCode }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert('오류: ' + (err.detail || '삭제 실패'));
        }}
      }}
      function editBranch(branchCode, branchName, loginId, password, branchType) {{
        document.getElementById('editBranchCode').value = branchCode;
        document.getElementById('editBranchName').value = branchName;
        document.getElementById('editLoginId').value = loginId;
        document.getElementById('editPassword').value = password;
        document.querySelector('input[name="editBranchType"][value="' + branchType + '"]').checked = true;
        document.getElementById('editBranchModal').style.display = 'flex';
      }}
      function closeEditBranch() {{
        document.getElementById('editBranchModal').style.display = 'none';
      }}
      async function saveEditBranch() {{
        const branchCode = document.getElementById('editBranchCode').value;
        const branchName = document.getElementById('editBranchName').value.trim();
        const loginId = document.getElementById('editLoginId').value.trim();
        const password = document.getElementById('editPassword').value.trim();
        const branchType = document.querySelector('input[name="editBranchType"]:checked').value;
        if (!branchName || !loginId || !password) {{ alert('모든 항목을 입력하세요.'); return; }}
        const res = await fetch('/master/branch-manage/update', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: branchCode, branch_name: branchName, login_id: loginId, password: password, branch_type: branchType }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert('오류: ' + (err.detail || '수정 실패'));
        }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/branch-manage/add")
async def master_branch_manage_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    branch_name = data.get("branch_name", "").strip()
    branch_code = data.get("branch_code", "").strip()
    login_id = data.get("login_id", "").strip() or branch_code
    password = data.get("password", "").strip() or "1234"
    branch_type = data.get("branch_type", "branch").strip() or "branch"

    if not branch_name or not branch_code:
        return JSONResponse(status_code=400, content={"detail": "지점명과 지점코드를 입력하세요."})

    err = add_branch(branch_code, branch_name, login_id, password, branch_type)
    if err:
        return JSONResponse(status_code=400, content={"detail": err})
    return JSONResponse(content={"status": "ok"})


@app.post("/master/branch-manage/delete")
async def master_branch_manage_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    if not branch_code:
        return JSONResponse(status_code=400, content={"detail": "지점코드가 지정되지 않았습니다."})

    err = delete_branch(branch_code)
    if err:
        return JSONResponse(status_code=400, content={"detail": err})
    return JSONResponse(content={"status": "ok"})

@app.post("/master/branch-manage/update")
async def master_branch_manage_update(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    branch_name = data.get("branch_name", "").strip()
    login_id = data.get("login_id", "").strip()
    password = data.get("password", "").strip()
    branch_type = data.get("branch_type", "branch").strip() or "branch"

    if not branch_code or not branch_name or not login_id or not password:
        return JSONResponse(status_code=400, content={"detail": "모든 항목을 입력하세요."})

    err = update_branch_account(branch_code, branch_name, login_id, password, branch_type)
    if err:
        return JSONResponse(status_code=400, content={"detail": err})
    return JSONResponse(content={"status": "ok"})

@app.get("/master/custom-channel", response_class=HTMLResponse)
async def master_custom_channel_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    channels = conn.execute("SELECT * FROM teams_custom_channel ORDER BY created_at DESC").fetchall()
    schedules = conn.execute("SELECT * FROM teams_scheduled_message ORDER BY created_at DESC").fetchall()
    conn.close()

    channel_rows = ""
    for c in channels:
        channel_rows += f"""
        <tr>
            <td><input type="checkbox" class="cc-check" value="{c['id']}" style="width:16px;height:16px;"></td>
            <td>{c['channel_name']}</td>
            <td style="font-size:12px;color:#888;">{c['webhook_url'][:40]}...</td>
            <td><button class="btn btn-red" style="font-size:11px;padding:4px 8px;" onclick="deleteChannel({c['id']}, '{c['channel_name']}')">삭제</button></td>
        </tr>
        """

    schedule_rows = ""
    for s in schedules:
        status_badge = '<span class="badge-green">활성</span>' if s['active'] else '<span class="badge-red">중지</span>'
        last_sent = s['last_sent_date'] if s['last_sent_date'] else '없음'
        schedule_rows += f"""
        <tr>
            <td>{s['title'] or '(제목없음)'}</td>
            <td style="font-size:12px;">{s['message'][:30]}...</td>
            <td>{s['start_date']} / {s['interval_days']}일 간격</td>
            <td style="font-size:12px;">{last_sent}</td>
            <td>{status_badge}</td>
            <td>
              <button class="btn" style="font-size:11px;padding:4px 8px;background:#64748B;" onclick="toggleSchedule({s['id']}, {str(not s['active']).lower()})">{'중지' if s['active'] else '재개'}</button>
              <button class="btn btn-red" style="font-size:11px;padding:4px 8px;" onclick="deleteSchedule({s['id']})">삭제</button>
            </td>
        </tr>
        """

    all_targets_json = json.dumps(
        [{"code": "master", "name": "마스터(본사)", "type": "branch"}] +
        [{"code": b["branch_code"], "name": b["branch_name"], "type": "branch"} for b in get_branches()] +
        [{"code": str(c["id"]), "name": c["channel_name"], "type": "custom"} for c in channels],
        ensure_ascii=False
    )

    content = f"""
    <h2 style="margin-bottom:16px;">📡 자유 채널 & 반복 메시지</h2>

    <div class="card">
      <h3 style="margin-bottom:8px;">새 채널 추가</h3>
      <div style="display:flex;flex-direction:column;gap:8px;max-width:500px;">
        <input type="text" id="newChannelName" placeholder="채널 이름 (예: 운영기획팀 공지방)">
        <input type="text" id="newChannelUrl" placeholder="Teams 웹훅 URL (https://로 시작)">
        <button class="btn" type="button" onclick="addChannel()">채널 추가</button>
      </div>
      <div id="addChannelResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">등록된 채널</h3>
      <div style="margin-bottom:8px;">
        <button type="button" class="btn" id="ccSelectAllBtn" style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
        <button type="button" class="btn" id="ccSendBtn" style="background:#2563eb;font-size:12px;padding:6px 12px;">✉️ 선택 채널에 메시지 보내기</button>
      </div>
      <table>
        <thead><tr><th style="width:40px;"><input type="checkbox" id="ccAllCheck" style="width:16px;height:16px;"></th><th>채널명</th><th>웹훅</th><th></th></tr></thead>
        <tbody>{channel_rows if channel_rows else '<tr><td colspan="4" style="text-align:center;color:#888;">등록된 채널이 없습니다.</td></tr>'}</tbody>
      </table>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">🔁 반복 메시지 예약</h3>
      <div id="scheduleResult" style="margin-bottom:8px;font-size:13px;"></div>
      <button class="btn" type="button" onclick="openScheduleModal()">새 반복 메시지 만들기</button>
      <table style="margin-top:12px;">
        <thead><tr><th>제목</th><th>내용</th><th>주기</th><th>마지막 발송</th><th>상태</th><th></th></tr></thead>
        <tbody>{schedule_rows if schedule_rows else '<tr><td colspan="6" style="text-align:center;color:#888;">등록된 반복 메시지가 없습니다.</td></tr>'}</tbody>
      </table>
    </div>

    <div id="ccMessageModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:480px;width:90%;">
        <h3 style="margin-bottom:12px;">메시지 보내기</h3>
        <p id="ccSelectedList" style="font-size:12px;color:#888;margin-bottom:12px;"></p>
        <input type="text" id="ccMsgTitle" placeholder="제목" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <textarea id="ccMsgBody" placeholder="메시지 내용" style="width:100%;min-height:100px;padding:10px;border:1px solid #ccc;border-radius:6px;box-sizing:border-box;font-size:14px;"></textarea>
        <div style="display:flex;gap:8px;margin-top:16px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeCcModal()">취소</button>
          <button class="btn" style="flex:1;" onclick="sendCcMessage()">보내기</button>
        </div>
      </div>
    </div>

    <div id="scheduleModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:520px;width:90%;max-height:85vh;overflow-y:auto;">
        <h3 style="margin-bottom:12px;">반복 메시지 예약</h3>
        <p style="font-size:12px;color:#888;margin-bottom:8px;">대상 선택</p>
        <div id="scheduleTargets" style="max-height:150px;overflow-y:auto;border:1px solid #ddd;border-radius:6px;padding:8px;margin-bottom:12px;"></div>
        <input type="text" id="schTitle" placeholder="제목" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <textarea id="schMessage" placeholder="메시지 내용" style="width:100%;min-height:80px;padding:10px;border:1px solid #ccc;border-radius:6px;box-sizing:border-box;font-size:14px;margin-bottom:8px;"></textarea>
        <label style="font-size:12px;color:#888;">시작일</label>
        <input type="date" id="schStartDate" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <label style="font-size:12px;color:#888;">반복 간격 (일)</label>
        <input type="number" id="schInterval" value="5" min="1" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:12px;box-sizing:border-box;">
        <div style="display:flex;gap:8px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeScheduleModal()">취소</button>
          <button class="btn" style="flex:1;" onclick="saveSchedule()">저장</button>
        </div>
      </div>
    </div>

    <script>
      const allTargets = {all_targets_json};

      (function() {{
        var allCheck = document.getElementById('ccAllCheck');
        var selectBtn = document.getElementById('ccSelectAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.cc-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        document.getElementById('ccSendBtn').addEventListener('click', function() {{
          var checked = Array.from(document.querySelectorAll('.cc-check:checked')).map(c => c.value);
          if (checked.length === 0) {{ alert('메시지를 보낼 채널을 선택하세요.'); return; }}
          document.getElementById('ccSelectedList').innerText = checked.length + '개 채널에 발송됩니다.';
          document.getElementById('ccMessageModal').style.display = 'flex';
          document.getElementById('ccMessageModal').dataset.targets = JSON.stringify(checked);
        }});
      }})();

      function closeCcModal() {{
        document.getElementById('ccMessageModal').style.display = 'none';
        document.getElementById('ccMsgTitle').value = '';
        document.getElementById('ccMsgBody').value = '';
      }}

      async function sendCcMessage() {{
        const targets = JSON.parse(document.getElementById('ccMessageModal').dataset.targets || '[]');
        const title = document.getElementById('ccMsgTitle').value.trim();
        const body = document.getElementById('ccMsgBody').value.trim();
        if (!body) {{ alert('메시지 내용을 입력하세요.'); return; }}
        const res = await fetch('/master/custom-channel/broadcast', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ channel_ids: targets, title: title, message: body }})
        }});
        const result = await res.json();
        if (res.ok) {{
          alert('발송 완료: 성공 ' + result.success + '건, 실패 ' + result.failed + '건');
          closeCcModal();
        }} else {{ alert('오류: ' + (result.detail || '발송 실패')); }}
      }}

      async function addChannel() {{
        const name = document.getElementById('newChannelName').value.trim();
        const url = document.getElementById('newChannelUrl').value.trim();
        if (!name || !url) {{ alert('채널명과 URL을 입력하세요.'); return; }}
        if (!url.startsWith('https://')) {{ alert('올바른 URL 형식이 아닙니다.'); return; }}
        const res = await fetch('/master/custom-channel/add', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ channel_name: name, webhook_url: url }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('addChannelResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}

      async function deleteChannel(id, name) {{
        if (!confirm(name + ' 채널을 삭제합니다. 계속할까요?')) return;
        const res = await fetch('/master/custom-channel/delete', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ channel_id: id }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('삭제 실패'); }}
      }}

      function openScheduleModal() {{
        const container = document.getElementById('scheduleTargets');
        container.innerHTML = '';
        allTargets.forEach(t => {{
          const label = document.createElement('label');
          label.style.display = 'block';
          label.style.fontSize = '13px';
          label.style.marginBottom = '4px';
          label.innerHTML = '<input type="checkbox" class="sch-target" value="' + t.type + ':' + t.code + '" style="margin-right:6px;">' + t.name;
          container.appendChild(label);
        }});
        document.getElementById('schStartDate').value = new Date().toISOString().slice(0, 10);
        document.getElementById('scheduleModal').style.display = 'flex';
      }}

      function closeScheduleModal() {{
        document.getElementById('scheduleModal').style.display = 'none';
      }}

      async function saveSchedule() {{
        const targets = Array.from(document.querySelectorAll('.sch-target:checked')).map(c => c.value);
        const title = document.getElementById('schTitle').value.trim();
        const message = document.getElementById('schMessage').value.trim();
        const startDate = document.getElementById('schStartDate').value;
        const interval = parseInt(document.getElementById('schInterval').value);
        if (targets.length === 0) {{ alert('대상을 선택하세요.'); return; }}
        if (!message) {{ alert('메시지 내용을 입력하세요.'); return; }}
        if (!startDate || !interval) {{ alert('시작일과 반복 간격을 입력하세요.'); return; }}
        const res = await fetch('/master/custom-channel/schedule/add', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ targets: targets, title: title, message: message, start_date: startDate, interval_days: interval }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('scheduleResult').innerText = '오류: ' + (err.detail || '저장 실패');
        }}
      }}

      async function toggleSchedule(id, newActive) {{
        const res = await fetch('/master/custom-channel/schedule/toggle', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ schedule_id: id, active: newActive }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('변경 실패'); }}
      }}

      async function deleteSchedule(id) {{
        if (!confirm('이 반복 메시지를 삭제할까요?')) return;
        const res = await fetch('/master/custom-channel/schedule/delete', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ schedule_id: id }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('삭제 실패'); }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/custom-channel/add")
async def master_custom_channel_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    channel_name = data.get("channel_name", "").strip()
    webhook_url = data.get("webhook_url", "").strip()
    if not channel_name or not webhook_url:
        return JSONResponse(status_code=400, content={"detail": "채널명과 URL을 입력하세요."})
    conn = get_conn()
    conn.execute(
        "INSERT INTO teams_custom_channel (channel_name, webhook_url) VALUES (?, ?)",
        (channel_name, webhook_url)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/custom-channel/delete")
async def master_custom_channel_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    channel_id = data.get("channel_id")
    if not channel_id:
        return JSONResponse(status_code=400, content={"detail": "채널이 지정되지 않았습니다."})
    conn = get_conn()
    conn.execute("DELETE FROM teams_custom_channel WHERE id=?", (channel_id,))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/custom-channel/broadcast")
async def master_custom_channel_broadcast(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    channel_ids = data.get("channel_ids", [])
    title = data.get("title", "").strip() or "알림"
    message = data.get("message", "").strip()
    if not channel_ids or not message:
        return JSONResponse(status_code=400, content={"detail": "대상 또는 메시지가 비어있습니다."})

    device_info = request.headers.get("user-agent", "")[:200]
    conn = get_conn()
    success_count = 0
    fail_count = 0
    for cid in channel_ids:
        row = conn.execute("SELECT webhook_url, channel_name FROM teams_custom_channel WHERE id=?", (cid,)).fetchone()
        if not row:
            fail_count += 1
            continue
        ok, detail = send_teams_notification_to_url(row["webhook_url"], title, message, sent_by=user["branch_code"], device_info=device_info, target_label=f"custom:{row['channel_name']}")
        if ok:
            success_count += 1
        else:
            fail_count += 1
    conn.close()
    return JSONResponse(content={"status": "ok", "success": success_count, "failed": fail_count})


@app.post("/master/custom-channel/schedule/add")
async def master_schedule_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    targets = data.get("targets", [])
    title = data.get("title", "").strip()
    message = data.get("message", "").strip()
    start_date = data.get("start_date", "").strip()
    interval_days = data.get("interval_days")

    if not targets or not message or not start_date or not interval_days:
        return JSONResponse(status_code=400, content={"detail": "모든 필드를 입력하세요."})

    branch_targets = [t.split(":", 1)[1] for t in targets if t.startswith("branch:")]
    custom_targets = [t.split(":", 1)[1] for t in targets if t.startswith("custom:")]

    conn = get_conn()
    if branch_targets:
        conn.execute(
            "INSERT INTO teams_scheduled_message (title, message, target_type, target_codes, start_date, interval_days, created_by) VALUES (?, ?, 'branch', ?, ?, ?, ?)",
            (title, message, ",".join(branch_targets), start_date, interval_days, user["branch_code"])
        )
    if custom_targets:
        conn.execute(
            "INSERT INTO teams_scheduled_message (title, message, target_type, target_codes, start_date, interval_days, created_by) VALUES (?, ?, 'custom', ?, ?, ?, ?)",
            (title, message, ",".join(custom_targets), start_date, interval_days, user["branch_code"])
        )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/custom-channel/schedule/toggle")
async def master_schedule_toggle(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    schedule_id = data.get("schedule_id")
    active = data.get("active")
    conn = get_conn()
    conn.execute("UPDATE teams_scheduled_message SET active=? WHERE id=?", (active, schedule_id))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/custom-channel/schedule/delete")
async def master_schedule_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    schedule_id = data.get("schedule_id")
    conn = get_conn()
    conn.execute("DELETE FROM teams_scheduled_message WHERE id=?", (schedule_id,))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.get("/api/cron/send-scheduled-messages")
async def cron_send_scheduled_messages(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    from datetime import date
    today = date.today()
    conn = get_conn()
    schedules = conn.execute("SELECT * FROM teams_scheduled_message WHERE active = TRUE").fetchall()

    sent_count = 0
    for s in schedules:
        start = s["start_date"] if isinstance(s["start_date"], date) else datetime.strptime(str(s["start_date"]), "%Y-%m-%d").date()
        if today < start:
            continue
        days_since_start = (today - start).days
        if days_since_start % s["interval_days"] != 0:
            continue
        if s["last_sent_date"]:
            last = s["last_sent_date"] if isinstance(s["last_sent_date"], date) else datetime.strptime(str(s["last_sent_date"]), "%Y-%m-%d").date()
            if last == today:
                continue

        codes = s["target_codes"].split(",")
        for code in codes:
            if s["target_type"] == "branch":
                send_teams_notification(code, s["title"] or "반복 알림", s["message"], sent_by="system_cron")
            else:
                row = conn.execute("SELECT webhook_url, channel_name FROM teams_custom_channel WHERE id=?", (code,)).fetchone()
                if row:
                    send_teams_notification_to_url(row["webhook_url"], s["title"] or "반복 알림", s["message"], sent_by="system_cron", target_label=f"custom:{row['channel_name']}")
        conn.execute("UPDATE teams_scheduled_message SET last_sent_date=? WHERE id=?", (today.isoformat(), s["id"]))
        sent_count += 1

    conn.commit()
    conn.close()
    return JSONResponse(content={"processed": sent_count, "date": today.isoformat()})

@app.get("/api/cron/send-unsubmitted-reminder")
async def cron_send_unsubmitted_reminder(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    from datetime import date
    today = date.today()

    conn = get_conn()

    setting_row = conn.execute(
        "SELECT value FROM system_settings WHERE key='unsubmitted_reminder_enabled'"
    ).fetchone()
    reminder_enabled = (setting_row["value"] == "true") if setting_row else True

    if not reminder_enabled:
        conn.close()
        return JSONResponse(content={"skipped": True, "reason": "reminder disabled", "date": today.isoformat()})

    # 5, 10, 15, 20, 25, 30일에만 발송 (그 외 날짜에 호출되면 조용히 스킵)
    if today.day not in (5, 10, 15, 20, 25, 30):
        conn.close()
        return JSONResponse(content={"skipped": True, "reason": "not a reminder day", "date": today.isoformat()})

    # vendor-eval/status와 동일한 "전월" 계산 로직
    prev_month_num = today.month - 1 if today.month > 1 else 12
    prev_month_year = today.year if today.month > 1 else today.year - 1
    month = f"{prev_month_year}-{prev_month_num:02d}"

    total_vendors = conn.execute("SELECT COUNT(*) as cnt FROM vendor_master").fetchone()["cnt"]

    unsubmitted_branches = []
    branches = get_branches(branch_type='branch')
    for b in branches:
        done_cnt = conn.execute("""
            SELECT COUNT(DISTINCT vendor_name) as cnt FROM vendor_evaluation_v2
            WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
        """, (b["branch_code"], month)).fetchone()["cnt"]

        is_complete = (done_cnt >= total_vendors and total_vendors > 0)
        if not is_complete:
            unsubmitted_branches.append({
                "branch_code": b["branch_code"],
                "branch_name": b["branch_name"],
                "done_cnt": done_cnt,
                "total_vendors": total_vendors
            })

    sent_count = 0
    for ub in unsubmitted_branches:
        message = (
            f"{month} 거래처평가 미제출 안내드립니다.\n"
            f"현재 진행률: {ub['done_cnt']} / {ub['total_vendors']}\n"
            f"빠른 시일 내 제출 부탁드립니다."
        )
        send_teams_notification(
            ub["branch_code"],
            "거래처평가 미제출 알림",
            message,
            link_url="https://inventory-sync-teal.vercel.app/vendor-eval",
            link_text="거래처 평가하러 가기",
            sent_by="system_cron_unsubmitted"
        )
        sent_count += 1

    conn.close()
    return JSONResponse(content={
        "processed": sent_count,
        "month": month,
        "date": today.isoformat(),
        "unsubmitted_branches": [ub["branch_code"] for ub in unsubmitted_branches]
    })

@app.get("/api/cron/check-qr-raw-mismatch")
async def cron_check_qr_raw_mismatch(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    conn = get_conn()
    teams_setting_row = conn.execute(
        "SELECT value FROM system_settings WHERE key='qr_raw_mismatch_teams_enabled'"
    ).fetchone()
    teams_enabled = (teams_setting_row["value"] == "true") if teams_setting_row else False

    inventory_rows = conn.execute("SELECT * FROM inventory").fetchall()
    conn.close()

    raw_rows = fetch_raw_inventory()
    raw_map = {f"{r['branch_code']}|{r['item_code']}": r["quantity"] for r in raw_rows}

    mismatch_by_branch = {}
    for r in inventory_rows:
        key = f"{r['branch_code']}|{r['item_code']}"
        raw_qty = raw_map.get(key, 0)
        diff = r["quantity"] - raw_qty
        if diff != 0:
            mismatch_by_branch.setdefault(r["branch_code"], []).append({
                "item_name": r["item_name"],
                "item_code": r["item_code"],
                "diff": diff
            })

    branches = get_branches(branch_type='branch')
    branch_codes_set = {b["branch_code"] for b in branches}

    sent_count = 0
    for branch_code, items in mismatch_by_branch.items():
        if branch_code not in branch_codes_set:
            continue

        count = len(items)
        preview = "\n".join(f"{it['item_name']}_{it['diff']:+d}" for it in items[:5])
        more_note = f"\n...외 {count - 5}건" if count > 5 else ""

        title = "재고 불일치 알림"
        body = f"불일치 품목 {count}건 발견\n{preview}{more_note}"

        send_push_notification(branch_code, title, body, event_type="qr_raw_mismatch", url="/")
        if teams_enabled:
            send_teams_notification(branch_code, title, body, sent_by="system_cron_qr_raw_mismatch")
        sent_count += 1


    return JSONResponse(content={
        "processed_branches": sent_count,
        "total_mismatch_branches": len(mismatch_by_branch)
    }) 

@app.get("/api/cron/check-new-purchase")
async def cron_check_new_purchase(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    conn = get_conn()
    teams_setting_row = conn.execute(
        "SELECT value FROM system_settings WHERE key='purchase_new_teams_enabled'"
    ).fetchone()
    teams_enabled = (teams_setting_row["value"] == "true") if teams_setting_row else False

    last_check_row = conn.execute(
        "SELECT value FROM system_settings WHERE key='purchase_new_last_checked_at'"
    ).fetchone()
    conn.close()

    from datetime import datetime, timedelta, timezone
    now = datetime.now(timezone.utc)
    if last_check_row and last_check_row["value"]:
        try:
            last_checked = datetime.fromisoformat(last_check_row["value"])
        except Exception:
            last_checked = now - timedelta(minutes=10)
    else:
        last_checked = now - timedelta(minutes=10)

    rows, err = await fetch_purchase_history(limit=500)
    if err:
        return JSONResponse(status_code=500, content={"detail": err})

    new_rows = [r for r in rows if r.get("registered_at") and r["registered_at"] > last_checked.isoformat()]

    branches = get_branches(branch_type='branch')
    branch_name_to_code = {b["branch_name"]: b["branch_code"] for b in branches}

    # 지점별로 신규 발주를 묶음 (요약 알림 1건으로 발송하기 위함)
    rows_by_branch: Dict[str, list] = {}
    for r in new_rows:
        branch_name = r.get("branch", "")
        branch_code = branch_name_to_code.get(branch_name)
        if not branch_code:
            continue
        rows_by_branch.setdefault(branch_code, []).append(r)

    sent_count = 0
    for branch_code, branch_rows in rows_by_branch.items():
        branch_name = branch_rows[0].get("branch", "")
        count = len(branch_rows)

        preview = "\n".join(
            f"{r.get('vendor', '-')} {r.get('product_name', '-')}_{r.get('quantity', '-')}개"
            for r in branch_rows[:10]
        )
        more_note = f"\n...외 {count - 10}건" if count > 10 else ""

        title = "발주내역 알림"
        body = f"{branch_name} 신규 발주 {count}건\n{preview}{more_note}"

        send_push_notification(branch_code, title, body, event_type="purchase_new", url="/purchase-history")
        if teams_enabled:
            send_teams_notification(branch_code, title, body, sent_by="system_cron_purchase_new")
        sent_count += 1

    conn2 = get_conn()
    existing_ts = conn2.execute(
        "SELECT value FROM system_settings WHERE key='purchase_new_last_checked_at'"
    ).fetchone()
    if existing_ts:
        conn2.execute(
            "UPDATE system_settings SET value=?, updated_at=NOW() WHERE key='purchase_new_last_checked_at'",
            (now.isoformat(),)
        )
    else:
        conn2.execute(
            "INSERT INTO system_settings (key, value) VALUES ('purchase_new_last_checked_at', ?)",
            (now.isoformat(),)
        )
    conn2.commit()
    conn2.close()

    return JSONResponse(content={"new_count": len(new_rows), "sent_count": sent_count})

@app.post("/master/teams-webhook/test-unsubmitted-reminder")
async def master_test_unsubmitted_reminder(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    data = await request.json()
    target_branch_code = data.get("branch_code", "").strip()
    if not target_branch_code:
        return JSONResponse(status_code=400, content={"detail": "테스트 대상 채널 코드가 필요합니다."})

    from datetime import date
    today = date.today()

    prev_month_num = today.month - 1 if today.month > 1 else 12
    prev_month_year = today.year if today.month > 1 else today.year - 1
    month = f"{prev_month_year}-{prev_month_num:02d}"

    conn = get_conn()
    total_vendors = conn.execute("SELECT COUNT(*) as cnt FROM vendor_master").fetchone()["cnt"]

    unsubmitted_branches = []
    branches = get_branches(branch_type='branch')
    for b in branches:
        done_cnt = conn.execute("""
            SELECT COUNT(DISTINCT vendor_name) as cnt FROM vendor_evaluation_v2
            WHERE branch_code = ? AND eval_month = ? AND status = 'completed'
        """, (b["branch_code"], month)).fetchone()["cnt"]

        is_complete = (done_cnt >= total_vendors and total_vendors > 0)
        if not is_complete:
            unsubmitted_branches.append({
                "branch_code": b["branch_code"],
                "branch_name": b["branch_name"],
                "done_cnt": done_cnt,
                "total_vendors": total_vendors
            })
    conn.close()

    # 실제 운영 로직(cron_send_unsubmitted_reminder)과 동일한 메시지를 지점별로 개별 생성,
    # 실제 지점이 아닌 target_branch_code(테스트채널) 하나로만 전부 발송
    sent_count = 0
    for ub in unsubmitted_branches:
        message = (
            f"[테스트] {month} 거래처평가 미제출 안내드립니다.\n"
            f"(실제 대상 지점: {ub['branch_name']})\n"
            f"현재 진행률: {ub['done_cnt']} / {ub['total_vendors']}\n"
            f"빠른 시일 내 제출 부탁드립니다."
        )
        send_teams_notification(
            target_branch_code,
            f"🧪[테스트] 거래처평가 미제출 알림 ({ub['branch_name']})",
            message,
            link_url="https://inventory-sync-teal.vercel.app/vendor-eval",
            link_text="거래처 평가하러 가기",
            sent_by=f"test_by_{user['login_id']}"
        )
        sent_count += 1

    return JSONResponse(content={
        "success": True,
        "month": month,
        "unsubmitted_count": len(unsubmitted_branches),
        "sent_count": sent_count,
        "unsubmitted_branches": [ub["branch_code"] for ub in unsubmitted_branches]
    })

@app.post("/master/teams-webhook/test-qr-raw-mismatch")
async def master_test_qr_raw_mismatch(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    data = await request.json()
    target_branch_code = data.get("branch_code", "").strip()
    if not target_branch_code:
        return JSONResponse(status_code=400, content={"detail": "테스트 대상 채널 코드가 필요합니다."})

    conn = get_conn()
    inventory_rows = conn.execute("SELECT * FROM inventory").fetchall()
    conn.close()

    raw_rows = fetch_raw_inventory()
    raw_map = {f"{r['branch_code']}|{r['item_code']}": r["quantity"] for r in raw_rows}

    mismatch_by_branch = {}
    for r in inventory_rows:
        key = f"{r['branch_code']}|{r['item_code']}"
        raw_qty = raw_map.get(key, 0)
        diff = r["quantity"] - raw_qty
        if diff != 0:
            mismatch_by_branch.setdefault(r["branch_code"], []).append({
                "item_name": r["item_name"],
                "item_code": r["item_code"],
                "diff": diff
            })

    branches = get_branches(branch_type='branch')
    branch_names = {b["branch_code"]: b["branch_name"] for b in branches}

    sent_count = 0
    for branch_code, items in mismatch_by_branch.items():
        if branch_code not in branch_names:
            continue
        count = len(items)
        preview = "\n".join(f"- {it['item_name']} ({it['item_code']}): {it['diff']:+d}" for it in items[:5])
        more_note = f"\n...외 {count - 5}건" if count > 5 else ""

        title = f"🧪[테스트] 재고 불일치 알림 ({branch_names[branch_code]})"
        body = f"[테스트] 실제 대상 지점: {branch_names[branch_code]}\n불일치 품목 {count}건 발견\n{preview}{more_note}"

        send_teams_notification(target_branch_code, title, body, sent_by=f"test_by_{user['login_id']}")
        sent_count += 1

    return JSONResponse(content={
        "success": True,
        "mismatch_branch_count": sent_count
    })

@app.post("/master/teams-webhook/test-purchase-new")
async def master_test_purchase_new(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    data = await request.json()
    target_branch_code = data.get("branch_code", "").strip()
    if not target_branch_code:
        return JSONResponse(status_code=400, content={"detail": "테스트 대상 채널 코드가 필요합니다."})

    rows, err = await fetch_purchase_history(limit=10)
    if err:
        return JSONResponse(status_code=500, content={"detail": err})

    if not rows:
        return JSONResponse(content={
            "success": True,
            "purchase_count": 0
        })

    branches = get_branches(branch_type='branch')
    branch_name_to_code = {b["branch_name"]: b["branch_code"] for b in branches}

    sent_count = 0
    for r in rows[:5]:
        branch_name = r.get("branch", "-")
        title = f"🧪[테스트] 발주내역 알림 ({branch_name})"
        body = (
            f"[테스트] 실제 대상 지점: {branch_name}\n"
            f"거래처: {r.get('vendor', '-')}\n"
            f"상품: {r.get('product_name', '-')}\n"
            f"수량: {r.get('quantity', '-')}"
        )
        send_teams_notification(target_branch_code, title, body, sent_by=f"test_by_{user['login_id']}")
        sent_count += 1

    return JSONResponse(content={
        "success": True,
        "purchase_count": sent_count
    })

@app.post("/master/toggle-unsubmitted-reminder")
async def master_toggle_unsubmitted_reminder(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM system_settings WHERE key='unsubmitted_reminder_enabled'"
    ).fetchone()
    current = (row["value"] == "true") if row else True
    new_value = "false" if current else "true"

    if row:
        conn.execute(
            "UPDATE system_settings SET value=?, updated_at=NOW() WHERE key='unsubmitted_reminder_enabled'",
            (new_value,)
        )
    else:
        conn.execute(
            "INSERT INTO system_settings (key, value) VALUES ('unsubmitted_reminder_enabled', ?)",
            (new_value,)
        )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/teams-webhook", status_code=303)

@app.post("/master/toggle-qr-raw-mismatch-teams")
async def master_toggle_qr_raw_mismatch_teams(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM system_settings WHERE key='qr_raw_mismatch_teams_enabled'"
    ).fetchone()
    current = (row["value"] == "true") if row else False
    new_value = "false" if current else "true"

    if row:
        conn.execute(
            "UPDATE system_settings SET value=?, updated_at=NOW() WHERE key='qr_raw_mismatch_teams_enabled'",
            (new_value,)
        )
    else:
        conn.execute(
            "INSERT INTO system_settings (key, value) VALUES ('qr_raw_mismatch_teams_enabled', ?)",
            (new_value,)
        )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/notification-settings", status_code=303)

@app.post("/master/toggle-purchase-new-teams")
async def master_toggle_purchase_new_teams(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    row = conn.execute(
        "SELECT value FROM system_settings WHERE key='purchase_new_teams_enabled'"
    ).fetchone()
    current = (row["value"] == "true") if row else False
    new_value = "false" if current else "true"

    if row:
        conn.execute(
            "UPDATE system_settings SET value=?, updated_at=NOW() WHERE key='purchase_new_teams_enabled'",
            (new_value,)
        )
    else:
        conn.execute(
            "INSERT INTO system_settings (key, value) VALUES ('purchase_new_teams_enabled', ?)",
            (new_value,)
        )
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/notification-settings", status_code=303)

@app.get("/master/notification-settings", response_class=HTMLResponse)
async def master_notification_settings_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    toggle_cards_html = ""
    for alert_key, alert_info in TEAMS_ALERT_TYPES.items():
        conn = get_conn()
        row = conn.execute(
            "SELECT value FROM system_settings WHERE key=?", (alert_info["toggle_key"],)
        ).fetchone()
        conn.close()
        is_enabled = (row["value"] == "true") if row else alert_info["default_enabled"]
        status_text = f"🟢 켜짐 ({alert_info['on_desc']})" if is_enabled else f"🔴 꺼짐 ({alert_info['off_desc']})"
        btn_label = "끄기" if is_enabled else "켜기"
        toggle_cards_html += f"""
        <div class="card" style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;margin-bottom:12px;">
          <div>
            <div style="font-weight:bold;margin-bottom:4px;">{alert_info['label']}</div>
            <div style="font-size:13px;color:#888;">{status_text}</div>
          </div>
          <form method="post" action="{alert_info['toggle_route']}" style="margin:0;">
            <button type="submit" class="btn" style="font-size:13px;padding:8px 16px;">{btn_label}</button>
          </form>
        </div>
        """

    content = f"""
    <h2 style="margin-bottom:16px;">⏰ 알림 설정</h2>
    {toggle_cards_html}
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">이 설정은 <a href="/master/teams-webhook">팀즈웹훅 관리</a> 페이지의 토글과 동일하게 연동됩니다.</p>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "notif-settings"))

@app.get("/master/teams-webhook", response_class=HTMLResponse)
async def master_teams_webhook_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    existing_rows = conn.execute("SELECT branch_code, webhook_url, channel_name FROM teams_webhook").fetchall()
    existing = {r["branch_code"]: {"url": r["webhook_url"], "name": r["channel_name"] or ""} for r in existing_rows}

    draft_rows = conn.execute("SELECT * FROM teams_draft_message ORDER BY updated_at DESC").fetchall()
    drafts_map: Dict[str, list] = {}
    for d in draft_rows:
        drafts_map.setdefault(d["branch_code"], []).append(dict(d))

    schedules = conn.execute("SELECT * FROM teams_scheduled_message WHERE target_type='branch'").fetchall()
    schedule_map: Dict[str, list] = {}
    for s in schedules:
        for code in s["target_codes"].split(","):
            schedule_map.setdefault(code, []).append(dict(s))
    conn.close()

    known_codes = {"master"} | {b["branch_code"] for b in get_branches()}
    free_channels = [
        {"branch_code": code, "branch_name": info["name"] or code}
        for code, info in existing.items()
        if code not in known_codes
    ]
    targets = [{"branch_code": "master", "branch_name": "마스터(본사)"}] + get_branches() + free_channels

    rows_html = ""
    reminder_setting_row = None
    conn2 = get_conn()
    reminder_setting_row = conn2.execute(
        "SELECT value FROM system_settings WHERE key='unsubmitted_reminder_enabled'"
    ).fetchone()
    conn2.close()
    reminder_enabled = (reminder_setting_row["value"] == "true") if reminder_setting_row else True
    reminder_status_text = "🟢 켜짐 (매월 5,10,15,20,25,30일 자동발송)" if reminder_enabled else "🔴 꺼짐 (자동발송 안 함)"
    reminder_btn_label = "끄기" if reminder_enabled else "켜기"

    for t in targets:
        info = existing.get(t["branch_code"])
        is_registered = bool(info)
        has_webhook = bool(info and info["url"])
        channel_display = (info["name"] if info and info["name"] else "(이름 미설정)") if is_registered else "-"
        url_display = (info["url"][:35] + "...") if has_webhook else "(URL 미등록)" if is_registered else "(미등록)"
        safe_channel_name = (info["name"] if info else "").replace("'", "")

        branch_drafts = drafts_map.get(t["branch_code"], [])
        branch_scheds = schedule_map.get(t["branch_code"], [])
        draft_badge = f'<span class="badge-green">임시저장 {len(branch_drafts)}건</span>' if branch_drafts else ''
        sched_badge = f'<span class="badge-green">반복 {len(branch_scheds)}건</span>' if branch_scheds else ''

        expand_btn = ""
        detail_row = ""
        if has_webhook:
            expand_btn = f'<button class="btn" style="font-size:11px;padding:4px 8px;background:#8B5CF6;" onclick="toggleAccordion(\'{t["branch_code"]}\')">▼ 상세</button>'

            draft_items_html = ""
            for d in branch_drafts:
                d_title = (d["title"] or "(제목없음)").replace("'", "")
                d_msg_preview = (d["message"] or "")[:40].replace("'", "")
                draft_items_html += f"""
                <div style="border:1px solid #ddd;border-radius:6px;padding:8px;margin-bottom:6px;font-size:12px;">
                  <b>{d_title}</b><br>
                  <span style="color:#888;">{d_msg_preview}{'...' if len(d['message'] or '') > 40 else ''}</span>
                  <div style="margin-top:6px;display:flex;gap:4px;">
                    <button class="btn" style="font-size:11px;padding:3px 8px;" onclick="editDraftItem({d['id']}, '{t['branch_code']}')">수정</button>
                    <button class="btn" style="font-size:11px;padding:3px 8px;background:#2563eb;" onclick="sendDraftItem({d['id']}, '{t['branch_code']}')">바로 발송</button>
                    <button class="btn btn-red" style="font-size:11px;padding:3px 8px;" onclick="deleteDraftItem({d['id']}, '{t['branch_code']}')">삭제</button>
                  </div>
                </div>
                """
            if not draft_items_html:
                draft_items_html = '<p style="font-size:12px;color:#888;">임시저장된 메시지가 없습니다.</p>'

            sched_items_html = ""
            for s in branch_scheds:
                s_title = (s["title"] or "(제목없음)").replace("'", "")
                sched_items_html += f"""
                <div style="border:1px solid #ddd;border-radius:6px;padding:8px;margin-bottom:6px;font-size:12px;">
                  <b>{s_title}</b> — {s['interval_days']}일 간격 ({'활성' if s['active'] else '중지'})<br>
                  <span style="color:#888;">{(s['message'] or '')[:40]}{'...' if len(s['message'] or '') > 40 else ''}</span>
                  <div style="margin-top:6px;display:flex;gap:4px;flex-wrap:wrap;">
                    <button class="btn" style="font-size:11px;padding:3px 8px;" onclick="editSchedItem({s['id']}, '{t['branch_code']}')">수정</button>
                    <button class="btn" style="font-size:11px;padding:3px 8px;background:#64748B;" onclick="toggleSchedItem({s['id']}, {str(not s['active']).lower()}, '{t['branch_code']}')">{'중지' if s['active'] else '재개'}</button>
                    <button class="btn btn-red" style="font-size:11px;padding:3px 8px;" onclick="deleteSchedItem({s['id']}, '{t['branch_code']}')">삭제</button>
                  </div>
                </div>
                """
            if not sched_items_html:
                sched_items_html = '<p style="font-size:12px;color:#888;">등록된 반복 메시지가 없습니다.</p>'

            detail_row = f"""
            <tr id="accordion_{t['branch_code']}" style="display:none;">
              <td colspan="6" style="background:#f8fafc;padding:16px;">
                <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">
                  <div>
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                      <h4 style="font-size:13px;">📝 임시저장 목록</h4>
                      <button class="btn" style="font-size:11px;padding:4px 8px;background:#F59E0B;" onclick="openNewDraft('{t['branch_code']}', '{t['branch_name']}')">+ 새 임시저장</button>
                    </div>
                    <div id="draftList_{t['branch_code']}">{draft_items_html}</div>
                  </div>
                  <div>
                    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px;">
                      <h4 style="font-size:13px;">🔁 반복 메시지 목록</h4>
                      <button class="btn" style="font-size:11px;padding:4px 8px;background:#8B5CF6;" onclick="openNewSchedule('{t['branch_code']}', '{t['branch_name']}')">+ 새 예약</button>
                    </div>
                    <div id="schedList_{t['branch_code']}">{sched_items_html}</div>
                  </div>
                </div>
              </td>
            </tr>
            """

        rows_html += f"""
        <tr>
            <td style="text-align:center;">
              <input type="checkbox" class="tw-check" value="{t['branch_code']}" {'disabled' if not is_registered else ''} style="width:16px;height:16px;">
            </td>
            <td>{t['branch_name']}</td>
            <td style="font-size:12px;">{channel_display}</td>
            <td style="font-size:12px;color:#888;">{url_display}</td>
            <td>{draft_badge} {sched_badge}</td>
            <td style="display:flex;gap:4px;flex-wrap:wrap;">
              <button class="btn" style="font-size:11px;padding:4px 8px;" onclick="editWebhook('{t['branch_code']}', '{t['branch_name']}', '{safe_channel_name}')">등록/수정</button>
              {'<button class="btn" style="font-size:11px;padding:4px 8px;background:#64748B;" onclick="testSend(\'' + t['branch_code'] + '\', \'' + t['branch_name'] + '\')">테스트</button>' if has_webhook else ''}
              {expand_btn}
            </td>
        </tr>
        {detail_row}
        """

    branch_options_html = ""
    for t in targets:
        branch_options_html += f'<option value="{t["branch_code"]}">{t["branch_name"]}</option>'

    teams_alert_toggles_html = ""
    for alert_key, alert_info in TEAMS_ALERT_TYPES.items():
        conn2b = get_conn()
        setting_row = conn2b.execute(
            "SELECT value FROM system_settings WHERE key=?", (alert_info["toggle_key"],)
        ).fetchone()
        conn2b.close()
        is_enabled = (setting_row["value"] == "true") if setting_row else alert_info["default_enabled"]
        status_icon = "🟢" if is_enabled else "🔴"
        status_desc = alert_info["on_desc"] if is_enabled else alert_info["off_desc"]
        btn_label = "끄기" if is_enabled else "켜기"
        teams_alert_toggles_html += f"""
        <div style="display:flex;justify-content:space-between;align-items:center;padding:8px 0;border-bottom:1px solid #f0f0f0;">
          <div style="font-size:13px;">{alert_info['label']}: {status_icon} {status_desc}</div>
          <form method="post" action="{alert_info['toggle_route']}" style="margin:0;">
            <button type="submit" class="btn" style="font-size:11px;padding:5px 12px;">{btn_label}</button>
          </form>
        </div>
        """

    teams_alert_test_cards_html = ""
    for alert_key, alert_info in TEAMS_ALERT_TYPES.items():
        teams_alert_test_cards_html += f"""
        <div style="display:flex;justify-content:space-between;align-items:center;gap:8px;padding:8px 0;border-bottom:1px solid #f0f0f0;flex-wrap:wrap;">
          <div style="font-size:13px;min-width:140px;">{alert_info['label']}</div>
          <div style="display:flex;gap:6px;flex:1;min-width:220px;">
            <select id="testTarget_{alert_key}" style="flex:1;padding:6px;font-size:12px;">
              {branch_options_html}
            </select>
            <button class="btn" style="font-size:11px;padding:6px 10px;" onclick="{alert_info['test_js_func']}('{alert_key}')">발송</button>
          </div>
        </div>
        """

    content = f"""
    <h2 style="margin-bottom:16px;">🔔 Teams 웹훅 관리</h2>
    <div class="card" style="margin-bottom:16px;">
      <h3 style="margin-bottom:10px;font-size:14px;">📢 자동 발송 알림 설정</h3>
      {teams_alert_toggles_html}
    </div>
    <div class="card" style="margin-bottom:16px;">
      <h3 style="margin-bottom:8px;font-size:14px;">🧪 알림 테스트 발송</h3>
      <p style="font-size:12px;color:#888;margin-bottom:12px;">실제 지점에는 발송되지 않고, 선택한 채널로 현재 상황 미리보기만 전송됩니다.</p>
      {teams_alert_test_cards_html}
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">새 채널 추가</h3>
      <div style="display:flex;gap:8px;flex-wrap:wrap;max-width:600px;">
        <input type="text" id="newChannelCode" placeholder="채널 코드 (예: 운영기획팀, 공백없이)" style="flex:1;min-width:160px;">
        <input type="text" id="newChannelName" placeholder="채널 이름 (예: 운영기획팀 공지방)" style="flex:1;min-width:160px;">
        <button class="btn" type="button" onclick="addFreeChannel()">추가</button>
      </div>
      <div id="addChannelResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;">
        <button type="button" class="btn" id="twSelectAllBtn" style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
        <div style="display:flex;gap:8px;">
          <button type="button" class="btn" id="twSendBtn" style="background:#2563eb;font-size:12px;padding:6px 12px;">✉️ 선택 채널에 메시지 보내기</button>
          <button type="button" class="btn btn-red" id="twDeleteBtn" style="font-size:12px;padding:6px 12px;">선택 채널 삭제</button>
        </div>
      </div>
      <table>
        <thead><tr>
          <th style="width:36px;text-align:center;"><input type="checkbox" id="twAllCheck" style="width:16px;height:16px;"></th>
          <th>대상</th><th>채널 이름</th><th>등록된 웹훅</th><th>설정</th><th></th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>

    <div id="twMessageModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:480px;width:90%;">
        <h3 style="margin-bottom:12px;">메시지 보내기</h3>
        <p id="twSelectedList" style="font-size:12px;color:#888;margin-bottom:12px;"></p>
        <input type="text" id="twMsgTitle" placeholder="제목" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <textarea id="twMsgBody" placeholder="메시지 내용을 입력하세요" style="width:100%;min-height:100px;padding:10px;border:1px solid #ccc;border-radius:6px;box-sizing:border-box;font-size:14px;"></textarea>
        <input type="text" id="twMsgLinkUrl" placeholder="링크 URL (선택)" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-top:8px;box-sizing:border-box;">
        <input type="text" id="twMsgLinkText" placeholder="링크 버튼 텍스트 (선택)" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-top:8px;box-sizing:border-box;">
        <div style="display:flex;gap:8px;margin-top:16px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeMessageModal()">취소</button>
          <button class="btn" style="flex:1;" onclick="sendMessage()">보내기</button>
        </div>
      </div>
    </div>

    <div id="branchScheduleModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:480px;width:90%;max-height:85vh;overflow-y:auto;">
        <h3 id="schBranchTitle" style="margin-bottom:12px;">반복 메시지 예약</h3>
        <input type="text" id="schTitle" placeholder="제목" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <textarea id="schMessage" placeholder="메시지 내용" style="width:100%;min-height:80px;padding:10px;border:1px solid #ccc;border-radius:6px;box-sizing:border-box;font-size:14px;margin-bottom:8px;"></textarea>
        <label style="font-size:12px;color:#888;">시작일</label>
        <input type="date" id="schStartDate" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <label style="font-size:12px;color:#888;">반복 간격 (일)</label>
        <input type="number" id="schInterval" value="5" min="1" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:12px;box-sizing:border-box;">
        <input type="hidden" id="schBranchCode">
        <input type="hidden" id="schEditId">
        <div style="display:flex;gap:8px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeBranchSchedule()">취소</button>
          <button class="btn" style="flex:1;" id="schSaveBtn" onclick="saveBranchSchedule()">저장</button>
        </div>
      </div>
    </div>

    <div id="draftModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:480px;width:90%;">
        <h3 id="draftTitle" style="margin-bottom:12px;">임시 저장 메시지</h3>
        <input type="text" id="draftMsgTitle" placeholder="제목" style="width:100%;padding:10px;border:1px solid #ccc;border-radius:6px;margin-bottom:8px;box-sizing:border-box;">
        <textarea id="draftMsgBody" placeholder="메시지 내용" style="width:100%;min-height:100px;padding:10px;border:1px solid #ccc;border-radius:6px;box-sizing:border-box;font-size:14px;"></textarea>
        <input type="hidden" id="draftBranchCode">
        <input type="hidden" id="draftEditId">
        <div style="display:flex;gap:8px;margin-top:16px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeDraft()">취소</button>
          <button class="btn" style="flex:1;background:#F59E0B;" onclick="saveDraft()">임시저장</button>
          <button class="btn" style="flex:1;" onclick="sendFromDraft()">바로 발송</button>
        </div>
      </div>
    </div>

    <script>
      (function() {{
        var allCheck = document.getElementById('twAllCheck');
        var selectBtn = document.getElementById('twSelectAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.tw-check:not([disabled])').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{ allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }}); }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        document.getElementById('twSendBtn').addEventListener('click', function() {{
          var checked = Array.from(document.querySelectorAll('.tw-check:checked')).map(c => c.value);
          if (checked.length === 0) {{ alert('메시지를 보낼 대상을 선택하세요.'); return; }}
          document.getElementById('twSelectedList').innerText = checked.length + '개 채널에 발송됩니다.';
          document.getElementById('twMessageModal').style.display = 'flex';
          document.getElementById('twMessageModal').dataset.targets = JSON.stringify(checked);
        }});
        document.getElementById('twDeleteBtn').addEventListener('click', async function() {{
          var checked = Array.from(document.querySelectorAll('.tw-check:checked')).map(c => c.value);
          if (checked.length === 0) {{ alert('삭제할 채널을 선택하세요.'); return; }}
          if (!confirm(checked.length + '개 채널의 웹훅 등록을 삭제합니다. 계속할까요?')) return;
          for (const code of checked) {{
            await fetch('/master/teams-webhook/save', {{
              method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
              body: JSON.stringify({{ branch_code: code, webhook_url: '', channel_name: '', delete: true }})
            }});
          }}
          location.reload();
        }});
      }})();

      function toggleAccordion(branchCode) {{
        var row = document.getElementById('accordion_' + branchCode);
        if (!row) return;
        row.style.display = (row.style.display === 'none') ? 'table-row' : 'none';
      }}

      function closeMessageModal() {{
        document.getElementById('twMessageModal').style.display = 'none';
        document.getElementById('twMsgTitle').value = '';
        document.getElementById('twMsgBody').value = '';
        document.getElementById('twMsgLinkUrl').value = '';
        document.getElementById('twMsgLinkText').value = '';
      }}

      async function sendMessage() {{
        const targets = JSON.parse(document.getElementById('twMessageModal').dataset.targets || '[]');
        const title = document.getElementById('twMsgTitle').value.trim();
        const body = document.getElementById('twMsgBody').value.trim();
        const linkUrl = document.getElementById('twMsgLinkUrl').value.trim();
        const linkText = document.getElementById('twMsgLinkText').value.trim();
        if (!body) {{ alert('메시지 내용을 입력하세요.'); return; }}
        const res = await fetch('/master/teams-webhook/broadcast', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_codes: targets, title: title, message: body, link_url: linkUrl, link_text: linkText }})
        }});
        const result = await res.json();
        if (res.ok) {{
          alert('발송 완료: 성공 ' + result.success + '건, 실패 ' + result.failed + '건');
          closeMessageModal();
        }} else {{ alert('오류: ' + (result.detail || '발송 실패')); }}
      }}

      async function testSend(branchCode, branchName) {{
        if (!confirm(branchName + ' 채널로 테스트 메시지를 보낼까요?')) return;
        const res = await fetch('/master/teams-webhook/broadcast', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_codes: [branchCode], title: '테스트 알림', message: '이 메시지가 보이면 웹훅이 정상 연결된 것입니다.', link_url: 'https://inventory-sync-teal.vercel.app', link_text: '앱 바로가기' }})
        }});
        const result = await res.json();
        if (res.ok && result.success > 0) {{
          alert('테스트 발송 성공! Teams 채널을 확인하세요.');
        }} else {{
          alert('발송 실패:\\n' + (result.fail_details ? result.fail_details.join('\\n') : '알 수 없는 오류'));
        }}
      }}

      async function testUnsubmittedReminder(alertKey) {{
        const target = document.getElementById('testTarget_' + alertKey).value;
        if (!target) {{ alert('테스트 채널을 선택하세요.'); return; }}
        const res = await fetch('/master/teams-webhook/test-unsubmitted-reminder', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: target }})
        }});
        const result = await res.json();
        if (res.ok) {{
          alert('테스트 발송 완료! 미제출 지점 ' + result.unsubmitted_count + '곳 (기준월: ' + result.month + ')\\nTeams 채널을 확인하세요.');
        }} else {{
          alert('오류: ' + (result.detail || '발송 실패'));
        }}
      }}
      async function testQrRawMismatch(alertKey) {{
        const target = document.getElementById('testTarget_' + alertKey).value;
        if (!target) {{ alert('테스트 채널을 선택하세요.'); return; }}
        const res = await fetch('/master/teams-webhook/test-qr-raw-mismatch', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: target }})
        }});
        const result = await res.json();
        if (res.ok) {{
          alert('테스트 발송 완료! 불일치 있는 지점 ' + result.mismatch_branch_count + '곳\\nTeams 채널을 확인하세요.');
        }} else {{
          alert('오류: ' + (result.detail || '발송 실패'));
        }}
      }}
      async function testPurchaseNew(alertKey) {{
        const target = document.getElementById('testTarget_' + alertKey).value;
        if (!target) {{ alert('테스트 채널을 선택하세요.'); return; }}
        const res = await fetch('/master/teams-webhook/test-purchase-new', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: target }})
        }});
        const result = await res.json();
        if (res.ok) {{
          alert('테스트 발송 완료! 최근 발주 ' + result.purchase_count + '건 미리보기 발송\\nTeams 채널을 확인하세요.');
        }} else {{
          alert('오류: ' + (result.detail || '발송 실패'));
        }}
      }}

      async function editWebhook(branchCode, branchName, currentChannelName) {{
        const url = prompt(branchName + '의 Teams 웹훅 URL을 입력하세요:');
        if (url === null) return;
        if (url.trim() && !url.trim().startsWith('https://')) {{ alert('올바른 URL 형식이 아닙니다.'); return; }}
        let channelName = currentChannelName;
        if (url.trim()) {{
          const nameInput = prompt('채널 이름을 입력하세요:', currentChannelName);
          if (nameInput === null) return;
          channelName = nameInput.trim();
        }}
        const res = await fetch('/master/teams-webhook/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: branchCode, webhook_url: url.trim(), channel_name: channelName }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('저장 실패'); }}
      }}

      async function addFreeChannel() {{
        const code = document.getElementById('newChannelCode').value.trim();
        const name = document.getElementById('newChannelName').value.trim();
        if (!code) {{ alert('채널 코드를 입력하세요.'); return; }}
        if (code.includes(' ')) {{ alert('채널 코드에는 공백을 사용할 수 없습니다.'); return; }}
        const res = await fetch('/master/teams-webhook/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: code, webhook_url: '', channel_name: name }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('addChannelResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}

      // ── 반복 메시지: 신규/수정 공용 모달 ──
      function openNewSchedule(branchCode, branchName) {{
        document.getElementById('schBranchTitle').innerText = branchName + ' — 새 반복 메시지';
        document.getElementById('schBranchCode').value = branchCode;
        document.getElementById('schEditId').value = '';
        document.getElementById('schStartDate').value = new Date().toISOString().slice(0, 10);
        document.getElementById('schTitle').value = '';
        document.getElementById('schMessage').value = '';
        document.getElementById('schInterval').value = '5';
        document.getElementById('schSaveBtn').innerText = '저장';
        document.getElementById('branchScheduleModal').style.display = 'flex';
      }}

      function editSchedItem(scheduleId, branchCode) {{
        document.getElementById('schBranchTitle').innerText = '반복 메시지 수정';
        document.getElementById('schBranchCode').value = branchCode;
        document.getElementById('schEditId').value = scheduleId;
        document.getElementById('schSaveBtn').innerText = '수정 저장';
        document.getElementById('branchScheduleModal').style.display = 'flex';
        fetch('/master/teams-webhook/schedule/list?branch_code=' + encodeURIComponent(branchCode))
          .then(r => r.json())
          .then(data => {{
            const target = (data.schedules || []).find(s => s.id === scheduleId);
            if (!target) return;
            document.getElementById('schTitle').value = target.title || '';
            document.getElementById('schMessage').value = target.message || '';
            document.getElementById('schStartDate').value = target.start_date || '';
            document.getElementById('schInterval').value = target.interval_days || 5;
          }});
      }}

      function closeBranchSchedule() {{
        document.getElementById('branchScheduleModal').style.display = 'none';
      }}

      async function saveBranchSchedule() {{
        const branchCode = document.getElementById('schBranchCode').value;
        const editId = document.getElementById('schEditId').value;
        const title = document.getElementById('schTitle').value.trim();
        const message = document.getElementById('schMessage').value.trim();
        const startDate = document.getElementById('schStartDate').value;
        const interval = parseInt(document.getElementById('schInterval').value);
        if (!message || !startDate || !interval) {{ alert('메시지, 시작일, 간격을 입력하세요.'); return; }}

        let res;
        if (editId) {{
          res = await fetch('/master/teams-webhook/schedule/edit', {{
            method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{ schedule_id: parseInt(editId), title: title, message: message, start_date: startDate, interval_days: interval }})
          }});
        }} else {{
          res = await fetch('/master/teams-webhook/schedule/add', {{
            method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{ branch_code: branchCode, title: title, message: message, start_date: startDate, interval_days: interval }})
          }});
        }}
        if (res.ok) {{ location.reload(); }} else {{ alert('저장 실패'); }}
      }}

      async function toggleSchedItem(id, active, branchCode) {{
        await fetch('/master/teams-webhook/schedule/toggle', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ schedule_id: id, active: active }})
        }});
        location.reload();
      }}

      async function deleteSchedItem(id, branchCode) {{
        if (!confirm('삭제할까요?')) return;
        await fetch('/master/teams-webhook/schedule/delete', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ schedule_id: id }})
        }});
        location.reload();
      }}

      // ── 임시저장: 신규/수정 공용 모달 ──
      function openNewDraft(branchCode, branchName) {{
        document.getElementById('draftTitle').innerText = branchName + ' — 새 임시저장';
        document.getElementById('draftBranchCode').value = branchCode;
        document.getElementById('draftEditId').value = '';
        document.getElementById('draftMsgTitle').value = '';
        document.getElementById('draftMsgBody').value = '';
        document.getElementById('draftModal').style.display = 'flex';
      }}

      function editDraftItem(draftId, branchCode) {{
        document.getElementById('draftTitle').innerText = '임시저장 수정';
        document.getElementById('draftBranchCode').value = branchCode;
        document.getElementById('draftEditId').value = draftId;
        fetch('/master/teams-webhook/draft/list?branch_code=' + encodeURIComponent(branchCode))
          .then(r => r.json())
          .then(data => {{
            const target = (data.drafts || []).find(d => d.id === draftId);
            if (!target) return;
            document.getElementById('draftMsgTitle').value = target.title || '';
            document.getElementById('draftMsgBody').value = target.message || '';
          }});
        document.getElementById('draftModal').style.display = 'flex';
      }}

      function closeDraft() {{
        document.getElementById('draftModal').style.display = 'none';
      }}

      async function saveDraft() {{
        const branchCode = document.getElementById('draftBranchCode').value;
        const editId = document.getElementById('draftEditId').value;
        const title = document.getElementById('draftMsgTitle').value.trim();
        const message = document.getElementById('draftMsgBody').value.trim();
        const res = await fetch('/master/teams-webhook/draft/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_code: branchCode, draft_id: editId ? parseInt(editId) : null, title: title, message: message }})
        }});
        if (res.ok) {{ alert('임시저장 완료'); location.reload(); }} else {{ alert('저장 실패'); }}
      }}

      async function sendFromDraft() {{
        const branchCode = document.getElementById('draftBranchCode').value;
        const editId = document.getElementById('draftEditId').value;
        const title = document.getElementById('draftMsgTitle').value.trim();
        const message = document.getElementById('draftMsgBody').value.trim();
        if (!message) {{ alert('메시지 내용을 입력하세요.'); return; }}
        const res = await fetch('/master/teams-webhook/broadcast', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_codes: [branchCode], title: title, message: message }})
        }});
        const result = await res.json();
        if (res.ok && result.success > 0) {{
          if (editId) {{
            await fetch('/master/teams-webhook/draft/delete', {{
              method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
              body: JSON.stringify({{ draft_id: parseInt(editId) }})
            }});
          }}
          alert('발송 완료');
          closeDraft();
          location.reload();
        }} else {{ alert('발송 실패'); }}
      }}

      async function sendDraftItem(draftId, branchCode) {{
        if (!confirm('이 임시저장 메시지를 바로 발송할까요?')) return;
        const listRes = await fetch('/master/teams-webhook/draft/list?branch_code=' + encodeURIComponent(branchCode));
        const listData = await listRes.json();
        const target = (listData.drafts || []).find(d => d.id === draftId);
        if (!target) {{ alert('메시지를 찾을 수 없습니다.'); return; }}
        const res = await fetch('/master/teams-webhook/broadcast', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_codes: [branchCode], title: target.title, message: target.message }})
        }});
        const result = await res.json();
        if (res.ok && result.success > 0) {{
          await fetch('/master/teams-webhook/draft/delete', {{
            method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
            body: JSON.stringify({{ draft_id: draftId }})
          }});
          alert('발송 완료');
          location.reload();
        }} else {{ alert('발송 실패'); }}
      }}

      async function deleteDraftItem(draftId, branchCode) {{
        if (!confirm('이 임시저장 메시지를 삭제할까요?')) return;
        await fetch('/master/teams-webhook/draft/delete', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ draft_id: draftId }})
        }});
        location.reload();
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/teams-webhook/save")
async def master_teams_webhook_save(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    webhook_url = data.get("webhook_url", "").strip()
    channel_name = data.get("channel_name", "").strip()

    if not branch_code:
        return JSONResponse(status_code=400, content={"detail": "대상이 지정되지 않았습니다."})

    delete_requested = data.get("delete", False)

    conn = get_conn()
    existing = conn.execute("SELECT id FROM teams_webhook WHERE branch_code=?", (branch_code,)).fetchone()
    if existing:
        if delete_requested:
            conn.execute("DELETE FROM teams_webhook WHERE branch_code=?", (branch_code,))
        else:
            conn.execute(
                "UPDATE teams_webhook SET webhook_url=?, channel_name=?, updated_at=NOW() WHERE branch_code=?",
                (webhook_url, channel_name, branch_code)
            )
    else:
        conn.execute(
            "INSERT INTO teams_webhook (branch_code, webhook_url, channel_name) VALUES (?, ?, ?)",
            (branch_code, webhook_url, channel_name)
        )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/teams-webhook/broadcast")
async def master_teams_webhook_broadcast(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    branch_codes = data.get("branch_codes", [])
    title = data.get("title", "").strip() or "알림"
    message = data.get("message", "").strip()
    link_url = data.get("link_url", "").strip()
    link_text = data.get("link_text", "").strip()

    if not branch_codes or not message:
        return JSONResponse(status_code=400, content={"detail": "대상 또는 메시지가 비어있습니다."})

    device_info = request.headers.get("user-agent", "")[:200]

    success_count = 0
    fail_count = 0
    fail_details = []
    for bc in branch_codes:
        ok, detail = send_teams_notification(bc, title, message, link_url, link_text, user["branch_code"], device_info)
        if ok:
            success_count += 1
        else:
            fail_count += 1
            fail_details.append(f"{bc}: {detail}")

    return JSONResponse(content={
        "status": "ok",
        "success": success_count,
        "failed": fail_count,
        "fail_details": fail_details
    })


@app.post("/master/teams-webhook/schedule/add")
async def master_branch_schedule_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    title = data.get("title", "").strip()
    message = data.get("message", "").strip()
    start_date = data.get("start_date", "").strip()
    interval_days = data.get("interval_days")

    if not branch_code or not message or not start_date or not interval_days:
        return JSONResponse(status_code=400, content={"detail": "모든 필드를 입력하세요."})

    conn = get_conn()
    conn.execute(
        "INSERT INTO teams_scheduled_message (title, message, target_type, target_codes, start_date, interval_days, created_by) VALUES (?, ?, 'branch', ?, ?, ?, ?)",
        (title, message, branch_code, start_date, interval_days, user["branch_code"])
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.get("/master/teams-webhook/schedule/list")
async def master_branch_schedule_list(branch_code: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM teams_scheduled_message WHERE target_type='branch' AND target_codes=?", (branch_code,)
    ).fetchall()
    conn.close()
    return JSONResponse(content={"schedules": [dict(r) for r in rows]})


@app.post("/master/teams-webhook/schedule/toggle")
async def master_branch_schedule_toggle(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    conn = get_conn()
    conn.execute("UPDATE teams_scheduled_message SET active=? WHERE id=?", (data.get("active"), data.get("schedule_id")))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})

@app.post("/master/teams-webhook/schedule/edit")
async def master_branch_schedule_edit(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    schedule_id = data.get("schedule_id")
    title = data.get("title", "").strip()
    message = data.get("message", "").strip()
    start_date = data.get("start_date", "").strip()
    interval_days = data.get("interval_days")

    if not schedule_id or not message or not start_date or not interval_days:
        return JSONResponse(status_code=400, content={"detail": "모든 필드를 입력하세요."})

    conn = get_conn()
    conn.execute(
        "UPDATE teams_scheduled_message SET title=?, message=?, start_date=?, interval_days=? WHERE id=?",
        (title, message, start_date, interval_days, schedule_id)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/teams-webhook/schedule/delete")
async def master_branch_schedule_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    conn = get_conn()
    conn.execute("DELETE FROM teams_scheduled_message WHERE id=?", (data.get("schedule_id"),))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.get("/master/teams-webhook/draft/list")
async def master_draft_list(branch_code: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM teams_draft_message WHERE branch_code=? ORDER BY updated_at DESC",
        (branch_code,)
    ).fetchall()
    conn.close()
    return JSONResponse(content={"drafts": [dict(r) for r in rows]})


@app.post("/master/teams-webhook/draft/save")
async def master_draft_save(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    draft_id = data.get("draft_id")
    title = data.get("title", "").strip()
    message = data.get("message", "").strip()
    if not branch_code:
        return JSONResponse(status_code=400, content={"detail": "대상이 지정되지 않았습니다."})
    conn = get_conn()
    if draft_id:
        conn.execute(
            "UPDATE teams_draft_message SET title=?, message=?, updated_at=NOW() WHERE id=? AND branch_code=?",
            (title, message, draft_id, branch_code)
        )
    else:
        conn.execute(
            "INSERT INTO teams_draft_message (branch_code, title, message) VALUES (?, ?, ?)",
            (branch_code, title, message)
        )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/teams-webhook/draft/delete")
async def master_draft_delete(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    draft_id = data.get("draft_id")
    if not draft_id:
        return JSONResponse(status_code=400, content={"detail": "삭제할 항목이 지정되지 않았습니다."})
    conn = get_conn()
    conn.execute("DELETE FROM teams_draft_message WHERE id=?", (draft_id,))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.get("/master/eval-criteria", response_class=HTMLResponse)
async def eval_criteria_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    criteria_list = conn.execute("SELECT * FROM eval_criteria ORDER BY display_order").fetchall()
    criteria_html = ""
    for c in criteria_list:
        options = conn.execute(
            "SELECT * FROM eval_criteria_option WHERE criteria_id=? ORDER BY score", (c["id"],)
        ).fetchall()
        opt_rows = ""
        for o in options:
            comment_badge = '<span class="badge-red">사유필수</span>' if o["requires_comment"] else '<span class="badge-green">사유없음</span>'
            safe_label = o['label'].replace("'", "")
            safe_desc = (o['description'] or '').replace("'", "")
            opt_rows += f"""
            <tr>
              <td>{o['score']}</td><td>{o['label']}</td><td style="font-size:12px;color:#888;">{o['description'] or '-'}</td>
              <td>{comment_badge}</td>
              <td>
                <button class="btn" style="font-size:11px;padding:4px 8px;" onclick="editOption({o['id']}, '{safe_label}', '{safe_desc}', {str(o['requires_comment']).lower()})">수정</button>
                <button class="btn btn-red" style="font-size:11px;padding:4px 8px;" onclick="deleteOption({o['id']})">삭제</button>
              </td>
            </tr>
            """
        active_badge = '<span class="badge-green">사용중</span>' if c["active"] else '<span class="badge-red">비활성</span>'
        safe_c_label = c['label'].replace("'", "")
        safe_c_desc = (c['description'] or '').replace("'", "")
        criteria_html += f"""
        <div class="card">
          <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;">
            <div>
              <b>{c['display_order']}. {c['label']}</b> {active_badge}
              <span style="color:#888;font-size:12px;">(최대 {c['max_score']}단계)</span>
            </div>
            <div style="display:flex;gap:6px;">
              <button class="btn" style="font-size:12px;padding:6px 10px;" onclick="editCriteriaInfo({c['id']}, '{safe_c_label}', '{safe_c_desc}')">문항 수정</button>
              <button class="btn" style="font-size:12px;padding:6px 10px;" onclick="toggleActive({c['id']}, {str(not c['active']).lower()})">{'비활성화' if c['active'] else '활성화'}</button>
              <button class="btn" style="font-size:12px;padding:6px 10px;" onclick="addOption({c['id']})">단계 추가</button>
              <button class="btn btn-red" style="font-size:12px;padding:6px 10px;" onclick="deleteCriteria({c['id']})">문항 삭제</button>
            </div>
          </div>
          <p style="color:#888;font-size:12px;margin-bottom:8px;">{c['description'] or '(설명 없음 — "문항 수정" 버튼으로 추가 가능)'}</p>
          <table>
            <thead><tr><th>점수</th><th>라벨</th><th>설명</th><th>사유</th><th></th></tr></thead>
            <tbody>{opt_rows}</tbody>
          </table>
        </div>
        """

    content = f"""
    <h2 style="margin-bottom:16px;">📝 거래처평가 문항 관리</h2>
    <div class="card" style="background:#FFF7ED;border:1px solid #FCD34D;">
      <p style="font-size:13px;color:#92400E;">⚠️ 여기서 만든 문항은 새로 작성되는 평가부터 적용됩니다. 기존 평가 기록(레거시)은 영향받지 않습니다.</p>
    </div>
    <div class="card">
      <h3 style="margin-bottom:8px;">새 문항 추가</h3>
      <div style="display:flex;gap:8px;">
        <input type="text" id="newCriteriaLabel" placeholder="문항명 (예: 배송 포장 상태)" style="flex:1;">
        <button class="btn" type="button" onclick="addCriteria()">추가</button>
      </div>
      <div id="addCriteriaResult" style="margin-top:8px;font-size:13px;"></div>
    </div>
    {criteria_html}

    <script>
      async function addCriteria() {{
        const input = document.getElementById('newCriteriaLabel');
        const label = input.value.trim();
        if (!label) {{ alert('문항명을 입력하세요.'); return; }}
        const res = await fetch('/master/eval-criteria/add', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ label: label }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('addCriteriaResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}

      async function editCriteriaInfo(id, currentLabel, currentDesc) {{
        const label = prompt('문항명 수정:', currentLabel);
        if (label === null || !label.trim()) return;
        const description = prompt('문항 설명 수정 (전체 안내문, 선택 입력):', currentDesc);
        if (description === null) return;
        const res = await fetch('/master/eval-criteria/' + id + '/info', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ label: label.trim(), description: description.trim() }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert('오류: ' + (err.detail || '수정 실패'));
        }}
      }}

      async function deleteCriteria(id) {{
        if (!confirm('이 문항과 모든 단계 옵션을 삭제합니다. 계속할까요?')) return;
        const res = await fetch('/master/eval-criteria/' + id + '/delete', {{ method: 'POST' }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert(err.detail || '삭제 실패');
        }}
      }}

      async function toggleActive(id, newActive) {{
        const res = await fetch('/master/eval-criteria/' + id + '/toggle', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ active: newActive }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('변경 실패'); }}
      }}

      async function addOption(criteriaId) {{
        const label = prompt('새 단계의 라벨을 입력하세요 (예: 매우 우수)');
        if (!label) return;
        const requiresComment = confirm('이 단계 선택 시 사유 입력을 필수로 할까요? (확인=필수, 취소=선택)');
        const res = await fetch('/master/eval-criteria/' + criteriaId + '/option/add', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ label: label, description: '', requires_comment: requiresComment }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert('오류: ' + (err.detail || '추가 실패'));
        }}
      }}

      async function editOption(optionId, currentLabel, currentDesc, currentRequires) {{
        const label = prompt('라벨 수정:', currentLabel);
        if (label === null) return;
        const description = prompt('설명 수정 (선택):', currentDesc);
        if (description === null) return;
        const requiresComment = confirm('이 단계 선택 시 사유 입력을 필수로 할까요? (확인=필수, 취소=선택)');
        const res = await fetch('/master/eval-criteria/option/' + optionId + '/edit', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ label: label, description: description || '', requires_comment: requiresComment }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('수정 실패'); }}
      }}

      async function deleteOption(optionId) {{
        if (!confirm('이 단계를 삭제합니다. 계속할까요?')) return;
        const res = await fetch('/master/eval-criteria/option/' + optionId + '/delete', {{ method: 'POST' }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          alert(err.detail || '삭제 실패');
        }}
      }}


    </script>
    """
    conn.close()
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/eval-criteria/add")
async def eval_criteria_add(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    label = data.get("label", "").strip()
    if not label:
        return JSONResponse(status_code=400, content={"detail": "문항명을 입력하세요."})

    import re
    criteria_key = re.sub(r'[^a-z0-9_]', '', label.lower().replace(' ', '_'))[:40] or f"criteria_{int(datetime.now().timestamp())}"

    conn = get_conn()
    max_order = conn.execute("SELECT COALESCE(MAX(display_order), 0) as m FROM eval_criteria").fetchone()["m"]
    conn.execute(
        "INSERT INTO eval_criteria (criteria_key, label, display_order, active, max_score) VALUES (?, ?, ?, TRUE, 5)",
        (f"{criteria_key}_{int(datetime.now().timestamp())}", label, max_order + 1)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/eval-criteria/{criteria_id}/info")
async def eval_criteria_update_info(criteria_id: int, request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    label = data.get("label", "").strip()
    description = data.get("description", "").strip()
    if not label:
        return JSONResponse(status_code=400, content={"detail": "문항명을 입력하세요."})

    conn = get_conn()
    conn.execute(
        "UPDATE eval_criteria SET label=?, description=? WHERE id=?",
        (label, description, criteria_id)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/eval-criteria/{criteria_id}/delete")
async def eval_criteria_delete(criteria_id: int, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    conn = get_conn()
    usage_count = conn.execute(
        "SELECT COUNT(*) as cnt FROM vendor_evaluation_answer WHERE criteria_id=?", (criteria_id,)
    ).fetchone()["cnt"]

    if usage_count > 0:
        conn.close()
        return JSONResponse(status_code=400, content={
            "detail": f"이미 {usage_count}건의 평가에 사용된 문항이라 삭제할 수 없습니다. 대신 '비활성화'를 사용해주세요."
        })

    conn.execute("DELETE FROM eval_criteria WHERE id=?", (criteria_id,))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/eval-criteria/{criteria_id}/toggle")
async def eval_criteria_toggle(criteria_id: int, request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    active = data.get("active", True)
    conn = get_conn()
    conn.execute("UPDATE eval_criteria SET active=? WHERE id=?", (active, criteria_id))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/eval-criteria/{criteria_id}/option/add")
async def eval_criteria_option_add(criteria_id: int, request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    label = data.get("label", "").strip()
    description = data.get("description", "").strip()
    requires_comment = data.get("requires_comment", False)
    if not label:
        return JSONResponse(status_code=400, content={"detail": "라벨을 입력하세요."})

    conn = get_conn()
    max_score = conn.execute(
        "SELECT COALESCE(MAX(score), 0) as m FROM eval_criteria_option WHERE criteria_id=?", (criteria_id,)
    ).fetchone()["m"]
    new_score = max_score + 1
    conn.execute(
        "INSERT INTO eval_criteria_option (criteria_id, score, label, description, requires_comment) VALUES (?, ?, ?, ?, ?)",
        (criteria_id, new_score, label, description, requires_comment)
    )
    conn.execute("UPDATE eval_criteria SET max_score=? WHERE id=?", (new_score, criteria_id))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})

@app.post("/master/eval-criteria/option/{option_id}/delete")
async def eval_criteria_option_delete(option_id: int, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    conn = get_conn()
    usage_count = conn.execute(
        "SELECT COUNT(*) as cnt FROM vendor_evaluation_answer WHERE score IN "
        "(SELECT score FROM eval_criteria_option WHERE id=?) AND criteria_id IN "
        "(SELECT criteria_id FROM eval_criteria_option WHERE id=?)",
        (option_id, option_id)
    ).fetchone()["cnt"]

    if usage_count > 0:
        conn.close()
        return JSONResponse(status_code=400, content={
            "detail": f"이미 {usage_count}건의 평가에 사용된 단계라 삭제할 수 없습니다."
        })

    conn.execute("DELETE FROM eval_criteria_option WHERE id=?", (option_id,))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/eval-criteria/option/{option_id}/edit")
async def eval_criteria_option_edit(option_id: int, request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    label = data.get("label", "").strip()
    description = data.get("description", "").strip()
    requires_comment = data.get("requires_comment", False)
    if not label:
        return JSONResponse(status_code=400, content={"detail": "라벨을 입력하세요."})

    conn = get_conn()
    conn.execute(
        "UPDATE eval_criteria_option SET label=?, description=?, requires_comment=? WHERE id=?",
        (label, description, requires_comment, option_id)
    )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})

@app.post("/scan-log/delete")
async def scan_log_delete(
    request: Request,
    session_token: str = Cookie(default=None)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("log_ids")
    if ids:
        conn = get_conn()
        logs_to_delete = conn.execute(
            f"SELECT * FROM scan_log WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        ).fetchall()
        for lg in logs_to_delete:
            revert_delta = -1 if lg["scan_type"] == "IN" else 1
            adjust_quantity(lg["branch_code"], lg["item_code"], revert_delta)
        conn.execute(
            f"DELETE FROM scan_log WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/scan-log", status_code=303)


@app.post("/scan-log/delete-all")
async def scan_log_delete_all(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/scan-log", status_code=303)
    conn = get_conn()
    if user["role"] == "master":
        query = "SELECT * FROM scan_log"
        params = []
    logs_to_delete = conn.execute(query, params).fetchall()
    for lg in logs_to_delete:
        revert_delta = -1 if lg["scan_type"] == "IN" else 1
        adjust_quantity(lg["branch_code"], lg["item_code"], revert_delta)
    conn.execute("DELETE FROM scan_log")
    conn.commit()
    conn.close()
    return RedirectResponse(url="/scan-log", status_code=303)

# ── 헬스체크 ────────────────────────────────────────────

# ── PWA ──────────────────────────────────────────────

@app.get("/manifest.json")
async def pwa_manifest():
    from fastapi.responses import JSONResponse
    return JSONResponse({
        "name": "재고 관리 시스템",
        "short_name": "재고관리",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f5f7fa",
        "theme_color": "#1E2761",
        "icons": [
            {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png"}
        ]
    })


@app.get("/sw.js")
async def pwa_service_worker():
    from fastapi.responses import Response
    js = """
const CACHE_NAME = 'inventory-sync-v1';
self.addEventListener('install', function(event) {
  self.skipWaiting();
});
self.addEventListener('activate', function(event) {
  self.clients.claim();
});
self.addEventListener('fetch', function(event) {
  // 네트워크 우선, 실패 시 아무 것도 안 함 (재고 데이터는 항상 최신이어야 하므로 캐시 저장 안 함)
  event.respondWith(
    fetch(event.request).catch(function() {
      return new Response('오프라인 상태입니다. 네트워크 연결을 확인해주세요.', {
        status: 503,
        headers: { 'Content-Type': 'text/plain; charset=utf-8' }
      });
    })
  );
});
self.addEventListener('push', function(event) {
  var data = {};
  try { data = event.data ? event.data.json() : {}; } catch (e) {}
  var title = data.title || '재고 관리 시스템';
  var body = data.body || '새 알림이 있습니다.';
  event.waitUntil(
    self.registration.showNotification(title, {
      body: body,
      icon: '/icon-192.png',
      badge: '/icon-192.png',
      data: { url: data.url || '/' }
    })
  );
});
self.addEventListener('notificationclick', function(event) {
  event.notification.close();
  var url = (event.notification.data && event.notification.data.url) || '/';
  event.waitUntil(
    clients.openWindow(url)
  );
});
"""
    return Response(content=js, media_type="application/javascript")


@app.get("/icon-192.png")
async def pwa_icon_192():
    return _generate_app_icon(192)


@app.get("/icon-512.png")
async def pwa_icon_512():
    return _generate_app_icon(512)


def _generate_app_icon(size: int):
    from fastapi.responses import Response
    from PIL import Image, ImageDraw, ImageFont
    import io
    img = Image.new("RGB", (size, size), "#1E2761")
    draw = ImageDraw.Draw(img)
    text = "📦"
    try:
        font_path = os.path.join(os.path.dirname(__file__), "fonts", "NanumGothic-Bold.ttf")
        font = ImageFont.truetype(font_path, int(size * 0.3))
        label = "재고"
        bbox = draw.textbbox((0, 0), label, font=font)
        w, h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        draw.text(((size - w) / 2, (size - h) / 2 - bbox[1]), label, fill="white", font=font)
    except Exception:
        pass
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(content=buf.getvalue(), media_type="image/png")


@app.get("/health")
async def health_check():
    return {"status": "ok"}


# ── 마스터 전용 페이지 ──────────────────────────────────

@app.get("/master", response_class=HTMLResponse)
async def master_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    raw_count = conn.execute("SELECT COUNT(*) AS cnt FROM raw_inventory").fetchone()["cnt"]
    item_count = conn.execute("SELECT COUNT(*) AS cnt FROM items").fetchone()["cnt"]
    vendor_count = conn.execute("SELECT COUNT(*) AS cnt FROM vendor_master").fetchone()["cnt"]
    conn.close()

    content = f"""
    <h2 style="margin-bottom:16px;">⚙️ 마스터 관리</h2>
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;">
      <a href="/master/data" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📋</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">데이터 관리</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">품목 {item_count}개 등록됨</div>
        </div>
      </a>
      <a href="/master/qr-init" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🔄</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">QR 재고 업로드</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">엑셀로 초기 수량 업로드</div>
        </div>
      </a>
      <a href="/master/vendor-master" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🏢</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">거래처 관리</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">거래처 {vendor_count}개 등록됨</div>
        </div>
      </a>
      <a href="/master/eval-criteria" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📝</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">평가 문항 관리</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">문항 추가/삭제/수정</div>
        </div>
      </a>
      <a href="/master/vendor-eval/status" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📊</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">거래처평가 제출현황</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">지점별 제출/미제출 확인</div>
        </div>
      </a>
      <a href="/master/branch-manage" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🏬</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">지점 관리</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">지점 추가/삭제</div>
        </div>
      </a>
            <a href="/master/notification-settings" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">⏰</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">알림 설정</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">자동 알림 켜기/끄기</div>
        </div>
      </a>
      <a href="/master/webhook-send-log" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📨</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">웹훅 발송 이력</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">수동 발송 내역 조회</div>
        </div>
      </a>
      <a href="/master/login-history" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🔐</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">접속 이력</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">계정별 로그인 기록</div>
        </div>
      </a>
      <a href="/master/purchase-tracking" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📈</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">발주 주기 트래킹</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">구매 패턴 분석/추천</div>
        </div>
      </a>
      <a href="/master/purchase-order/product-settings" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">📝</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">상품 설정</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">리드타임/MOQ/소모품/예외</div>
        </div>
      </a>
      <a href="/master/purchase-order/branch-exceptions" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🏬</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">발주서 지점 예외</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">소모품 포함 지점 설정</div>
        </div>
      </a>
      <a href="/master/purchase-order/safety-stock" style="text-decoration:none;">
        <div class="card" style="text-align:center;padding:24px;cursor:pointer;">
          <div style="font-size:32px;">🎯</div>
          <div style="font-weight:bold;color:#1E2761;margin-top:8px;">안전재고 관리</div>
          <div style="color:#888;font-size:12px;margin-top:4px;">발주서 생성 기준값</div>
        </div>
      </a>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


# ── 마스터 > 데이터 관리 ────────────────────────────────

@app.get("/master/data", response_class=HTMLResponse)
async def master_data_page(
    session_token: str = Cookie(default=None),
    filter_branch: str = "",
    search: str = ""
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    query = "SELECT * FROM items WHERE 1=1"
    params: list = []
    if filter_branch:
        query += " AND branch_code=?"
        params.append(filter_branch)
    if search:
        query += " AND (item_name LIKE ? OR item_code LIKE ? OR branch_name LIKE ?)"
        params += [f"%{search}%", f"%{search}%", f"%{search}%"]
    query += " ORDER BY branch_code, item_name"
    items = conn.execute(query, params).fetchall()
    all_branches = conn.execute(
        "SELECT DISTINCT branch_code, branch_name FROM items ORDER BY branch_code"
    ).fetchall()
    conn.close()

    branch_options = '<option value="">전체 지점</option>'
    for b in all_branches:
        sel = "selected" if filter_branch == b["branch_code"] else ""
        branch_options += f'<option value="{b["branch_code"]}" {sel}>{b["branch_name"]}</option>'

    rows_html = ""
    if not items:
        rows_html = '<tr><td colspan="5" style="text-align:center;padding:20px;color:#888;">데이터 없음</td></tr>'
    else:
        for it in items:
            rows_html += f"""
            <tr>
              <td style="text-align:center;">
                <input type="checkbox" name="selected_ids" value="{it['id']}"
                       class="master-data-check" style="width:16px;height:16px;">
              </td>
              <td>{it['branch_name']}</td>
              <td>{it['item_name']}</td>
              <td>{it['item_code']}</td>
              <td>{it['created_at'][:10] if it['created_at'] else '-'}</td>
            </tr>"""

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>📋 데이터 관리</h2>
    </div>

    <div class="card">
      <h3 style="margin-bottom:12px;">➕ 수기 등록</h3>
      <form method="post" action="/master/data/add">
        <div style="display:flex;gap:8px;flex-wrap:wrap;">
          <div style="flex:1;min-width:130px;">
            <label style="font-size:12px;color:#888;">지점</label>
            <select name="branch_code" required style="margin-top:4px;">
              <option value="">선택</option>
              {''.join(f'<option value="{b["branch_code"]}">{b["branch_name"]}</option>' for b in get_branches())}
            </select>
          </div>
          <div style="flex:1;min-width:130px;">
            <label style="font-size:12px;color:#888;">상품명</label>
            <input name="item_name" required placeholder="상품명" style="margin-top:4px;">
          </div>
          <div style="flex:1;min-width:130px;">
            <label style="font-size:12px;color:#888;">품번</label>
            <input name="item_code" required placeholder="품번" style="margin-top:4px;">
          </div>
          <div style="flex:1;min-width:100px;">
            <label style="font-size:12px;color:#888;">초기 수량</label>
            <input name="init_quantity" type="number" value="0" style="margin-top:4px;">
          </div>
        </div>
        <button class="btn" type="submit" style="margin-top:12px;">등록</button>
      </form>
    </div>

    <div class="card">
      <h3 style="margin-bottom:12px;">📂 엑셀 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:8px;">컬럼: A=지점명 | B=상품명 | C=품번 (1행 헤더)</p>
      <form method="post" action="/master/data/upload" enctype="multipart/form-data"
            style="display:flex;gap:8px;align-items:center;">
        <input type="file" name="file" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="submit">업로드</button>
      </form>
    </div>

    <div class="card">
      <form method="get" action="/master/data"
            style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        <div style="flex:1;min-width:120px;">
          <label style="font-size:12px;color:#888;">지점 필터</label>
          <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
        </div>
        <div style="flex:2;min-width:160px;">
          <label style="font-size:12px;color:#888;">검색</label>
          <input name="search" value="{search}" placeholder="상품명/품번 검색"
                 style="margin-top:4px;">
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/master/data" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
    </div>

    <div class="card">
      <form method="post" action="/master/data/delete-selected" id="listForm">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
          <h3>품목 목록 ({len(items)}개)</h3>
          <div style="display:flex;gap:8px;">
            <button type="button" class="btn" id="masterDataSelectAllBtn"
                    style="background:#64748B;font-size:12px;padding:6px 12px;">전체선택</button>
            <button type="submit" class="btn btn-red"
                    style="font-size:12px;padding:6px 12px;"
                    onclick="return confirm('선택 항목을 삭제할까요?')">선택삭제</button>
            <button type="button" class="btn btn-red" id="masterDataDeleteAllBtn"
                    style="font-size:12px;padding:6px 12px;">전체삭제</button>
          </div>
        </div>
        <table>
          <thead><tr>
            <th style="width:40px;text-align:center;">
              <input type="checkbox" id="masterDataAllCheck" style="width:16px;height:16px;">
            </th>
            <th>지점명</th><th>상품명</th><th>품번</th><th>등록일</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>
      </form>
    </div>
    <form method="post" action="/master/data/delete-all" id="masterDataDeleteAllForm"></form>
    <script>
      (function() {{
        var allCheck = document.getElementById('masterDataAllCheck');
        var selectBtn = document.getElementById('masterDataSelectAllBtn');
        var deleteAllBtn = document.getElementById('masterDataDeleteAllBtn');
        function applyAll(checked) {{
          document.querySelectorAll('.master-data-check').forEach(function(c) {{ c.checked = checked; }});
          if (allCheck) allCheck.checked = checked;
        }}
        if (allCheck) {{
          allCheck.addEventListener('click', function() {{ applyAll(allCheck.checked); }});
        }}
        if (selectBtn) {{
          selectBtn.addEventListener('click', function() {{
            var next = !(allCheck && allCheck.checked);
            applyAll(next);
          }});
        }}
        if (deleteAllBtn) {{
          deleteAllBtn.addEventListener('click', function() {{
            if (confirm('전체 품목을 삭제합니다. 되돌릴 수 없습니다.')) {{
              document.getElementById('masterDataDeleteAllForm').submit();
            }}
          }});
        }}
      }})();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/data/add")
async def master_data_add(
    session_token: str = Cookie(default=None),
    branch_code: str = Form(...),
    item_name: str = Form(...),
    item_code: str = Form(...),
    init_quantity: int = Form(0)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    branch_name = next(
        (b["branch_name"] for b in get_branches() if b["branch_code"] == branch_code), branch_code)
    conn = get_conn()
    now = datetime.now().isoformat()
    try:
        conn.execute(
            """INSERT INTO items (branch_code, branch_name, item_name, item_code, created_at)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(branch_code, item_code) DO UPDATE SET item_name=excluded.item_name""",
            (branch_code, branch_name, item_name, item_code, now)
        )
        conn.execute(
            """INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated)
               VALUES (?, ?, ?, ?, ?)
               ON CONFLICT(branch_code, item_code) DO UPDATE SET
                 quantity=excluded.quantity, item_name=excluded.item_name, last_updated=excluded.last_updated""",
            (branch_code, item_name, item_code, init_quantity, now)
        )
        # ⚠️ RAW 재고에 0으로 자동 생성 (이미 있으면 건드리지 않음, source='master')
        existing_raw = conn.execute(
            "SELECT id FROM raw_inventory WHERE branch_code=? AND item_code=? AND source='master'",
            (branch_code, item_code)
        ).fetchone()
        if not existing_raw:
            conn.execute(
                """INSERT INTO raw_inventory
                   (branch_code, branch_name, item_name, item_code, quantity, source, uploaded_at)
                   VALUES (?, ?, ?, ?, 0, 'master', ?)""",
                (branch_code, branch_name, item_name, item_code, now)
            )
        conn.commit()
    except Exception:
        pass
    conn.close()
    return RedirectResponse(url="/master/data", status_code=303)


@app.post("/master/data/upload")
async def master_data_upload(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return RedirectResponse(url="/master/data", status_code=303)
    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]
    conn = get_conn()
    now = datetime.now().isoformat()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        branch_name = str(row[0]).strip()
        item_name = str(row[1]).strip() if row[1] else ""
        item_code = str(row[2]).strip() if row[2] else ""
        branch_code = (branch_map.get(branch_name)
                       or branch_map.get(branch_name.replace(" ", ""))
                       or branch_name)
        try:
            conn.execute(
                """INSERT INTO items (branch_code, branch_name, item_name, item_code, created_at)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(branch_code, item_code) DO UPDATE SET item_name=excluded.item_name""",
                (branch_code, branch_name, item_name, item_code, now)
            )
            # ⚠️ inventory에도 없으면 0으로 생성 (기존 코드에 누락돼 있던 부분 보강)
            conn.execute(
                """INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated)
                   VALUES (?, ?, ?, 0, ?)
                   ON CONFLICT(branch_code, item_code) DO NOTHING""",
                (branch_code, item_name, item_code, now)
            )
            # ⚠️ RAW 재고에 0으로 자동 생성 (이미 있으면 건드리지 않음, source='master')
            existing_raw = conn.execute(
                "SELECT id FROM raw_inventory WHERE branch_code=? AND item_code=? AND source='master'",
                (branch_code, item_code)
            ).fetchone()
            if not existing_raw:
                conn.execute(
                    """INSERT INTO raw_inventory
                       (branch_code, branch_name, item_name, item_code, quantity, source, uploaded_at)
                       VALUES (?, ?, ?, ?, 0, 'master', ?)""",
                    (branch_code, branch_name, item_name, item_code, now)
                )
        except Exception:
            continue
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/data", status_code=303)


@app.post("/master/data/delete-selected")
async def master_data_delete_selected(
    request: Request,
    session_token: str = Cookie(default=None)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    form = await request.form()
    ids = form.getlist("selected_ids")
    if ids:
        conn = get_conn()
        conn.execute(
            f"DELETE FROM items WHERE id IN ({','.join('?' for _ in ids)})",
            [int(i) for i in ids]
        )
        conn.commit()
        conn.close()
    return RedirectResponse(url="/master/data", status_code=303)


@app.post("/master/data/delete-all")
async def master_data_delete_all(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)
    conn = get_conn()
    conn.execute("DELETE FROM items")
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/data", status_code=303)


@app.get("/master/purchase-tracking", response_class=HTMLResponse)
async def purchase_tracking_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    record_count = conn.execute("SELECT COUNT(*) as cnt FROM purchase_records").fetchone()["cnt"]
    leadtime_count = conn.execute("SELECT COUNT(*) as cnt FROM product_lead_time").fetchone()["cnt"]
    latest_purchase = conn.execute("SELECT MAX(purchase_datetime) as t FROM purchase_records").fetchone()["t"]
    conn.close()

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>📈 발주 주기 트래킹</h2>
    </div>

    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">
        구매내역과 상품별 리드타임(목표 구매주기)을 업로드하면, 지점×상품별 실제 구매 주기와 목표 주기를 비교해 발주 시기/수량을 추천합니다.
      </p>
    </div>

    <div class="card" style="display:flex;gap:24px;padding:16px;">
      <div><span style="color:#888;font-size:12px;">구매내역 (최근 3개월)</span> <b style="font-size:16px;">{record_count}건</b></div>
      <div><span style="color:#888;font-size:12px;">리드타임 등록 상품</span> <b style="font-size:16px;">{leadtime_count}개</b></div>
      <div><span style="color:#888;font-size:12px;">최근 구매일</span> <b style="font-size:16px;">{(latest_purchase or '없음')[:16]}</b></div>
    </div>

    <div class="card" style="display:flex;gap:8px;">
      <a href="/master/purchase-tracking/status" class="btn" style="text-decoration:none;flex:1;text-align:center;">📊 발주 추천 현황 보기</a>
      <a href="/master/purchase-tracking/history" class="btn" style="text-decoration:none;flex:1;text-align:center;background:#64748B;">📈 편차 추이 히스토리</a>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">📦 구매내역 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:12px;">
        컬럼: 구매일시 / 구분 / 거래처 / 지점 / 담당자 / 상품명 / 품번 / 수량 / 단가 / 합계금액 / 메모 / 등록일시 (2행 헤더)
      </p>
      <div style="display:flex;gap:8px;align-items:center;">
        <input type="file" id="recordsFile" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="button" onclick="uploadRecords()">업로드</button>
      </div>
      <div id="recordsResult" style="display:none;margin-top:12px;padding:12px;border-radius:8px;font-size:13px;"></div>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">🎯 상품별 리드타임(목표 구매주기) 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:12px;">
        '상품명', '리드타임' 헤더가 포함된 상품내역 엑셀을 그대로 업로드하면 됩니다 (2행 헤더). 리드타임이 비어있거나 0인 상품은 건너뜁니다.
      </p>
      <div style="display:flex;gap:8px;align-items:center;">
        <input type="file" id="leadtimeFile" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="button" onclick="uploadLeadtime()">업로드</button>
      </div>
      <div id="leadtimeResult" style="display:none;margin-top:12px;padding:12px;border-radius:8px;font-size:13px;"></div>
    </div>

    <script>
      async function uploadRecords() {{
        const file = document.getElementById('recordsFile').files[0];
        if (!file) {{ alert('파일을 선택해주세요.'); return; }}
        const fd = new FormData();
        fd.append('file', file);
        const btn = event.target;
        btn.textContent = '업로드 중...';
        btn.disabled = true;
        try {{
          const res = await fetch('/master/purchase-tracking/upload-records', {{ method: 'POST', body: fd }});
          const data = await res.json();
          const box = document.getElementById('recordsResult');
          box.style.display = 'block';
          box.style.background = data.errors && data.errors.length ? '#FEF9C3' : '#D1FAE5';
          box.innerHTML = `<b>${{data.errors && data.errors.length ? '⚠️' : '✅'}} 업로드 완료</b><br>
          성공: <b style="color:#22C55E">${{data.success}}건</b> &nbsp;
          실패: <b style="color:#EF4444">${{data.skipped}}건</b>
          ${{data.errors && data.errors.length ? '<ul>' + data.errors.map(e=>`<li style="color:#EF4444;font-size:12px;">${{e}}</li>`).join('') + '</ul>' : ''}}`;
          setTimeout(() => location.reload(), 2000);
        }} catch(e) {{
          alert('업로드 중 오류가 발생했습니다.');
        }} finally {{
          btn.textContent = '업로드';
          btn.disabled = false;
        }}
      }}

      async function uploadLeadtime() {{
        const file = document.getElementById('leadtimeFile').files[0];
        if (!file) {{ alert('파일을 선택해주세요.'); return; }}
        const fd = new FormData();
        fd.append('file', file);
        const btn = event.target;
        btn.textContent = '업로드 중...';
        btn.disabled = true;
        try {{
          const res = await fetch('/master/purchase-tracking/upload-leadtime', {{ method: 'POST', body: fd }});
          const data = await res.json();
          const box = document.getElementById('leadtimeResult');
          box.style.display = 'block';
          box.style.background = data.errors && data.errors.length ? '#FEF9C3' : '#D1FAE5';
          box.innerHTML = `<b>${{data.errors && data.errors.length ? '⚠️' : '✅'}} 업로드 완료</b><br>
          성공: <b style="color:#22C55E">${{data.success}}건</b> &nbsp;
          건너뜀(리드타임 없음/0): <b style="color:#888">${{data.skipped}}건</b>
          ${{data.errors && data.errors.length ? '<ul>' + data.errors.map(e=>`<li style="color:#EF4444;font-size:12px;">${{e}}</li>`).join('') + '</ul>' : ''}}`;
          setTimeout(() => location.reload(), 2000);
        }} catch(e) {{
          alert('업로드 중 오류가 발생했습니다.');
        }} finally {{
          btn.textContent = '업로드';
          btn.disabled = false;
        }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/purchase-tracking/upload-records")
async def purchase_tracking_upload_records(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return {"success": 0, "skipped": 0, "errors": ["로그인이 필요합니다"]}
    return await _process_purchase_records_upload(file)

# ── 유비플러스 재고 (RAW 업로드) ────────────────────────

@app.get("/master/raw-upload", response_class=HTMLResponse)
async def raw_upload_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    raws = conn.execute(
        "SELECT * FROM raw_inventory ORDER BY branch_code, item_code"
    ).fetchall()
    conn.close()

    rows_html = ""
    if not raws:
        rows_html = '<tr><td colspan="5" style="text-align:center;padding:20px;color:#888;">업로드된 데이터 없음</td></tr>'
    else:
        for r in raws:
            rows_html += f"""
            <tr>
              <td>{r['branch_name']}</td>
              <td>{r['item_name']}</td>
              <td>{r['item_code']}</td>
              <td style="font-weight:bold;">{r['quantity']}</td>
              <td>{r['uploaded_at'][:10] if r['uploaded_at'] else '-'}</td>
            </tr>"""

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>📤 유비플러스 재고</h2>
    </div>
    <div class="card" style="background:#FFF7ED;border:1px solid #FCD34D;">
      <p style="font-size:13px;color:#92400E;">
        ⚠️ MSSQL 연동 전 임시 기능입니다. 업로드한 데이터가 대시보드 비교 기준(RAW)으로 사용됩니다.
        H열/Q열 값은 QR재고에 자동 가산됩니다.
      </p>
    </div>
    <div class="card">
      <h3 style="margin-bottom:8px;">엑셀 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:12px;">
        컬럼 위치: <b>A=지점명 / B=상품명 / D=품번 / N=현재수량 / H,Q=QR재고 가산분</b> (1행 헤더)
      </p>
      <div style="display:flex;gap:8px;align-items:center;">
        <input type="file" id="rawFile" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="button" onclick="uploadRaw()">업로드</button>
      </div>
      <div id="uploadResult" style="display:none;margin-top:12px;padding:12px;
           border-radius:8px;font-size:13px;"></div>
      <script>
      async function uploadRaw() {{
        const file = document.getElementById('rawFile').files[0];
        if (!file) {{ alert('파일을 선택해주세요.'); return; }}
        const fd = new FormData();
        fd.append('file', file);
        const btn = event.target;
        btn.textContent = '업로드 중...';
        btn.disabled = true;
        try {{
          const res = await fetch('/master/raw-upload/ajax', {{
            method: 'POST', body: fd
          }});
          const data = await res.json();
          const box = document.getElementById('uploadResult');
          box.style.display = 'block';
          box.style.background = data.errors.length ? '#FEF9C3' : '#D1FAE5';
          box.innerHTML = `<b>${{data.errors.length ? '⚠️' : '✅'}} 업로드 완료</b><br>
          헤더 인식 행: <b>${{data.header_row_used ?? '?'}}행</b><br>
          컬럼 매핑: <b>${{JSON.stringify(data.col_map_debug)}}</b><br>
          성공: <b style="color:#22C55E">${{data.success}}건</b> &nbsp;
          실패: <b style="color:#EF4444">${{data.skipped}}건</b>
          ${{data.errors.length ? '<ul>' + data.errors.map(e=>`<li style="color:#EF4444;font-size:12px;">${{e}}</li>`).join('') + '</ul>' : ''}}
          ${{data.hq_debug && data.hq_debug.length ? '<br><b>H/Q 반영 내역:</b><ul>' + data.hq_debug.map(e=>`<li style="font-size:12px;">${{e}}</li>`).join('') + '</ul>' : '<br><span style="color:#F59E0B;">⚠️ H/Q 반영된 품목 없음</span>'}}`;
          setTimeout(() => location.reload(), 2000);
        }} catch(e) {{
          alert('업로드 중 오류가 발생했습니다.');
        }} finally {{
          btn.textContent = '업로드';
          btn.disabled = false;
        }}
      }}
      </script>
    </div>
    <div class="card">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">
        <h3>현재 유비플러스 데이터 ({len(raws)}개)</h3>
        <form method="post" action="/master/raw-upload/clear">
          <button type="submit" class="btn btn-red"
                  style="font-size:12px;padding:6px 12px;"
                  onclick="return confirm('전체 데이터를 삭제합니다.')">전체 초기화</button>
        </form>
      </div>
      <table>
        <thead><tr>
          <th>지점명</th><th>상품명</th><th>품번</th><th>수량</th><th>업로드일</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))

@app.post("/master/purchase-tracking/upload-leadtime")
async def purchase_tracking_upload_leadtime(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return {"success": 0, "skipped": 0, "errors": ["로그인이 필요합니다"]}
    return await _process_lead_time_upload(file)

# ============================================================
# CSS는 f-string 밖에서 별도 상수로 관리 (중괄호 이스케이프 문제 원천 차단)
# ============================================================
PRODUCT_SETTINGS_EXTRA_CSS = """
<style>
#colCheckboxList input[type="checkbox"] { width: auto !important; }
.pm-table input[type="text"], .pm-table input[type="number"] { width: 100%; padding: 6px; box-sizing: border-box; }
.pm-upload-box { border: 2px dashed #93C5FD; border-radius: 10px; padding: 16px; background: #EFF6FF; }
</style>
"""


@app.get("/master/purchase-order/product-settings", response_class=HTMLResponse)
async def purchase_order_product_settings_page(
    session_token: str = Cookie(default=None),
    search_item: str = "",
    filter_branch: str = "",
    filter_consumable: str = "",
    sort_by: str = "item_name",
    sort_dir: str = "asc"
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    allowed_sort_cols = {"item_name", "branch_name", "lead_time_days", "moq", "is_consumable", "purchase_price"}
    if sort_by not in allowed_sort_cols:
        sort_by = "item_name"
    sort_dir_sql = "DESC" if sort_dir == "desc" else "ASC"

    conn = get_conn()
    query = "SELECT * FROM product_master WHERE 1=1"
    params: list = []
    if search_item:
        query += " AND (item_name LIKE ? OR item_code LIKE ?)"
        params.append(f"%{search_item}%")
        params.append(f"%{search_item}%")
    if filter_branch:
        query += " AND branch_name = ?"
        params.append(filter_branch)
    if filter_consumable == "yes":
        query += " AND is_consumable = TRUE"
    elif filter_consumable == "no":
        query += " AND is_consumable = FALSE"
    query += f" ORDER BY {sort_by} {sort_dir_sql} LIMIT 300"
    rows = conn.execute(query, params).fetchall()

    monthly_exc_rows = conn.execute(
        "SELECT item_name, branch_code FROM purchase_order_monthly_exception"
    ).fetchall()
    conn.close()

    monthly_exc_map: Dict[str, list] = {}
    for r in monthly_exc_rows:
        monthly_exc_map.setdefault(r["item_name"], []).append(r["branch_code"])

    branches = get_branches(branch_type='branch')
    branch_options_json = json.dumps(
        [{"code": b["branch_code"], "name": b["branch_name"]} for b in branches], ensure_ascii=False
    )
    branch_filter_options = '<option value="">전체 지점</option>'
    for b in branches:
        sel = "selected" if filter_branch == b["branch_name"] else ""
        branch_filter_options += f'<option value="{b["branch_name"]}" {sel}>{b["branch_name"]}</option>'

    rows_html = ""
    if not rows:
        rows_html = '<tr><td colspan="9" style="text-align:center;padding:20px;color:#888;">등록된 상품이 없습니다.</td></tr>'
    else:
        for r in rows:
            consumable_checked = "checked" if r["is_consumable"] else ""
            exc_branches = monthly_exc_map.get(r["item_name"], [])
            exc_badge = (
                f'<span class="badge-green">{len(exc_branches)}개 지점</span>'
                if exc_branches else '<span style="color:#888;font-size:11px;">없음</span>'
            )
            safe_item_name = (r["item_name"] or "").replace("'", "")
            rows_html += f"""
            <tr>
              <td>{r['branch_name'] or ''}</td>
              <td>{r['item_name'] or ''}</td>
              <td>{r['item_code'] or ''}</td>
              <td><input type="number" class="pm-price" data-id="{r['id']}" value="{r['purchase_price'] or 0}"></td>
              <td>{r['supplier'] or ''}</td>
              <td>{r['tax_setting'] or ''}</td>
              <td><input type="number" class="pm-leadtime" data-id="{r['id']}" value="{r['lead_time_days'] or 0}"></td>
              <td><input type="number" class="pm-moq" data-id="{r['id']}" value="{r['moq'] or 1}"></td>
              <td style="text-align:center;"><input type="checkbox" class="pm-consumable" data-id="{r['id']}" {consumable_checked} style="width:18px;height:18px;"></td>
              <td>
                {exc_badge}
                <button class="btn" style="font-size:11px;padding:4px 8px;background:#8B5CF6;" onclick="openMonthlyExc('{safe_item_name}')">설정</button>
              </td>
              <td><button class="btn" style="font-size:12px;padding:6px 12px;" onclick="savePmRow({r['id']})">저장</button></td>
            </tr>
            """

    content = f"""
    {PRODUCT_SETTINGS_EXTRA_CSS}
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master/purchase-tracking" style="color:#1E2761;">← 발주 주기 트래킹</a>
      <h2>📝 상품 마스터 (단가/거래처/리드타임/MOQ/소모품)</h2>
    </div>

    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">상품별 단가, 거래처, 과세설정, 리드타임, MOQ, 소모품 여부를 관리합니다. 발주서 자동생성 시 이 정보를 사용합니다.</p>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">새 상품 수동 등록</h3>
      <div style="display:flex;gap:8px;flex-wrap:wrap;">
        <input type="text" id="pmBranch" placeholder="지점명" style="flex:1;min-width:100px;">
        <input type="text" id="pmItemName" placeholder="상품명" style="flex:2;min-width:180px;">
        <input type="text" id="pmItemCode" placeholder="품번" style="flex:1;min-width:120px;">
        <input type="number" id="pmPrice" placeholder="구매금액" style="flex:1;min-width:100px;">
        <input type="text" id="pmSupplier" placeholder="거래처명" style="flex:1;min-width:100px;">
        <input type="text" id="pmTax" placeholder="과세설정(예: 과세)" style="flex:1;min-width:100px;">
        <input type="number" id="pmLeadTime" placeholder="리드타임(일)" style="flex:1;min-width:100px;">
        <input type="number" id="pmMoq" placeholder="MOQ(기본 1)" style="flex:1;min-width:100px;">
        <label style="display:flex;align-items:center;gap:6px;font-size:13px;"><input type="checkbox" id="pmConsumable"> 소모품</label>
        <button class="btn" type="button" onclick="addPmItem()">추가</button>
      </div>
      <div id="pmAddResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card pm-upload-box">
      <h3 style="margin-bottom:8px;">엑셀 업로드 (상품마스터 일괄 등록/갱신)</h3>
      <p style="font-size:12px;color:#1E40AF;margin-bottom:8px;">
        헤더 포함, 컬럼: 지점명 | 상품분류 | 상품명 | 상품내용 | 출고금액 | 구매금액 | 상품타입 | 품번 | 규격 | 판매금액 | 단위 | 과세설정 | ... | 리드타임 | ... | 구매처<br>
        (지점명+품번 기준으로 이미 있으면 덮어씁니다.)
      </p>
      <input type="file" id="pmExcelFile" accept=".xlsx,.xls">
      <button class="btn" type="button" onclick="uploadPmExcel()">업로드</button>
      <div id="pmUploadResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card">
      <form method="get" action="/master/purchase-order/product-settings" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;">
        <input name="search_item" value="{search_item}" placeholder="상품명/품번 검색" style="flex:2;min-width:180px;">
        <select name="filter_branch" style="flex:1;min-width:120px;">{branch_filter_options}</select>
        <select name="filter_consumable" style="flex:1;min-width:120px;">
          <option value="" {'selected' if not filter_consumable else ''}>전체</option>
          <option value="yes" {'selected' if filter_consumable == 'yes' else ''}>소모품만</option>
          <option value="no" {'selected' if filter_consumable == 'no' else ''}>일반상품만</option>
        </select>
        <button class="btn" type="submit">검색</button>
        <a href="/master/purchase-order/product-settings" style="padding:10px 14px;background:#eee;border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
      <p style="font-size:13px;color:#888;margin-bottom:12px;">{len(rows)}건 (최대 300건까지 표시)</p>
      <table class="pm-table">
        <thead><tr>
          <th>지점</th>
          <th style="cursor:pointer;" onclick="sortPm('item_name')">상품명 {'▲' if sort_by=='item_name' and sort_dir=='asc' else ('▼' if sort_by=='item_name' else '')}</th>
          <th>품번</th>
          <th style="cursor:pointer;" onclick="sortPm('purchase_price')">구매금액 {'▲' if sort_by=='purchase_price' and sort_dir=='asc' else ('▼' if sort_by=='purchase_price' else '')}</th>
          <th>거래처</th>
          <th>과세설정</th>
          <th style="cursor:pointer;" onclick="sortPm('lead_time_days')">리드타임(일) {'▲' if sort_by=='lead_time_days' and sort_dir=='asc' else ('▼' if sort_by=='lead_time_days' else '')}</th>
          <th style="cursor:pointer;" onclick="sortPm('moq')">MOQ {'▲' if sort_by=='moq' and sort_dir=='asc' else ('▼' if sort_by=='moq' else '')}</th>
          <th style="cursor:pointer;" onclick="sortPm('is_consumable')">소모품 {'▲' if sort_by=='is_consumable' and sort_dir=='asc' else ('▼' if sort_by=='is_consumable' else '')}</th>
          <th>월1회예외 지점</th><th></th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>

    <div id="monthlyExcModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:420px;width:90%;max-height:80vh;overflow-y:auto;">
        <h3 id="monthlyExcTitle" style="margin-bottom:12px;">월1회 예외 지점 설정</h3>
        <div id="monthlyExcBranchList" style="margin-bottom:16px;"></div>
        <div style="display:flex;gap:8px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeMonthlyExc()">닫기</button>
          <button class="btn" style="flex:1;" onclick="saveMonthlyExc()">저장</button>
        </div>
      </div>
    </div>

    <script>
      const allBranches = {branch_options_json};

      function sortPm(col) {{
        const params = new URLSearchParams(window.location.search);
        const curSort = params.get('sort_by') || 'item_name';
        const curDir = params.get('sort_dir') || 'asc';
        const newDir = (curSort === col && curDir === 'asc') ? 'desc' : 'asc';
        params.set('sort_by', col);
        params.set('sort_dir', newDir);
        window.location.search = params.toString();
      }}

      let currentExcItemName = '';

      async function openMonthlyExc(itemName) {{
        currentExcItemName = itemName;
        document.getElementById('monthlyExcTitle').innerText = itemName + ' — 월1회 예외 지점';
        const res = await fetch('/master/purchase-order/monthly-exception/list?item_name=' + encodeURIComponent(itemName));
        const data = await res.json();
        const checkedSet = new Set(data.branches || []);
        const container = document.getElementById('monthlyExcBranchList');
        container.innerHTML = allBranches.map(b =>
          '<label style="display:block;font-size:13px;margin-bottom:6px;"><input type="checkbox" class="mexc-branch" value="' + b.code + '" ' +
          (checkedSet.has(b.code) ? 'checked' : '') + ' style="margin-right:6px;">' + b.name + '</label>'
        ).join('');
        document.getElementById('monthlyExcModal').style.display = 'flex';
      }}

      function closeMonthlyExc() {{
        document.getElementById('monthlyExcModal').style.display = 'none';
      }}

      async function saveMonthlyExc() {{
        const checked = Array.from(document.querySelectorAll('.mexc-branch:checked')).map(c => c.value);
        const res = await fetch('/master/purchase-order/monthly-exception/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ item_name: currentExcItemName, branch_codes: checked }})
        }});
        if (res.ok) {{ location.reload(); }} else {{ alert('저장 실패'); }}
      }}

      async function addPmItem() {{
        const branch = document.getElementById('pmBranch').value.trim();
        const name = document.getElementById('pmItemName').value.trim();
        const code = document.getElementById('pmItemCode').value.trim();
        const price = parseFloat(document.getElementById('pmPrice').value) || 0;
        const supplier = document.getElementById('pmSupplier').value.trim();
        const tax = document.getElementById('pmTax').value.trim();
        const leadTime = parseInt(document.getElementById('pmLeadTime').value) || 0;
        const moq = parseInt(document.getElementById('pmMoq').value) || 1;
        const consumable = document.getElementById('pmConsumable').checked;
        if (!branch || !name) {{ alert('지점명과 상품명을 입력하세요.'); return; }}
        const res = await fetch('/master/purchase-order/product-master/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{
            branch_name: branch, item_name: name, item_code: code, purchase_price: price,
            supplier: supplier, tax_setting: tax, lead_time_days: leadTime, moq: moq, is_consumable: consumable
          }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('pmAddResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}

      async function savePmRow(id) {{
        const priceEl = document.querySelector('.pm-price[data-id="' + id + '"]');
        const leadTimeEl = document.querySelector('.pm-leadtime[data-id="' + id + '"]');
        const moqEl = document.querySelector('.pm-moq[data-id="' + id + '"]');
        const consumableEl = document.querySelector('.pm-consumable[data-id="' + id + '"]');
        const res = await fetch('/master/purchase-order/product-master/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{
            id: id,
            purchase_price: parseFloat(priceEl.value) || 0,
            lead_time_days: parseInt(leadTimeEl.value) || 0,
            moq: parseInt(moqEl.value) || 1,
            is_consumable: consumableEl.checked
          }})
        }});
        if (res.ok) {{ alert('저장되었습니다.'); }} else {{ alert('저장 실패'); }}
      }}

      async function uploadPmExcel() {{
        const fileEl = document.getElementById('pmExcelFile');
        if (!fileEl.files.length) {{ alert('파일을 선택하세요.'); return; }}
        const formData = new FormData();
        formData.append('file', fileEl.files[0]);
        document.getElementById('pmUploadResult').innerText = '업로드 중...';
        const res = await fetch('/master/purchase-order/product-master/upload', {{
          method: 'POST', body: formData
        }});
        const data = await res.json();
        if (res.ok) {{
          document.getElementById('pmUploadResult').innerText = '완료: ' + data.inserted + '건 등록/갱신';
          setTimeout(() => location.reload(), 1200);
        }} else {{
          document.getElementById('pmUploadResult').innerText = '오류: ' + (data.detail || '업로드 실패');
        }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/purchase-order/product-master/save")
async def purchase_order_product_master_save(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    row_id = data.get("id")
    branch_name = (data.get("branch_name") or "").strip()
    item_name = (data.get("item_name") or "").strip()
    item_code = (data.get("item_code") or "").strip()
    purchase_price = data.get("purchase_price", 0)
    supplier = (data.get("supplier") or "").strip()
    tax_setting = (data.get("tax_setting") or "").strip()
    lead_time_days = data.get("lead_time_days", 0)
    moq = data.get("moq", 1)
    is_consumable = bool(data.get("is_consumable", False))

    now = datetime.now().isoformat()
    conn = get_conn()
    if row_id:
        conn.execute(
            """UPDATE product_master
               SET purchase_price=?, lead_time_days=?, moq=?, is_consumable=?, updated_at=?
               WHERE id=?""",
            (purchase_price, lead_time_days, moq, is_consumable, now, row_id)
        )
    else:
        if not branch_name or not item_name:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "지점명과 상품명을 입력하세요."})
        conn.execute("""
            INSERT INTO product_master
                (branch_name, item_name, item_code, purchase_price, supplier,
                 tax_setting, lead_time_days, moq, is_consumable, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (branch_name, item_code) DO UPDATE SET
                item_name=excluded.item_name, purchase_price=excluded.purchase_price,
                supplier=excluded.supplier, tax_setting=excluded.tax_setting,
                lead_time_days=excluded.lead_time_days, moq=excluded.moq,
                is_consumable=excluded.is_consumable, updated_at=excluded.updated_at
        """, (branch_name, item_name, item_code, purchase_price, supplier,
              tax_setting, lead_time_days, moq, is_consumable, now))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/purchase-order/product-master/upload")
async def purchase_order_product_master_upload(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse(status_code=400, content={"detail": "파일이 없습니다."})

    import openpyxl
    from io import BytesIO
    raw = await file.read()
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    header_row_idx = None
    header_map = {}
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row and "지점명" in row and "상품명" in row:
            header_row_idx = idx
            for col_idx, col_name in enumerate(row):
                if col_name:
                    header_map[col_name.strip()] = col_idx
            break

    if header_row_idx is None:
        return JSONResponse(status_code=400, content={"detail": "헤더(지점명/상품명 등)를 찾을 수 없습니다."})

    required = ["지점명", "상품명"]
    for col in required:
        if col not in header_map:
            return JSONResponse(status_code=400, content={"detail": f"필수 컬럼 누락: {col}"})

    now = datetime.now().isoformat()
    conn = get_conn()
    inserted = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row:
            continue

        def get_col(name, default=None):
            idx = header_map.get(name)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return val if val is not None else default

        branch_name = str(get_col("지점명", "") or "").strip()
        item_name = str(get_col("상품명", "") or "").strip()
        if not branch_name or not item_name:
            continue

        item_code = str(get_col("품번", "") or "").strip()
        purchase_price = get_col("구매금액", 0) or 0
        supplier = str(get_col("구매처", "") or "").strip()
        tax_setting = str(get_col("과세설정", "") or "").strip()
        lead_time_raw = get_col("리드타임", 0)
        try:
            lead_time_days = int(lead_time_raw) if lead_time_raw not in (None, "") else 0
        except (ValueError, TypeError):
            lead_time_days = 0

        conn.execute("""
            INSERT INTO product_master
                (branch_name, item_name, item_code, purchase_price, supplier, tax_setting, lead_time_days, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (branch_name, item_code) DO UPDATE SET
                item_name=excluded.item_name, purchase_price=excluded.purchase_price,
                supplier=excluded.supplier, tax_setting=excluded.tax_setting,
                lead_time_days=excluded.lead_time_days, updated_at=excluded.updated_at
        """, (branch_name, item_name, item_code, purchase_price, supplier, tax_setting, lead_time_days, now))
        inserted += 1

    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok", "inserted": inserted})


@app.get("/master/purchase-order/monthly-exception/list")
async def purchase_order_monthly_exception_list(item_name: str, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    conn = get_conn()
    rows = conn.execute(
        "SELECT branch_code FROM purchase_order_monthly_exception WHERE item_name=?", (item_name,)
    ).fetchall()
    conn.close()
    return JSONResponse(content={"branches": [r["branch_code"] for r in rows]})


@app.post("/master/purchase-order/monthly-exception/save")
async def purchase_order_monthly_exception_save(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})
    data = await request.json()
    item_name = (data.get("item_name") or "").strip()
    branch_codes = data.get("branch_codes", [])

    if not item_name:
        return JSONResponse(status_code=400, content={"detail": "상품명이 없습니다."})

    conn = get_conn()
    conn.execute("DELETE FROM purchase_order_monthly_exception WHERE item_name=?", (item_name,))
    for bc in branch_codes:
        conn.execute(
            "INSERT INTO purchase_order_monthly_exception (item_name, branch_code) VALUES (?, ?)",
            (item_name, bc)
        )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})

@app.get("/master/purchase-tracking/status", response_class=HTMLResponse)
async def purchase_tracking_status_page(
    session_token: str = Cookie(default=None),
    filter_branch: str = "",
    search_item: str = "",
    filter_date: str = "",
    filter_deviation: str = "",
    sort_by: str = "deviation_abs",
    sort_dir: str = "desc"
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    latest_week_row = conn.execute(
        "SELECT MAX(snapshot_week) as w FROM purchase_tracking_snapshot"
    ).fetchone()
    latest_week = latest_week_row["w"] if latest_week_row else None

    if not latest_week:
        conn.close()
        content = """
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
          <a href="/master/purchase-tracking" style="color:#1E2761;">← 발주 주기 트래킹</a>
          <h2>📊 발주 추천 현황</h2>
        </div>
        <div class="card" style="text-align:center;padding:40px;">
          <div style="font-size:32px;">📭</div>
          <p style="color:#888;margin-top:12px;">아직 계산된 스냅샷이 없습니다. cron이 최초 1회 실행된 후 표시됩니다.</p>
        </div>
        """
        return HTMLResponse(content=render_page(content, user, "master"))

    query = "SELECT * FROM purchase_tracking_snapshot WHERE snapshot_week=?"
    params: list = [latest_week]
    if filter_branch:
        query += " AND branch_name=?"
        params.append(filter_branch)
    if search_item:
        query += " AND item_name LIKE ?"
        params.append(f"%{search_item}%")
    if filter_date:
        query += " AND last_purchase_date LIKE ?"
        params.append(f"{filter_date}%")
    if filter_deviation == "over":
        query += " AND deviation_days > 3"
    elif filter_deviation == "delay":
        query += " AND deviation_days < -3"
    elif filter_deviation == "ok":
        query += " AND deviation_days BETWEEN -3 AND 3"

    sort_dir_sql = "DESC" if sort_dir == "desc" else "ASC"
    allowed_sort_map = {
        "branch_name": "branch_name",
        "item_name": "item_name",
        "last_purchase_date": "last_purchase_date",
        "deviation": "deviation_days",
        "recommended_qty": "recommended_qty",
        "deviation_abs": "ABS(deviation_days)",
    }
    sort_col_sql = allowed_sort_map.get(sort_by, "ABS(deviation_days)")
    query += f" ORDER BY {sort_col_sql} {sort_dir_sql}"

    rows = conn.execute(query, params).fetchall()

    all_branches = conn.execute(
        "SELECT DISTINCT branch_name FROM purchase_tracking_snapshot WHERE snapshot_week=? ORDER BY branch_name",
        (latest_week,)
    ).fetchall()

    branch_name_to_code = {b["branch_name"]: b["branch_code"] for b in get_branches()}

    prev_purchase_map = {}
    for r in rows:
        key = (r["branch_name"], r["item_name"])
        if key in prev_purchase_map:
            continue
        hist = conn.execute("""
            SELECT purchase_datetime, quantity FROM purchase_records
            WHERE branch_name=? AND item_name=?
            ORDER BY purchase_datetime DESC LIMIT 2
        """, (r["branch_name"], r["item_name"])).fetchall()
        prev_purchase_map[key] = hist

    raw_stock_map = {}
    raw_rows = conn.execute("SELECT branch_code, item_name, quantity FROM raw_inventory").fetchall()
    for rr in raw_rows:
        raw_stock_map[(rr["branch_code"], rr["item_name"])] = rr["quantity"]

    conn.close()

    def fmt_int(val):
        if val is None:
            return "-"
        return f"{round(val):,}"

    branch_options = '<option value="">전체 지점</option>'
    for b in all_branches:
        sel = "selected" if filter_branch == b["branch_name"] else ""
        branch_options += f'<option value="{b["branch_name"]}" {sel}>{b["branch_name"]}</option>'

    dev_options = f"""
        <option value="" {'selected' if not filter_deviation else ''}>전체</option>
        <option value="over" {'selected' if filter_deviation == 'over' else ''}>과다구매 (너무 자주 구매)</option>
        <option value="ok" {'selected' if filter_deviation == 'ok' else ''}>적정</option>
        <option value="delay" {'selected' if filter_deviation == 'delay' else ''}>구매지연</option>
    """

    def sort_arrow(col):
        if sort_by != col:
            return ""
        return "▲" if sort_dir == "asc" else "▼"

    rows_html = ""
    if not rows:
        rows_html = '<tr><td colspan="9" style="text-align:center;padding:20px;color:#888;">조건에 맞는 데이터가 없습니다.</td></tr>'
    else:
        for idx, r in enumerate(rows):
            dev = r["deviation_days"]
            if dev is None:
                dev_display = '<span style="color:#888;">-</span>'
            elif dev > 3:
                dev_display = f'<span style="color:#F59E0B;font-weight:bold;">+{dev}</span>'
            elif dev < -3:
                dev_display = f'<span style="color:#EF4444;font-weight:bold;">{dev}</span>'
            else:
                dev_display = f'<span style="color:#22C55E;">{dev:+d}</span>'

            raw_qty = r["recommended_qty"]
            if raw_qty is not None and raw_qty < 5:
                qty_display = "5 (최소)"
            elif raw_qty is not None:
                qty_display = f"{round(raw_qty):,}"
            else:
                qty_display = "-"

            key = (r["branch_name"], r["item_name"])
            hist = prev_purchase_map.get(key, [])
            last_qty_display = fmt_int(hist[0]["quantity"]) if len(hist) >= 1 else "-"
            prev_date_display = hist[1]["purchase_datetime"][:10] if len(hist) >= 2 else "-"
            prev_qty_display = fmt_int(hist[1]["quantity"]) if len(hist) >= 2 else "-"

            branch_code = branch_name_to_code.get(r["branch_name"])
            current_stock = raw_stock_map.get((branch_code, r["item_name"])) if branch_code else None
            stock_display = fmt_int(current_stock)

            last_date_display = (r["last_purchase_date"] or "-")[:10]
            next_date_display = r["recommended_next_date"] or "-"
            detail_id = f"ptdetail_{idx}"

            rows_html += f"""
            <tr style="cursor:pointer;" onclick="togglePtDetail('{detail_id}')">
              <td class="pt-col-branch">{r['branch_name']}</td>
              <td class="pt-col-item">{r['item_name']}</td>
              <td class="pt-col-prevdate" style="font-size:11px;color:#888;">{prev_date_display}</td>
              <td class="pt-col-prevqty" style="text-align:center;color:#888;">{prev_qty_display}</td>
              <td class="pt-col-lastdate" style="font-size:11px;">{last_date_display}</td>
              <td class="pt-col-lastqty" style="text-align:center;">{last_qty_display}</td>
              <td class="pt-col-stock" style="text-align:right;">{stock_display}</td>
              <td class="pt-col-deviation">{dev_display}</td>
              <td class="pt-col-qty" style="text-align:right;font-weight:bold;">{qty_display}</td>
              <td class="pt-col-arrow" style="text-align:center;color:#aaa;" id="{detail_id}_arrow">▼</td>
            </tr>
            <tr id="{detail_id}" style="display:none;background:#f8fafc;">
              <td colspan="10" style="padding:10px 16px;">
                <div style="display:grid;grid-template-columns:repeat(auto-fit, minmax(130px, 1fr));gap:8px;font-size:11px;">
                  <div><span style="color:#888;">실제주기(A)</span><br><b>{r['actual_interval_days']}일</b></div>
                  <div><span style="color:#888;">목표주기(B)</span><br><b>{r['lead_time_days']}일</b></div>
                  <div><span style="color:#888;">추천발주일</span><br><b>{next_date_display}</b></div>
                </div>
              </td>
            </tr>
            """

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master/purchase-tracking" style="color:#1E2761;">← 발주 주기 트래킹</a>
      <h2>📊 발주 추천 현황</h2>
    </div>
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">
        기준 주차: <b>{latest_week}</b> · 컬럼 헤더를 클릭하면 정렬됩니다. 행을 클릭하면 상세정보가 펼쳐집니다.<br>
        <span style="color:#F59E0B;">주황(과다구매)</span> / <span style="color:#EF4444;">빨강(구매지연)</span> / <span style="color:#22C55E;">초록(적정)</span>
        · 추천수량은 최소 5개로 보정됩니다.
      </p>
    </div>
    <div class="card">
      <form method="get" action="/master/purchase-tracking/status" style="display:flex;gap:8px;flex-wrap:wrap;align-items:flex-end;">
        <input type="hidden" name="sort_by" value="{sort_by}">
        <input type="hidden" name="sort_dir" value="{sort_dir}">
        <div style="flex:1;min-width:130px;">
          <label style="font-size:12px;color:#888;">지점 필터</label>
          <select name="filter_branch" style="margin-top:4px;">{branch_options}</select>
        </div>
        <div style="flex:1;min-width:130px;">
          <label style="font-size:12px;color:#888;">상품명 검색</label>
          <input name="search_item" value="{search_item}" placeholder="상품명 검색" style="margin-top:4px;">
        </div>
        <div style="flex:1;min-width:130px;">
          <label style="font-size:12px;color:#888;">마지막구매일</label>
          <input name="filter_date" type="date" value="{filter_date}" style="margin-top:4px;">
        </div>
        <div style="flex:1;min-width:150px;">
          <label style="font-size:12px;color:#888;">구매 상태</label>
          <select name="filter_deviation" style="margin-top:4px;">{dev_options}</select>
        </div>
        <button class="btn" type="submit">검색</button>
        <a href="/master/purchase-tracking/status" style="padding:10px 14px;background:#eee;
           border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
        <button type="button" class="btn" style="background:#64748B;" onclick="openColSettings()">⚙️ 컬럼 설정</button>
        <a href="/master/purchase-tracking/status/export?filter_branch={filter_branch}&search_item={search_item}&filter_date={filter_date}&filter_deviation={filter_deviation}"
           class="btn" style="text-decoration:none;background:#22C55E;">⬇️ 엑셀 다운로드</a>
      </form>
    </div>
    <div class="card">
      <p style="font-size:13px;color:#888;margin-bottom:12px;">{len(rows)}건 (전체 조회)</p>
      <table id="ptTable">
        <thead><tr>
          <th class="pt-col-branch" style="cursor:pointer;" onclick="sortPt('branch_name')">지점 {sort_arrow('branch_name')}</th>
          <th class="pt-col-item" style="cursor:pointer;" onclick="sortPt('item_name')">상품명 {sort_arrow('item_name')}</th>
          <th class="pt-col-prevdate">직전구매일</th>
          <th class="pt-col-prevqty">직전수량</th>
          <th class="pt-col-lastdate" style="cursor:pointer;" onclick="sortPt('last_purchase_date')">마지막구매일 {sort_arrow('last_purchase_date')}</th>
          <th class="pt-col-lastqty">구매수량</th>
          <th class="pt-col-stock">현재고</th>
          <th class="pt-col-deviation" style="cursor:pointer;" onclick="sortPt('deviation')">편차 {sort_arrow('deviation')}</th>
          <th class="pt-col-qty" style="cursor:pointer;" onclick="sortPt('recommended_qty')">추천수량 {sort_arrow('recommended_qty')}</th>
          <th class="pt-col-arrow"></th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>

    <div id="colSettingsModal" style="display:none;position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.5);z-index:1000;align-items:center;justify-content:center;">
      <div style="background:#fff;border-radius:12px;padding:24px;max-width:360px;width:90%;">
        <h3 style="margin-bottom:12px;">⚙️ 표시할 컬럼 선택</h3>
        <div id="colCheckboxList" style="display:flex;flex-direction:column;gap:8px;font-size:14px;"></div>
        <style>#colCheckboxList input[type="checkbox"] {{ width: auto !important; }}</style>
        <div style="display:flex;gap:8px;margin-top:16px;">
          <button class="btn" style="flex:1;background:#eee;color:#333;" onclick="closeColSettings()">닫기</button>
          <button class="btn" style="flex:1;" onclick="applyColSettings()">적용</button>
        </div>
      </div>
    </div>

    <script>
      const PT_COLUMNS = [
        {{ key: 'branch', label: '지점', locked: true }},
        {{ key: 'item', label: '상품명', locked: true }},
        {{ key: 'prevdate', label: '직전구매일', locked: false }},
        {{ key: 'prevqty', label: '직전수량', locked: false }},
        {{ key: 'lastdate', label: '마지막구매일', locked: false }},
        {{ key: 'lastqty', label: '구매수량', locked: false }},
        {{ key: 'stock', label: '현재고', locked: false }},
        {{ key: 'deviation', label: '편차', locked: true }},
        {{ key: 'qty', label: '추천수량', locked: true }},
      ];
      const PT_STORAGE_KEY = 'pt_status_col_settings';

      function getColSettings() {{
        try {{
          const saved = localStorage.getItem(PT_STORAGE_KEY);
          if (saved) return JSON.parse(saved);
        }} catch(e) {{}}
        const defaults = {{}};
        PT_COLUMNS.forEach(c => defaults[c.key] = true);
        return defaults;
      }}

      function applyColVisibility() {{
        const settings = getColSettings();
        PT_COLUMNS.forEach(c => {{
          const visible = settings[c.key] !== false;
          document.querySelectorAll('.pt-col-' + c.key).forEach(el => {{
            el.style.display = visible ? '' : 'none';
          }});
        }});
      }}

      function openColSettings() {{
        const settings = getColSettings();
        const container = document.getElementById('colCheckboxList');
        container.innerHTML = PT_COLUMNS.map(c =>
          '<label style="display:flex;align-items:center;gap:8px;' + (c.locked ? 'opacity:0.5;' : '') + '">' +
          '<input type="checkbox" class="col-setting-check" data-key="' + c.key + '" style="width:18px;height:18px;flex-shrink:0;" ' +
          (settings[c.key] !== false ? 'checked' : '') + (c.locked ? ' disabled' : '') + '>' +
          '<span>' + c.label + (c.locked ? ' (필수)' : '') + '</span></label>'
        ).join('');
        document.getElementById('colSettingsModal').style.display = 'flex';
      }}

      function closeColSettings() {{
        document.getElementById('colSettingsModal').style.display = 'none';
      }}

      function applyColSettings() {{
        const settings = {{}};
        document.querySelectorAll('.col-setting-check').forEach(cb => {{
          settings[cb.dataset.key] = cb.checked;
        }});
        localStorage.setItem(PT_STORAGE_KEY, JSON.stringify(settings));
        applyColVisibility();
        closeColSettings();
      }}

      function sortPt(col) {{
        const params = new URLSearchParams(window.location.search);
        const curSort = params.get('sort_by') || 'deviation_abs';
        const curDir = params.get('sort_dir') || 'desc';
        const newDir = (curSort === col && curDir === 'asc') ? 'desc' : 'asc';
        params.set('sort_by', col);
        params.set('sort_dir', newDir);
        window.location.search = params.toString();
      }}

      function togglePtDetail(detailId) {{
        var el = document.getElementById(detailId);
        var arrow = document.getElementById(detailId + '_arrow');
        if (!el) return;
        var isOpen = el.style.display !== 'none';
        el.style.display = isOpen ? 'none' : 'table-row';
        if (arrow) arrow.innerText = isOpen ? '▼' : '▲';
      }}

      applyColVisibility();
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.get("/master/purchase-tracking/status/export")
async def purchase_tracking_status_export(
    session_token: str = Cookie(default=None),
    filter_branch: str = "",
    search_item: str = "",
    filter_date: str = "",
    filter_deviation: str = ""
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    latest_week_row = conn.execute(
        "SELECT MAX(snapshot_week) as w FROM purchase_tracking_snapshot"
    ).fetchone()
    latest_week = latest_week_row["w"] if latest_week_row else None

    if not latest_week:
        conn.close()
        return RedirectResponse(url="/master/purchase-tracking/status", status_code=303)

    query = "SELECT * FROM purchase_tracking_snapshot WHERE snapshot_week=?"
    params: list = [latest_week]
    if filter_branch:
        query += " AND branch_name=?"
        params.append(filter_branch)
    if search_item:
        query += " AND item_name LIKE ?"
        params.append(f"%{search_item}%")
    if filter_date:
        query += " AND last_purchase_date LIKE ?"
        params.append(f"{filter_date}%")
    if filter_deviation == "over":
        query += " AND deviation_days > 3"
    elif filter_deviation == "delay":
        query += " AND deviation_days < -3"
    elif filter_deviation == "ok":
        query += " AND deviation_days BETWEEN -3 AND 3"
    query += " ORDER BY ABS(deviation_days) DESC"

    rows = conn.execute(query, params).fetchall()

    branch_name_to_code = {b["branch_name"]: b["branch_code"] for b in get_branches()}

    prev_purchase_map = {}
    for r in rows:
        key = (r["branch_name"], r["item_name"])
        if key in prev_purchase_map:
            continue
        hist = conn.execute("""
            SELECT purchase_datetime, quantity FROM purchase_records
            WHERE branch_name=? AND item_name=?
            ORDER BY purchase_datetime DESC LIMIT 2
        """, (r["branch_name"], r["item_name"])).fetchall()
        prev_purchase_map[key] = hist

    raw_stock_map = {}
    raw_rows = conn.execute("SELECT branch_code, item_name, quantity FROM raw_inventory").fetchall()
    for rr in raw_rows:
        raw_stock_map[(rr["branch_code"], rr["item_name"])] = rr["quantity"]

    conn.close()

    import io
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "발주추천현황"
    headers = ["지점", "상품명", "마지막구매일", "마지막수량", "직전구매일", "직전수량",
               "실제주기(A)", "목표주기(B)", "편차", "추천발주일", "현재고", "추천수량"]
    ws.append(headers)

    for r in rows:
        key = (r["branch_name"], r["item_name"])
        hist = prev_purchase_map.get(key, [])
        last_qty = round(hist[0]["quantity"]) if len(hist) >= 1 else None
        prev_date = hist[1]["purchase_datetime"] if len(hist) >= 2 else None
        prev_qty = round(hist[1]["quantity"]) if len(hist) >= 2 else None

        branch_code = branch_name_to_code.get(r["branch_name"])
        current_stock = raw_stock_map.get((branch_code, r["item_name"])) if branch_code else None
        current_stock = round(current_stock) if current_stock is not None else None

        raw_qty = r["recommended_qty"]
        final_qty = 5 if (raw_qty is not None and raw_qty < 5) else (round(raw_qty) if raw_qty is not None else None)

        ws.append([
            r["branch_name"], r["item_name"], r["last_purchase_date"], last_qty,
            prev_date, prev_qty, r["actual_interval_days"], r["lead_time_days"],
            r["deviation_days"], r["recommended_next_date"], current_stock, final_qty
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"발주추천현황_{latest_week}.xlsx"
    from urllib.parse import quote
    encoded_filename = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )

@app.get("/master/purchase-tracking/history", response_class=HTMLResponse)
async def purchase_tracking_history_page(
    session_token: str = Cookie(default=None),
    branch: str = "",
    item: str = ""
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()

    if not branch or not item:
        combo_rows = conn.execute("""
            SELECT DISTINCT branch_name, item_name FROM purchase_tracking_snapshot
            ORDER BY branch_name, item_name LIMIT 1000
        """).fetchall()
        conn.close()

        combo_options = ""
        for c in combo_rows:
            combo_options += f'<option value="{c["branch_name"]}|{c["item_name"]}">{c["branch_name"]} / {c["item_name"]}</option>'

        content = f"""
        <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
          <a href="/master/purchase-tracking" style="color:#1E2761;">← 발주 주기 트래킹</a>
          <h2>📈 편차 추이 히스토리</h2>
        </div>
        <div class="card">
          <label style="font-size:13px;color:#555;">지점 / 상품 선택</label>
          <select id="comboSelect" style="margin-top:4px;">
            <option value="">-- 선택하세요 --</option>
            {combo_options}
          </select>
          <button class="btn" style="margin-top:12px;" onclick="goHistory()">조회</button>
        </div>
        <script>
          function goHistory() {{
            const val = document.getElementById('comboSelect').value;
            if (!val) {{ alert('지점/상품을 선택하세요.'); return; }}
            const [branch, item] = val.split('|');
            window.location.href = '/master/purchase-tracking/history?branch=' + encodeURIComponent(branch) + '&item=' + encodeURIComponent(item);
          }}
        </script>
        """
        return HTMLResponse(content=render_page(content, user, "master"))

    history_rows = conn.execute("""
        SELECT * FROM purchase_tracking_snapshot
        WHERE branch_name=? AND item_name=?
        ORDER BY snapshot_week ASC
    """, (branch, item)).fetchall()
    conn.close()

    rows_html = ""
    if not history_rows:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:20px;color:#888;">기록 없음</td></tr>'
    else:
        for r in history_rows:
            dev = r["deviation_days"]
            dev_color = "#22C55E" if (dev is not None and abs(dev) <= 3) else ("#F59E0B" if (dev is not None and dev > 3) else "#EF4444")
            dev_display = f"{dev:+d}" if dev is not None else "-"
            qty_val = r["recommended_qty"]
            qty_display = f"{qty_val:g}" if qty_val is not None else "-"
            rows_html += f"""
            <tr>
              <td>{r['snapshot_week']}</td>
              <td style="font-size:12px;">{(r['last_purchase_date'] or '-')[:16]}</td>
              <td style="text-align:center;">{r['actual_interval_days']}일</td>
              <td style="text-align:center;">{r['lead_time_days']}일</td>
              <td style="color:{dev_color};font-weight:bold;">{dev_display}</td>
              <td style="text-align:right;">{qty_display}</td>
            </tr>
            """

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master/purchase-tracking/history" style="color:#1E2761;">← 다른 상품 선택</a>
      <h2>📈 편차 추이 히스토리</h2>
    </div>
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;"><b>{branch}</b> / <b>{item}</b> — 주차별 편차가 0에 가까워질수록 발주 주기가 목표에 맞게 개선되고 있는 것입니다.</p>
    </div>
    <div class="card">
      <table>
        <thead><tr>
          <th>주차</th><th>마지막구매일</th><th>실제주기(A)</th><th>목표주기(B)</th><th>편차</th><th>추천수량</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))

@app.get("/master/purchase-order/branch-exceptions", response_class=HTMLResponse)
async def purchase_order_branch_exceptions_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    conn = get_conn()
    exception_rows = conn.execute(
        "SELECT branch_code FROM purchase_order_branch_exception WHERE exception_type='consumable_include'"
    ).fetchall()
    conn.close()

    exception_set = {r["branch_code"] for r in exception_rows}

    branches = get_branches(branch_type='branch')

    rows_html = ""
    for b in branches:
        consumable_checked = "checked" if b["branch_code"] in exception_set else ""
        rows_html += f"""
        <tr>
          <td>{b['branch_name']}</td>
          <td style="text-align:center;">
            <input type="checkbox" class="exc-check" data-branch="{b['branch_code']}" data-type="consumable_include" {consumable_checked} style="width:18px;height:18px;">
          </td>
        </tr>
        """

    content = f"""
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>🏬 발주서 지점 예외 설정</h2>
    </div>
    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">
        <b>소모품 포함</b>: 체크된 지점은 소모품도 발주 목록에서 제외하지 않고 포함시킵니다.<br>
        체크박스를 클릭하면 즉시 저장됩니다.<br>
        <b>월 1회 예외(지점+상품별 개별 지정)</b>는 <a href="/master/purchase-order/product-settings" style="color:#1E40AF;text-decoration:underline;">상품 설정 페이지</a>에서 상품별로 지정합니다.
      </p>
    </div>
    <div class="card">
      <table>
        <thead><tr>
          <th>지점</th><th style="text-align:center;">소모품 포함</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    <script>
      document.querySelectorAll('.exc-check').forEach(function(cb) {{
        cb.addEventListener('change', async function() {{
          const branchCode = cb.dataset.branch;
          const excType = cb.dataset.type;
          const enabled = cb.checked;
          try {{
            const res = await fetch('/master/purchase-order/branch-exceptions/toggle', {{
              method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
              body: JSON.stringify({{ branch_code: branchCode, exception_type: excType, enabled: enabled }})
            }});
            if (!res.ok) {{
              alert('저장 실패');
              cb.checked = !enabled;
            }}
          }} catch(e) {{
            alert('저장 중 오류가 발생했습니다.');
            cb.checked = !enabled;
          }}
        }});
      }});
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/purchase-order/branch-exceptions/toggle")
async def purchase_order_branch_exceptions_toggle(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    branch_code = data.get("branch_code", "").strip()
    exception_type = data.get("exception_type", "").strip()
    enabled = bool(data.get("enabled", False))

    if not branch_code or exception_type != "consumable_include":
        return JSONResponse(status_code=400, content={"detail": "잘못된 요청입니다."})

    conn = get_conn()
    if enabled:
        existing = conn.execute(
            "SELECT id FROM purchase_order_branch_exception WHERE branch_code=? AND exception_type=?",
            (branch_code, exception_type)
        ).fetchone()
        if not existing:
            conn.execute(
                "INSERT INTO purchase_order_branch_exception (branch_code, exception_type, created_at) VALUES (?, ?, ?)",
                (branch_code, exception_type, datetime.now().isoformat())
            )
    else:
        conn.execute(
            "DELETE FROM purchase_order_branch_exception WHERE branch_code=? AND exception_type=?",
            (branch_code, exception_type)
        )
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/raw-upload/ajax")
async def raw_upload_ajax(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    user = get_session(session_token)
    if not user:
        return {"success": 0, "skipped": 0, "errors": ["로그인이 필요합니다"]}
    return await _process_raw_upload_master(file)

async def _fetch_and_process_s3_csv():
    """S3의 유비플러스 재고현황 CSV를 다운로드하여 재고수불부 형식과 동일하게 파싱/반영"""
    import httpx
    import csv
    import io as io_module

    CSV_URL = "https://petdoc-ubiplus.s3.ap-northeast-2.amazonaws.com/stock/유비플러스_재고현황.csv"

    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.get(CSV_URL)
        resp.raise_for_status()
        raw_bytes = resp.content

    # 인코딩 자동 판별 (한글 CSV는 보통 cp949 또는 utf-8-sig)
    try:
        text = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw_bytes.decode('cp949')

    reader = csv.reader(io_module.StringIO(text))
    rows_raw = list(reader)

    header_row_idx = None
    col_map = {}
    KEYWORDS = {
        "branch": ["지점"],
        "item_name": ["상품명"],
        "item_code": ["품번"],
        "qty": ["기말수량"],
        "h": ["증가수량"],
        "q": ["재고조정"],
    }
    for row_idx in range(min(5, len(rows_raw))):
        row_vals = rows_raw[row_idx]
        found = {}
        for col_idx, cell_val in enumerate(row_vals):
            if not cell_val:
                continue
            text_val = str(cell_val).strip()
            for key, keywords in KEYWORDS.items():
                if key in found:
                    continue
                if any(kw in text_val for kw in keywords):
                    found[key] = col_idx
        if all(k in found for k in ("branch", "item_name", "item_code")):
            header_row_idx = row_idx
            col_map = found
            break

    if header_row_idx is None:
        return {"success": 0, "skipped": 0, "errors": ["CSV 헤더를 찾을 수 없습니다."]}

    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []
    hq_adjustments = []
    debug_hq_log = []
    data_start_row = header_row_idx + 1

    branches_in_file = set()
    parsed_rows = []
    for idx in range(data_start_row, len(rows_raw)):
        row = rows_raw[idx]
        branch_col = col_map.get("branch")
        if branch_col is None or branch_col >= len(row) or not row[branch_col]:
            continue
        branch_name = str(row[branch_col]).strip()
        branch_code = (branch_map.get(branch_name)
                       or branch_map.get(branch_name.replace(" ", ""))
                       or branch_name)
        branches_in_file.add(branch_code)
        parsed_rows.append((idx, row, branch_code, branch_name))

    if not branches_in_file:
        return {"success": 0, "skipped": 0, "errors": ["CSV에서 유효한 지점 데이터를 찾지 못했습니다."]}

    conn = get_conn()
    for bc in branches_in_file:
        conn.execute("DELETE FROM raw_inventory WHERE branch_code=?", (bc,))
    old_hq_rows = conn.execute("SELECT * FROM hq_bonus_log").fetchall()
    old_hq_map = {f"{r['branch_code']}|{r['item_code']}": r["last_hq_total"] for r in old_hq_rows}
    for bc in branches_in_file:
        conn.execute("DELETE FROM hq_bonus_log WHERE branch_code=?", (bc,))
    conn.commit()

    for idx, row, branch_code, branch_name in parsed_rows:
        try:
            item_name = str(row[col_map["item_name"]]).strip() if col_map["item_name"] < len(row) and row[col_map["item_name"]] else ""
            item_code = str(row[col_map["item_code"]]).strip() if col_map["item_code"] < len(row) and row[col_map["item_code"]] else ""

            if not item_code:
                import hashlib
                name_hash = hashlib.md5(item_name.encode('utf-8')).hexdigest()[:8]
                item_code = f"미지정_{name_hash}"

            qty_n = row[col_map["qty"]] if "qty" in col_map and col_map["qty"] < len(row) else None
            qty_h = row[col_map["h"]] if "h" in col_map and col_map["h"] < len(row) else None
            qty_q = row[col_map["q"]] if "q" in col_map and col_map["q"] < len(row) else None

            raw_quantity = int(float(str(qty_n))) if qty_n not in (None, "") else 0
            add_h = int(float(str(qty_h))) if qty_h not in (None, "") else 0
            add_q = int(float(str(qty_q))) if qty_q not in (None, "") else 0
            hq_total = add_h + add_q

            conn.execute("""
                INSERT INTO raw_inventory
                  (branch_code, branch_name, item_name, item_code, quantity, source, uploaded_at)
                VALUES (?, ?, ?, ?, ?, 'auto_s3', ?)
                ON CONFLICT(branch_code, item_code) DO UPDATE SET
                  quantity=excluded.quantity,
                  item_name=excluded.item_name,
                  branch_name=excluded.branch_name,
                  source=excluded.source,
                  uploaded_at=excluded.uploaded_at
            """, (branch_code, branch_name, item_name, item_code, raw_quantity, now))

            existing_item = conn.execute(
                "SELECT id FROM items WHERE branch_code=? AND item_code=?",
                (branch_code, item_code)
            ).fetchone()
            if not existing_item:
                conn.execute("""
                    INSERT INTO items (branch_code, branch_name, item_name, item_code, created_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(branch_code, item_code) DO UPDATE SET item_name=excluded.item_name
                """, (branch_code, branch_name, item_name, item_code, now))
                conn.execute("""
                    INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated)
                    VALUES (?, ?, ?, 0, ?)
                    ON CONFLICT(branch_code, item_code) DO NOTHING
                """, (branch_code, item_name, item_code, now))

            if item_code.startswith("미지정_"):
                conn.execute("""
                    UPDATE inventory SET quantity=?, last_updated=?
                    WHERE branch_code=? AND item_code=?
                """, (raw_quantity, now, branch_code, item_code))

            if hq_total != 0:
                hq_adjustments.append((branch_code, item_name, item_code, hq_total))
                debug_hq_log.append(f"{item_name}({item_code}): 증가={add_h}, 조정={add_q}, 합계={hq_total}")

            success += 1
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()
    conn.close()

    for branch_code, item_name, item_code, new_hq_total in hq_adjustments:
        conn2 = get_conn()
        prev_hq_total = old_hq_map.get(f"{branch_code}|{item_code}", 0)
        net_delta = new_hq_total - prev_hq_total
        if net_delta != 0:
            new_qty = adjust_quantity(branch_code, item_code, net_delta, absolute=False)
            conn2.execute(
                "INSERT INTO adjustment_log (branch_code, item_name, item_code, delta, result_quantity, adjusted_at) VALUES (?, ?, ?, ?, ?, ?)",
                (branch_code, item_name, item_code, net_delta, new_qty, now)
            )
        conn2.execute("""
            INSERT INTO hq_bonus_log (branch_code, item_code, last_hq_total, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(branch_code, item_code) DO UPDATE SET
              last_hq_total=excluded.last_hq_total, updated_at=excluded.updated_at
        """, (branch_code, item_code, new_hq_total, now))
        conn2.commit()
        conn2.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10],
            "header_row_used": header_row_idx,
            "hq_debug": debug_hq_log[:20],
            "synced_at": now}

@app.get("/api/cron/compute-purchase-tracking-snapshot")
async def cron_compute_purchase_tracking_snapshot(authorization: str = Header(default="")):
    expected = f"Bearer {os.environ.get('CRON_SECRET', '')}"
    if authorization != expected:
        return JSONResponse(status_code=401, content={"detail": "인증 실패"})

    result = await _compute_purchase_tracking_snapshot()
    return JSONResponse(content=result)

@app.get("/api/cron/sync-raw-inventory")
async def cron_sync_raw_inventory(request: Request):
    # Vercel Cron 요청 검증 (Authorization 헤더로 비인가 접근 차단)
    auth_header = request.headers.get("authorization", "")
    cron_secret = os.environ.get("CRON_SECRET", "")
    if cron_secret and auth_header != f"Bearer {cron_secret}":
        return JSONResponse(status_code=401, content={"detail": "unauthorized"})

    result = await _fetch_and_process_s3_csv()
    return JSONResponse(content=result)

async def _process_purchase_records_upload(file: UploadFile):
    """구매내역 엑셀 업로드 처리 — 헤더 2행, dedup UPSERT, 3개월 이전 데이터 자동 삭제"""
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return {"success": 0, "skipped": 0, "errors": ["시트를 찾을 수 없습니다."]}

    header_row = None
    for row_idx in range(1, 4):
        row_vals = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
        if not row_vals:
            continue
        if row_vals[0] and "구매일시" in str(row_vals[0]):
            header_row = row_idx
            break

    if header_row is None:
        return {"success": 0, "skipped": 0, "errors": ["'구매일시' 헤더를 찾을 수 없습니다. 파일 형식을 확인해주세요."]}

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []

    conn = get_conn()
    data_start_row = header_row + 1

    for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        if not row[0]:
            continue
        try:
            purchase_datetime = str(row[0]).strip()
            vendor = str(row[2]).strip() if row[2] else ""
            branch_name = str(row[3]).strip() if row[3] else ""
            item_name = str(row[5]).strip() if row[5] else ""
            item_code = str(row[6]).strip() if row[6] else ""
            quantity = float(row[7]) if row[7] is not None else 0
            unit_price = float(row[8]) if row[8] is not None else None
            total_price = float(row[9]) if row[9] is not None else None

            if not branch_name or not item_name:
                skipped += 1
                continue

            conn.execute("""
                INSERT INTO purchase_records
                    (purchase_datetime, branch_name, vendor, item_name, item_code, quantity, unit_price, total_price, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (purchase_datetime, branch_name, item_name, item_code, quantity)
                DO UPDATE SET
                    vendor=excluded.vendor,
                    unit_price=excluded.unit_price,
                    total_price=excluded.total_price,
                    created_at=excluded.created_at
            """, (purchase_datetime, branch_name, vendor, item_name, item_code, quantity, unit_price, total_price, now))
            success += 1
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()

    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(days=90)).isoformat()
    conn.execute("DELETE FROM purchase_records WHERE purchase_datetime < ?", (cutoff,))
    conn.commit()
    conn.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10], "header_row_used": header_row}

def _get_iso_week_str(dt: datetime) -> str:
    """ISO 주차 문자열 반환 (예: '2026-W32')"""
    iso_year, iso_week, _ = dt.isocalendar()
    return f"{iso_year}-W{iso_week:02d}"


async def _compute_purchase_tracking_snapshot():
    """지점×상품명별 A(실제간격) vs B(리드타임) 비교, 주간 스냅샷 계산 및 저장"""
    conn = get_conn()

    leadtime_rows = conn.execute("SELECT item_name, lead_time_days FROM product_lead_time").fetchall()
    leadtime_map = {r["item_name"]: r["lead_time_days"] for r in leadtime_rows}

    if not leadtime_map:
        conn.close()
        return {"processed": 0, "skipped_no_leadtime": 0, "message": "등록된 리드타임이 없습니다."}

    combos = conn.execute("""
        SELECT DISTINCT branch_name, item_name FROM purchase_records
    """).fetchall()

    now = datetime.now()
    snapshot_week = _get_iso_week_str(now)
    now_iso = now.isoformat()

    processed = 0
    skipped_no_leadtime = 0

    for combo in combos:
        branch_name = combo["branch_name"]
        item_name = combo["item_name"]

        lead_time_days = leadtime_map.get(item_name)
        if lead_time_days is None:
            skipped_no_leadtime += 1
            continue

        history = conn.execute("""
            SELECT purchase_datetime, quantity FROM purchase_records
            WHERE branch_name = ? AND item_name = ?
            ORDER BY purchase_datetime DESC
            LIMIT 2
        """, (branch_name, item_name)).fetchall()

        if not history:
            continue

        last_purchase_date = history[0]["purchase_datetime"]
        last_qty = float(history[0]["quantity"])

        if len(history) >= 2:
            try:
                d1 = datetime.fromisoformat(history[0]["purchase_datetime"])
                d2 = datetime.fromisoformat(history[1]["purchase_datetime"])
                actual_interval_days = abs((d1 - d2).days)
                if actual_interval_days == 0:
                    actual_interval_days = 1
            except Exception:
                actual_interval_days = lead_time_days
        else:
            actual_interval_days = lead_time_days

        deviation_days = lead_time_days - actual_interval_days

        try:
            recommended_qty = round(last_qty * (lead_time_days / actual_interval_days), 1)
        except ZeroDivisionError:
            recommended_qty = last_qty

        try:
            last_dt = datetime.fromisoformat(last_purchase_date)
            from datetime import timedelta
            recommended_next_date = (last_dt + timedelta(days=lead_time_days)).date().isoformat()
        except Exception:
            recommended_next_date = None

        conn.execute("""
            INSERT INTO purchase_tracking_snapshot
                (snapshot_week, branch_name, item_name, last_purchase_date, actual_interval_days,
                 lead_time_days, deviation_days, recommended_next_date, recommended_qty, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT (snapshot_week, branch_name, item_name) DO UPDATE SET
                last_purchase_date=excluded.last_purchase_date,
                actual_interval_days=excluded.actual_interval_days,
                lead_time_days=excluded.lead_time_days,
                deviation_days=excluded.deviation_days,
                recommended_next_date=excluded.recommended_next_date,
                recommended_qty=excluded.recommended_qty,
                created_at=excluded.created_at
        """, (snapshot_week, branch_name, item_name, last_purchase_date, actual_interval_days,
              lead_time_days, deviation_days, recommended_next_date, recommended_qty, now_iso))
        processed += 1

    conn.commit()
    conn.close()

    return {
        "processed": processed,
        "skipped_no_leadtime": skipped_no_leadtime,
        "snapshot_week": snapshot_week
    }

async def _process_lead_time_upload(file: UploadFile):
    """상품별 리드타임/MOQ(단위)/소모품여부 엑셀 업로드 처리 — 헤더 2행, 상품명 기준 UPSERT
    ⚠️ 2026-08-07 확장: 기존 리드타임 전용 업로드에 단위(MOQ)/소모품종류 파싱 추가 (같은 상품내역 파일 재사용)
    """
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return {"success": 0, "skipped": 0, "errors": ["시트를 찾을 수 없습니다."]}

    header_row = None
    col_map = {}
    for row_idx in range(1, 4):
        row_vals = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
        if not row_vals:
            continue
        found = {}
        for col_idx, cell_val in enumerate(row_vals):
            if not cell_val:
                continue
            text = str(cell_val).strip()
            if text == "상품명" and "item_name" not in found:
                found["item_name"] = col_idx
            if text == "리드타임" and "lead_time" not in found:
                found["lead_time"] = col_idx
            if text == "단위" and "moq" not in found:
                found["moq"] = col_idx
            if text == "소모품종류" and "consumable" not in found:
                found["consumable"] = col_idx
        if "item_name" in found and "lead_time" in found:
            header_row = row_idx
            col_map = found
            break

    if header_row is None:
        return {"success": 0, "skipped": 0, "errors": ["'상품명' 또는 '리드타임' 헤더를 찾을 수 없습니다."]}

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []
    data_start_row = header_row + 1

    conn = get_conn()
    for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        try:
            item_name_col = col_map["item_name"]
            lead_time_col = col_map["lead_time"]
            if item_name_col >= len(row) or not row[item_name_col]:
                continue
            item_name = str(row[item_name_col]).strip()

            lead_time_raw = row[lead_time_col] if lead_time_col < len(row) else None
            if lead_time_raw is None or str(lead_time_raw).strip() in ("", "0"):
                skipped += 1
                continue

            lead_time_days = int(float(str(lead_time_raw)))

            moq = 1
            if "moq" in col_map:
                moq_col = col_map["moq"]
                if moq_col < len(row) and row[moq_col] not in (None, ""):
                    try:
                        moq = int(float(str(row[moq_col])))
                        if moq <= 0:
                            moq = 1
                    except Exception:
                        moq = 1

            is_consumable = False
            if "consumable" in col_map:
                cons_col = col_map["consumable"]
                if cons_col < len(row) and row[cons_col] not in (None, ""):
                    is_consumable = True

            conn.execute("""
                INSERT INTO product_lead_time (item_name, lead_time_days, moq, is_consumable, updated_at)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT (item_name) DO UPDATE SET
                    lead_time_days=excluded.lead_time_days,
                    moq=excluded.moq,
                    is_consumable=excluded.is_consumable,
                    updated_at=excluded.updated_at
            """, (item_name, lead_time_days, moq, is_consumable, now))
            success += 1
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()
    conn.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10], "header_row_used": header_row,
            "col_map_debug": {k: v for k, v in col_map.items()}}

async def _fetch_and_process_purchase_records_csv():
    """구매내역 CSV를 외부 소스(URL)에서 다운로드하여 자동 반영. URL은 PURCHASE_RECORDS_CSV_URL 환경변수로 지정."""
    import httpx
    import csv
    import io as io_module

    csv_url = os.environ.get("PURCHASE_RECORDS_CSV_URL", "")
    if not csv_url:
        return {"success": 0, "skipped": 0, "errors": ["PURCHASE_RECORDS_CSV_URL 환경변수가 설정되지 않았습니다."]}

    async with httpx.AsyncClient(timeout=30.0) as client:
        resp = await client.get(csv_url)
        resp.raise_for_status()
        raw_bytes = resp.content

    try:
        text = raw_bytes.decode('utf-8-sig')
    except UnicodeDecodeError:
        text = raw_bytes.decode('cp949')

    reader = csv.reader(io_module.StringIO(text))
    rows_raw = list(reader)

    header_row_idx = None
    for row_idx in range(min(5, len(rows_raw))):
        row_vals = rows_raw[row_idx]
        if row_vals and row_vals[0] and "구매일시" in str(row_vals[0]):
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return {"success": 0, "skipped": 0, "errors": ["CSV에서 '구매일시' 헤더를 찾을 수 없습니다."]}

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []
    data_start_row = header_row_idx + 1

    conn = get_conn()
    for idx in range(data_start_row, len(rows_raw)):
        row = rows_raw[idx]
        if not row or not row[0]:
            continue
        try:
            purchase_datetime = str(row[0]).strip()
            vendor = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            branch_name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            item_name = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            item_code = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            quantity = float(row[7]) if len(row) > 7 and row[7] not in (None, "") else 0
            unit_price = float(row[8]) if len(row) > 8 and row[8] not in (None, "") else None
            total_price = float(row[9]) if len(row) > 9 and row[9] not in (None, "") else None

            if not branch_name or not item_name:
                skipped += 1
                continue

            conn.execute("""
                INSERT INTO purchase_records
                    (purchase_datetime, branch_name, vendor, item_name, item_code, quantity, unit_price, total_price, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT (purchase_datetime, branch_name, item_name, item_code, quantity)
                DO UPDATE SET
                    vendor=excluded.vendor,
                    unit_price=excluded.unit_price,
                    total_price=excluded.total_price,
                    created_at=excluded.created_at
            """, (purchase_datetime, branch_name, vendor, item_name, item_code, quantity, unit_price, total_price, now))
            success += 1
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()

    from datetime import timedelta
    cutoff = (datetime.now() - timedelta(days=90)).isoformat()
    conn.execute("DELETE FROM purchase_records WHERE purchase_datetime < ?", (cutoff,))
    conn.commit()
    conn.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10], "synced_at": now}


@app.get("/api/cron/sync-purchase-records")
async def cron_sync_purchase_records(request: Request):
    auth_header = request.headers.get("authorization", "")
    cron_secret = os.environ.get("CRON_SECRET", "")
    if cron_secret and auth_header != f"Bearer {cron_secret}":
        return JSONResponse(status_code=401, content={"detail": "unauthorized"})

    result = await _fetch_and_process_purchase_records_csv()
    return JSONResponse(content=result)

async def _process_raw_upload_master(file: UploadFile):
    """마스터 전용 — 헤더가 2행에 있고 컬럼명이 다른 '재고수불부' 형식 처리
    ⚠️ 2026-07 수정: 업로드된 엑셀에 실제로 존재하는 지점만 삭제/갱신하도록 변경
       (기존 버그: source='master' 전체를 지워서 다른 지점 데이터가 0으로 변함)
    """
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return {"success": 0, "skipped": 0, "errors": ["시트를 찾을 수 없습니다."]}

    header_row_idx = None
    col_map = {}
    KEYWORDS = {
        "branch": ["지점"],
        "item_name": ["상품명"],
        "item_code": ["품번"],
        "qty": ["기말수량"],
        "h": ["증가수량"],
        "q": ["재고조정"],
    }
    for row_idx in range(1, 6):
        row_vals = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
        if not row_vals:
            continue
        found = {}
        for col_idx, cell_val in enumerate(row_vals):
            if not cell_val:
                continue
            text = str(cell_val).strip()
            for key, keywords in KEYWORDS.items():
                if key in found:
                    continue
                if any(kw in text for kw in keywords):
                    found[key] = col_idx
        if all(k in found for k in ("branch", "item_name", "item_code")):
            header_row_idx = row_idx
            col_map = found
            break

    if header_row_idx is None:
        return {"success": 0, "skipped": 0,
                "errors": ["헤더를 찾을 수 없습니다. '지점/상품명/품번' 컬럼명이 포함된 행이 있는지 확인해주세요."]}

    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []
    hq_adjustments = []
    debug_hq_log = []
    data_start_row = header_row_idx + 1

    # ── ⚠️ 1단계: 삭제 전에 먼저 "이번 엑셀에 실제로 어떤 지점이 있는지" 수집 ──
    branches_in_file = set()
    parsed_rows = []
    for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        branch_col = col_map.get("branch")
        if branch_col is None or branch_col >= len(row) or not row[branch_col]:
            continue
        branch_name = str(row[branch_col]).strip()
        branch_code = (branch_map.get(branch_name)
                       or branch_map.get(branch_name.replace(" ", ""))
                       or branch_name)
        branches_in_file.add(branch_code)
        parsed_rows.append((idx, row, branch_code, branch_name))

    # ── ⚠️ 빈 데이터 방어: 엑셀에서 유효한 지점 행을 하나도 못 찾으면 삭제 자체를 하지 않고 즉시 중단 ──
    if not branches_in_file:
        return {"success": 0, "skipped": 0,
                "errors": ["엑셀에서 유효한 지점 데이터를 찾지 못했습니다. 기존 데이터는 보존되었으며 아무 것도 변경되지 않았습니다."],
                "header_row_used": header_row_idx,
                "col_map_debug": {k: v for k, v in col_map.items()}}

    conn = get_conn()
    # ── source 무관, 이번 엑셀에 있는 지점의 RAW 전체를 삭제 후 재삽입 (최신 업로드가 항상 기준) ──
    for bc in branches_in_file:
        conn.execute("DELETE FROM raw_inventory WHERE branch_code=?", (bc,))
    old_hq_rows = conn.execute("SELECT * FROM hq_bonus_log").fetchall()
    old_hq_map = {f"{r['branch_code']}|{r['item_code']}": r["last_hq_total"] for r in old_hq_rows}
    # ── ⚠️ hq_bonus_log도 전체 삭제 대신 해당 지점만 삭제 ──
    for bc in branches_in_file:
        conn.execute("DELETE FROM hq_bonus_log WHERE branch_code=?", (bc,))
    conn.commit()

    for idx, row, branch_code, branch_name in parsed_rows:
        try:
            item_name = str(row[col_map["item_name"]]).strip() if row[col_map["item_name"]] else ""
            item_code = str(row[col_map["item_code"]]).strip() if row[col_map["item_code"]] else ""

            if not item_code:
                # 품번 없는 상품: 지점+상품명 조합으로 안전한 고유 코드 생성 (50자 자르기로 인한 충돌 방지)
                import hashlib
                name_hash = hashlib.md5(item_name.encode('utf-8')).hexdigest()[:8]
                item_code = f"미지정_{name_hash}"

            qty_n = row[col_map["qty"]] if "qty" in col_map and col_map["qty"] < len(row) else None
            qty_h = row[col_map["h"]] if "h" in col_map and col_map["h"] < len(row) else None
            qty_q = row[col_map["q"]] if "q" in col_map and col_map["q"] < len(row) else None

            raw_quantity = int(float(str(qty_n))) if qty_n not in (None, "") else 0
            add_h = int(float(str(qty_h))) if qty_h not in (None, "") else 0
            add_q = int(float(str(qty_q))) if qty_q not in (None, "") else 0
            hq_total = add_h + add_q

            conn.execute("""
                INSERT INTO raw_inventory
                  (branch_code, branch_name, item_name, item_code, quantity, source, uploaded_at)
                VALUES (?, ?, ?, ?, ?, 'master', ?)
                ON CONFLICT(branch_code, item_code) DO UPDATE SET
                  quantity=excluded.quantity,
                  item_name=excluded.item_name,
                  branch_name=excluded.branch_name,
                  source=excluded.source,
                  uploaded_at=excluded.uploaded_at
            """, (branch_code, branch_name, item_name, item_code, raw_quantity, now))

            # ⚠️ 신규 상품 자동 등록: items에 없으면 items + inventory(초기수량 0)에 추가
            existing_item = conn.execute(
                "SELECT id FROM items WHERE branch_code=? AND item_code=?",
                (branch_code, item_code)
            ).fetchone()
            if not existing_item:
                conn.execute("""
                    INSERT INTO items (branch_code, branch_name, item_name, item_code, created_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(branch_code, item_code) DO UPDATE SET item_name=excluded.item_name
                """, (branch_code, branch_name, item_name, item_code, now))
                conn.execute("""
                    INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated)
                    VALUES (?, ?, ?, 0, ?)
                    ON CONFLICT(branch_code, item_code) DO NOTHING
                """, (branch_code, item_name, item_code, now))

            # ⚠️ 품번 없는 상품(미지정_): 업로드할 때마다 QR재고를 RAW재고와 항상 강제 동기화
            if item_code.startswith("미지정_"):
                conn.execute("""
                    UPDATE inventory SET quantity=?, last_updated=?
                    WHERE branch_code=? AND item_code=?
                """, (raw_quantity, now, branch_code, item_code))

            if hq_total != 0:
                hq_adjustments.append((branch_code, item_name, item_code, hq_total))
                debug_hq_log.append(f"{item_name}({item_code}): 증가={add_h}, 조정={add_q}, 합계={hq_total}")

            success += 1
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()
    conn.close()

    for branch_code, item_name, item_code, new_hq_total in hq_adjustments:
        conn2 = get_conn()
        prev_hq_total = old_hq_map.get(f"{branch_code}|{item_code}", 0)
        net_delta = new_hq_total - prev_hq_total
        if net_delta != 0:
            new_qty = adjust_quantity(branch_code, item_code, net_delta, absolute=False)
            conn2.execute(
                "INSERT INTO adjustment_log (branch_code, item_name, item_code, delta, result_quantity, adjusted_at) VALUES (?, ?, ?, ?, ?, ?)",
                (branch_code, item_name, item_code, net_delta, new_qty, now)
            )
        conn2.execute("""
            INSERT INTO hq_bonus_log (branch_code, item_code, last_hq_total, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(branch_code, item_code) DO UPDATE SET
              last_hq_total=excluded.last_hq_total, updated_at=excluded.updated_at
        """, (branch_code, item_code, new_hq_total, now))
        conn2.commit()
        conn2.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10],
            "header_row_used": header_row_idx,
            "hq_debug": debug_hq_log[:20],
            "col_map_debug": {k: v for k, v in col_map.items()}}


async def _process_raw_upload(file: UploadFile, restrict_branch: Optional[str] = None):
    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return {"success": 0, "skipped": 0, "errors": ["시트를 찾을 수 없습니다."]}

    # ── 헤더 행 자동 탐색 (1행 또는 2행, 병합 셀 대응) ──
    # 최대 5행까지 훑어서 "품번" 또는 "현재수량" 텍스트가 있는 행을 헤더로 인식
    header_row_idx = None
    col_map = {}  # {"branch":0, "item_name":1, "item_code":3, "qty":13, "h":7, "q":16}

    KEYWORDS = {
        "branch": ["지점", "지점명"],
        "item_name": ["상품명", "품명"],
        "item_code": ["품번", "품목코드"],
        "qty": ["현재고", "기말수량"],
        "h": ["H", "증가수량"],
        "q": ["Q", "재고조정"],
    }

    for row_idx in range(1, 6):
        row_vals = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), None)
        if not row_vals:
            continue
        found = {}
        for col_idx, cell_val in enumerate(row_vals):
            if not cell_val:
                continue
            text = str(cell_val).strip()
            for key, keywords in KEYWORDS.items():
                if key in found:
                    continue
                if any(kw in text for kw in keywords):
                    found[key] = col_idx
        # 최소한 지점/상품명/품번 세 개는 찾아야 이 행을 헤더로 인정
        if all(k in found for k in ("branch", "item_name", "item_code")):
            header_row_idx = row_idx
            col_map = found
            break

    if header_row_idx is None:
        return {"success": 0, "skipped": 0,
                "errors": ["헤더를 찾을 수 없습니다. '지점명/상품명/품번' 컬럼명이 포함된 행이 있는지 확인해주세요."]}

    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]

    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []
    hq_adjustments = []
    debug_hq_log = []


    conn = get_conn()
    data_start_row = header_row_idx + 1
    # ── 이전 업로드 데이터 삭제 (지점 제한 있으면 해당 지점만, 없으면 전체 - 마스터) ──
    if restrict_branch:
        conn.execute("DELETE FROM raw_inventory WHERE branch_code=?", (restrict_branch,))
    else:
        conn.execute("DELETE FROM raw_inventory")
    # 이전 H/Q 반영 이력도 초기화 (재계산 기준점 리셋)
    old_hq_rows = conn.execute("SELECT * FROM hq_bonus_log").fetchall()
    old_hq_map = {f"{r['branch_code']}|{r['item_code']}": r["last_hq_total"] for r in old_hq_rows}
    conn.execute("DELETE FROM hq_bonus_log")
    conn.commit()
    

    for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        branch_col = col_map.get("branch")
        if branch_col is None or branch_col >= len(row) or not row[branch_col]:
            continue
        try:
            branch_name = str(row[col_map["branch"]]).strip()
            item_name = str(row[col_map["item_name"]]).strip() if row[col_map["item_name"]] else ""
            item_code = str(row[col_map["item_code"]]).strip() if row[col_map["item_code"]] else ""

            if not item_code:
                import hashlib
                import hashlib
                name_hash = hashlib.md5(item_name.encode('utf-8')).hexdigest()[:8]
                item_code = f"미지정_{name_hash}"

            qty_n = row[col_map["qty"]] if "qty" in col_map and col_map["qty"] < len(row) else None
            qty_h = row[col_map["h"]] if "h" in col_map and col_map["h"] < len(row) else None
            qty_q = row[col_map["q"]] if "q" in col_map and col_map["q"] < len(row) else None

            raw_quantity = int(float(str(qty_n))) if qty_n not in (None, "") else 0
            add_h = int(float(str(qty_h))) if qty_h not in (None, "") else 0
            add_q = int(float(str(qty_q))) if qty_q not in (None, "") else 0
            hq_total = add_h + add_q

            branch_code = (branch_map.get(branch_name)
                           or branch_map.get(branch_name.replace(" ", ""))
                           or branch_name)

            if restrict_branch and branch_code != restrict_branch:
                skipped += 1
                continue

            conn.execute("""
                INSERT INTO raw_inventory
                  (branch_code, branch_name, item_name, item_code, quantity, source, uploaded_at)
                VALUES (?, ?, ?, ?, ?, 'branch', ?)
                ON CONFLICT(branch_code, item_code) DO UPDATE SET
                  quantity=excluded.quantity,
                  item_name=excluded.item_name,
                  branch_name=excluded.branch_name,
                  source=excluded.source,
                  uploaded_at=excluded.uploaded_at
            """, (branch_code, branch_name, item_name, item_code, raw_quantity, now))

            # ⚠️ 신규 상품 자동 등록: items에 없으면 items + inventory(초기수량 0)에 추가
            existing_item = conn.execute(
                "SELECT id FROM items WHERE branch_code=? AND item_code=?",
                (branch_code, item_code)
            ).fetchone()
            if not existing_item:
                conn.execute("""
                    INSERT INTO items (branch_code, branch_name, item_name, item_code, created_at)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(branch_code, item_code) DO UPDATE SET item_name=excluded.item_name
                """, (branch_code, branch_name, item_name, item_code, now))
                conn.execute("""
                    INSERT INTO inventory (branch_code, item_name, item_code, quantity, last_updated)
                    VALUES (?, ?, ?, 0, ?)
                    ON CONFLICT(branch_code, item_code) DO NOTHING
                """, (branch_code, item_name, item_code, now))

            # ⚠️ 품번 없는 상품(미지정_): 업로드할 때마다 QR재고를 RAW재고와 항상 강제 동기화
            if item_code.startswith("미지정_"):
                conn.execute("""
                    UPDATE inventory SET quantity=?, last_updated=?
                    WHERE branch_code=? AND item_code=?
                """, (raw_quantity, now, branch_code, item_code))

            if hq_total != 0:
                hq_adjustments.append((branch_code, item_name, item_code, hq_total))
                debug_hq_log.append(f"{item_name}({item_code}): H={add_h}, Q={add_q}, 합계={hq_total}")

            success += 1
            
        except Exception as e:
            errors.append(f"행 {idx}: {str(e)[:50]}")
            skipped += 1

    conn.commit()
    conn.close()

    for branch_code, item_name, item_code, new_hq_total in hq_adjustments:
        conn2 = get_conn()
        prev_hq_total = old_hq_map.get(f"{branch_code}|{item_code}", 0)

        # 이전에 반영했던 만큼 빼고, 새 값을 더함 → 결과적으로 "덮어쓰기" 효과
        net_delta = new_hq_total - prev_hq_total

        if net_delta != 0:
            new_qty = adjust_quantity(branch_code, item_code, net_delta, absolute=False)
            conn2.execute(
                "INSERT INTO adjustment_log (branch_code, item_name, item_code, delta, result_quantity, adjusted_at) VALUES (?, ?, ?, ?, ?, ?)",
                (branch_code, item_name, item_code, net_delta, new_qty, now)
            )

        conn2.execute("""
            INSERT INTO hq_bonus_log (branch_code, item_code, last_hq_total, updated_at)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(branch_code, item_code) DO UPDATE SET
              last_hq_total=excluded.last_hq_total,
              updated_at=excluded.updated_at
        """, (branch_code, item_code, new_hq_total, now))

        conn2.commit()
        conn2.close()

    return {"success": success, "skipped": skipped, "errors": errors[:10],
            "header_row_used": header_row_idx,
            "hq_debug": debug_hq_log[:20],
            "col_map_debug": {k: v for k, v in col_map.items()}}

@app.post("/master/raw-upload/clear")
async def raw_upload_clear(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    conn = get_conn()
    conn.execute("DELETE FROM raw_inventory")
    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/raw-upload", status_code=303)


@app.get("/raw-upload", response_class=HTMLResponse)
async def raw_upload_redirect(session_token: str = Cookie(default=None)):
    """구 경로 호환용 - 역할별 분기"""
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)
    if user["role"] == "master":
        return RedirectResponse(url="/master/raw-upload", status_code=303)
    return RedirectResponse(url="/raw-branch", status_code=303)

@app.get("/raw-branch", response_class=HTMLResponse)
async def raw_branch_page(session_token: str = Cookie(default=None)):
    """유비플러스 재고 — 지점 계정: 본인 지점만 업로드/조회, 마스터: 전체 접근 가능 (마스터 전용 화면과 별개로 지점 시점 확인용)"""
    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    target_branch = user["branch_code"] if user["role"] == "branch" else ""

    conn = get_conn()
    if target_branch:
        raws = conn.execute(
            "SELECT * FROM raw_inventory WHERE branch_code=? ORDER BY item_code",
            (target_branch,)
        ).fetchall()
    else:
        raws = conn.execute("SELECT * FROM raw_inventory ORDER BY branch_code, item_code").fetchall()
    conn.close()

    rows_html = ""
    if not raws:
        rows_html = '<tr><td colspan="5" style="text-align:center;padding:20px;color:#888;">데이터 없음</td></tr>'
    else:
        for r in raws:
            rows_html += f"""
            <tr>
              <td>{r['branch_name']}</td>
              <td>{r['item_name']}</td>
              <td>{r['item_code']}</td>
              <td style="font-weight:bold;">{r['quantity']}</td>
              <td>{r['uploaded_at'][:10] if r['uploaded_at'] else '-'}</td>
            </tr>"""

    upload_note = "본인 지점 데이터만 업로드/조회됩니다." if target_branch else "마스터 계정 — 전체 지점 데이터가 조회됩니다. 업로드 시 엑셀 내 지점 컬럼 기준으로 반영됩니다."

    content = f"""
    <h2 style="margin-bottom:16px;">📤 유비플러스 재고</h2>
    <div class="card" style="background:#FFF7ED;border:1px solid #FCD34D;">
      <p style="font-size:13px;color:#92400E;">⚠️ {upload_note} H열/Q열 값은 QR재고에 자동 가산됩니다.</p>
    </div>
    <div class="card">
      <h3 style="margin-bottom:8px;">엑셀 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:12px;">
        컬럼 위치: <b>A=지점명 / B=상품명 / D=품번 / N=현재수량 / H,Q=QR재고 가산분</b> (1행 헤더)
      </p>
      <div style="display:flex;gap:8px;align-items:center;">
        <input type="file" id="rawFile" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="button" onclick="uploadRawBranch()">업로드</button>
      </div>
      <div id="uploadResult" style="display:none;margin-top:12px;padding:12px;
           border-radius:8px;font-size:13px;"></div>
      <script>
      async function uploadRawBranch() {{
        const file = document.getElementById('rawFile').files[0];
        if (!file) {{ alert('파일을 선택해주세요.'); return; }}
        const fd = new FormData();
        fd.append('file', file);
        const btn = event.target;
        btn.textContent = '업로드 중...';
        btn.disabled = true;
        try {{
          const res = await fetch('/raw-branch/upload', {{ method: 'POST', body: fd }});
          const data = await res.json();
          const box = document.getElementById('uploadResult');
          box.style.display = 'block';
          box.style.background = data.errors && data.errors.length ? '#FEF9C3' : '#D1FAE5';
          box.innerHTML = `<b>${{data.errors && data.errors.length ? '⚠️' : '✅'}} 업로드 완료</b><br>
          성공: <b style="color:#22C55E">${{data.success}}건</b> &nbsp;
          실패: <b style="color:#EF4444">${{data.skipped}}건</b>
          ${{data.errors && data.errors.length ? '<ul>' + data.errors.map(e=>`<li style="color:#EF4444;font-size:12px;">${{e}}</li>`).join('') + '</ul>' : ''}}`;
          setTimeout(() => location.reload(), 2000);
        }} catch(e) {{
          alert('업로드 중 오류가 발생했습니다.');
        }} finally {{
          btn.textContent = '업로드';
          btn.disabled = false;
        }}
      }}
      </script>
    </div>
    <div class="card">
      <h3 style="margin-bottom:12px;">현재 데이터 ({len(raws)}개)</h3>
      <table>
        <thead><tr>
          <th>지점명</th><th>상품명</th><th>품번</th><th>수량</th><th>업로드일</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "raw-branch"))


@app.post("/raw-branch/upload")
async def raw_branch_upload(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    """지점 계정 업로드 — 본인 지점 데이터만 반영 (엑셀에 다른 지점 있어도 무시)"""
    user = get_session(session_token)
    if not user:
        return {"success": 0, "skipped": 0, "errors": ["로그인이 필요합니다"]}

    result = await _process_raw_upload(file, restrict_branch=(user["branch_code"] if user["role"] == "branch" else None))
    return result


# ── 마스터 > QR 재고 업로드(초기화) ────────────────────

@app.get("/master/qr-init", response_class=HTMLResponse)
async def qr_init_page(session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    content = """
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master" style="color:#1E2761;">← 마스터</a>
      <h2>🔄 QR 재고 업로드</h2>
    </div>
    <div class="card" style="background:#FFF7ED;border:1px solid #FCD34D;">
      <p style="font-size:13px;color:#92400E;">
        ⚠️ 엑셀로 초기 수량을 업로드하면 기존 QR 재고 수량이 덮어쓰기됩니다.
      </p>
    </div>
    <div class="card">
      <h3 style="margin-bottom:8px;">엑셀 업로드</h3>
      <p style="color:#666;font-size:12px;margin-bottom:12px;">
        컬럼: <b>A=지점명 / B=상품명 / C=품번 / D=초기수량</b> (1행 헤더)
      </p>
      <div style="display:flex;gap:8px;align-items:center;">
        <input type="file" id="qrInitFile" accept=".xlsx,.xls" style="width:auto;flex:1;">
        <button class="btn" type="button" onclick="uploadQrInit()" id="qrInitBtn">업로드</button>
      </div>
      <div id="qrInitResult" style="display:none;margin-top:12px;padding:12px;
           border-radius:8px;font-size:13px;"></div>
      <script>
      async function uploadQrInit() {
        const file = document.getElementById('qrInitFile').files[0];
        if (!file) { alert('파일을 선택해주세요.'); return; }
        const fd = new FormData();
        fd.append('file', file);
        const btn = document.getElementById('qrInitBtn');
        let seconds = 0;
        btn.disabled = true;
        const timerInterval = setInterval(() => {
          seconds++;
          btn.textContent = `업로드 중... (${seconds}초째)`;
        }, 1000);
        btn.textContent = '업로드 중... (0초째)';
        try {
          const res = await fetch('/master/qr-init/upload-ajax', { method: 'POST', body: fd });
          const data = await res.json();
          const box = document.getElementById('qrInitResult');
          box.style.display = 'block';
          box.style.background = data.errors && data.errors.length ? '#FEF9C3' : '#D1FAE5';
          box.innerHTML = `<b>${data.errors && data.errors.length ? '⚠️' : '✅'} 업로드 완료</b><br>
          성공: <b style="color:#22C55E">${data.success}건</b> &nbsp;
          실패: <b style="color:#EF4444">${data.skipped}건</b>
          ${data.errors && data.errors.length ? '<ul>' + data.errors.map(e=>`<li style="color:#EF4444;font-size:12px;">${e}</li>`).join('') + '</ul>' : ''}`;
        } catch(e) {
          alert('업로드 중 오류가 발생했습니다.');
        } finally {
          clearInterval(timerInterval);
          btn.textContent = '업로드';
          btn.disabled = false;
        }
      }
      </script>
    </div>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/qr-init/upload-ajax")
async def qr_init_upload_ajax(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return {"success": 0, "skipped": 0, "errors": ["로그인이 필요합니다"]}

    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return {"success": 0, "skipped": 0, "errors": ["시트를 찾을 수 없습니다."]}

    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]

    conn = get_conn()
    now = datetime.now().isoformat()
    success, skipped, errors = 0, 0, []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        try:
            branch_name = str(row[0]).strip()
            item_name = str(row[1]).strip() if row[1] else ""
            item_code = str(row[2]).strip() if row[2] else ""
            item_code = str(row[2]).strip() if row[2] else ""
            if not item_code:
                import hashlib
                name_hash = hashlib.md5(item_name.encode('utf-8')).hexdigest()[:8]
                item_code = f"미지정_{name_hash}"
            init_qty = int(float(str(row[3]))) if row[3] is not None else 0
            branch_code = (branch_map.get(branch_name)
                           or branch_map.get(branch_name.replace(" ", ""))
                           or branch_name)

            conn.execute(
                """INSERT INTO inventory
                   (branch_code, item_name, item_code, quantity, last_updated)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(branch_code, item_code) DO UPDATE SET
                     quantity=excluded.quantity,
                     item_name=excluded.item_name,
                     last_updated=excluded.last_updated""",
                (branch_code, item_name, item_code, init_qty, now)
            )
            conn.execute(
                "INSERT INTO qr_init_log (branch_code, item_code, init_quantity, initialized_at) VALUES (?, ?, ?, ?)",
                (branch_code, item_code, init_qty, now)
            )
            success += 1
        except Exception as e:
            errors.append(f"행 {row_idx}: {str(e)[:50]}")
            skipped += 1
            continue

    conn.commit()
    conn.close()
    return {"success": success, "skipped": skipped, "errors": errors[:10]}


@app.post("/master/qr-init/upload")
async def qr_init_upload(
    session_token: str = Cookie(default=None),
    file: UploadFile = File(...)
):
    """구버전 호환용 - 폼 제출 방식 (리다이렉트만)"""
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    contents = await file.read()
    import io
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    if ws is None:
        return RedirectResponse(url="/master/qr-init", status_code=303)

    branch_map = {}
    branches = get_branches()
    for b in branches:
        branch_map[b["branch_name"]] = b["branch_code"]
        branch_map[b["branch_name"].replace(" ", "")] = b["branch_code"]
        branch_map[b["branch_code"]] = b["branch_code"]

    conn = get_conn()
    now = datetime.now().isoformat()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        try:
            branch_name = str(row[0]).strip()
            item_name = str(row[1]).strip() if row[1] else ""
            item_code = str(row[2]).strip() if row[2] else ""
            item_code = str(row[2]).strip() if row[2] else ""
            if not item_code:
                import hashlib
                name_hash = hashlib.md5(item_name.encode('utf-8')).hexdigest()[:8]
                item_code = f"미지정_{name_hash}"
            init_qty = int(float(str(row[3]))) if row[3] is not None else 0
            branch_code = (branch_map.get(branch_name)
                           or branch_map.get(branch_name.replace(" ", ""))
                           or branch_name)

            conn.execute(
                """INSERT INTO inventory
                   (branch_code, item_name, item_code, quantity, last_updated)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(branch_code, item_code) DO UPDATE SET
                     quantity=excluded.quantity,
                     item_name=excluded.item_name,
                     last_updated=excluded.last_updated""",
                (branch_code, item_name, item_code, init_qty, now)
            )
            conn.execute(
                "INSERT INTO qr_init_log (branch_code, item_code, init_quantity, initialized_at) VALUES (?, ?, ?, ?)",
                (branch_code, item_code, init_qty, now)
            )
        except Exception:
            continue

    conn.commit()
    conn.close()
    return RedirectResponse(url="/master/qr-init", status_code=303)


# ── 마스터 > QR 일괄 생성 (ZIP) ────────────────────────

@app.post("/master/qr/generate-bulk")
async def master_qr_generate_bulk(
    session_token: str = Cookie(default=None),
    branch_code: str = Form(...)
):
    import zipfile
    import io
    from urllib.parse import quote

    user = get_session(session_token)
    if not user:
        return RedirectResponse(url="/login", status_code=303)

    if user["role"] != "master":
        branch_code = user["branch_code"]  # 지점 계정은 본인 지점 QR만 생성 가능

    conn = get_conn()
    if branch_code == "ALL":
        items = conn.execute("SELECT * FROM items ORDER BY branch_code, item_name").fetchall()
        zip_name = "전체지점_QR"
    else:
        items = conn.execute(
            "SELECT * FROM items WHERE branch_code=?", (branch_code,)
        ).fetchall()
        zip_name = f"{branch_code}_QR"
    conn.close()

    if not items:
        return HTMLResponse(content=render_page(
            '<div class="card"><p>❌ 등록된 품목이 없습니다.</p>'
            '<a href="/qr">← 돌아가기</a></div>', user, "qr"))

    hostname_env = os.getenv("PUBLIC_SERVER_URL")
    if hostname_env:
        server_url = hostname_env
    else:
        hostname = socket.gethostbyname(socket.gethostname())
        server_url = f"http://{hostname}:{SERVER_PORT}"

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for it in items:
            for scan_type in ["IN", "OUT"]:
                img_bytes = generate_qr_bytes(
                    server_url, it["branch_code"],
                    it["item_code"], scan_type, it["item_name"]
                )
                filename = f"{it['branch_code']}_{it['item_code']}_{scan_type}.png"
                zf.writestr(f"{it['branch_code']}/{filename}", img_bytes)

    zip_buffer.seek(0)
    encoded_name = quote(f"{zip_name}.zip")
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}
    )

SAFETY_STOCK_EXTRA_CSS = """
<style>
.ss-table input[type="number"] { width: 100%; padding: 6px; box-sizing: border-box; }
.ss-upload-box { border: 2px dashed #93C5FD; border-radius: 10px; padding: 16px; background: #EFF6FF; }
.ss-autocomplete { position: relative; flex: 2; min-width: 180px; }
.ss-autocomplete-list { display: none; position: absolute; top: 100%; left: 0; right: 0; background: #fff;
  border: 1px solid #ddd; border-radius: 8px; max-height: 220px; overflow-y: auto; z-index: 50;
  box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
.ss-autocomplete-item { padding: 8px 12px; cursor: pointer; font-size: 13px; border-bottom: 1px solid #f0f0f0; }
.ss-autocomplete-item:hover { background: #EFF6FF; }
</style>
"""


@app.get("/master/purchase-order/safety-stock", response_class=HTMLResponse)
async def safety_stock_page(
    session_token: str = Cookie(default=None),
    search_item: str = "",
    filter_branch: str = "",
    sort_by: str = "item_name",
    sort_dir: str = "asc"
):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return RedirectResponse(url="/login", status_code=303)

    allowed_sort_cols = {"item_name", "branch_name", "qty", "updated_at"}
    if sort_by not in allowed_sort_cols:
        sort_by = "item_name"
    sort_dir_sql = "DESC" if sort_dir == "desc" else "ASC"

    conn = get_conn()
    query = "SELECT * FROM safety_stock WHERE 1=1"
    params: list = []
    if search_item:
        query += " AND (item_name LIKE ? OR item_code LIKE ?)"
        params.append(f"%{search_item}%")
        params.append(f"%{search_item}%")
    if filter_branch:
        query += " AND branch_name = ?"
        params.append(filter_branch)
    query += f" ORDER BY {sort_by} {sort_dir_sql} LIMIT 300"
    rows = conn.execute(query, params).fetchall()

    # 상품 검색선택 UX를 위해 product_master 전체 목록을 JS로 내려보냄 (지점명+상품명+품번)
    product_rows = conn.execute(
        "SELECT branch_name, item_name, item_code FROM product_master ORDER BY branch_name, item_name LIMIT 5000"
    ).fetchall()
    conn.close()

    branches = get_branches(branch_type='branch')

    branch_filter_options = '<option value="">전체 지점</option>'
    for b in branches:
        sel = "selected" if filter_branch == b["branch_name"] else ""
        branch_filter_options += f'<option value="{b["branch_name"]}" {sel}>{b["branch_name"]}</option>'

    branch_select_options = '<option value="">지점 선택</option>'
    for b in branches:
        branch_select_options += f'<option value="{b["branch_name"]}">{b["branch_name"]}</option>'

    product_list_json = json.dumps(
        [{"branch": p["branch_name"], "name": p["item_name"], "code": p["item_code"] or ""} for p in product_rows],
        ensure_ascii=False
    )

    rows_html = ""
    if not rows:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:20px;color:#888;">등록된 안전재고가 없습니다.</td></tr>'
    else:
        for r in rows:
            updated_display = (r["updated_at"] or "-")[:16] if r["updated_at"] else "-"
            rows_html += f"""
            <tr>
              <td>{r['branch_name'] or ''}</td>
              <td>{r['item_name'] or ''}</td>
              <td>{r['item_code'] or ''}</td>
              <td><input type="number" class="ss-qty" data-id="{r['id']}" value="{r['qty'] or 0}"></td>
              <td style="font-size:11px;color:#888;">{updated_display}</td>
              <td><button class="btn" style="font-size:12px;padding:6px 12px;" onclick="saveSsRow({r['id']})">저장</button></td>
            </tr>
            """

    content = f"""
    {SAFETY_STOCK_EXTRA_CSS}
    <div style="display:flex;align-items:center;gap:8px;margin-bottom:16px;">
      <a href="/master/purchase-tracking" style="color:#1E2761;">← 발주 주기 트래킹</a>
      <h2>🎯 안전재고 관리</h2>
    </div>

    <div class="card" style="background:#EFF6FF;border:1px solid #93C5FD;">
      <p style="font-size:13px;color:#1E40AF;">지점×상품별 안전재고 수량을 관리합니다. 발주서 자동생성 시 "현재고 &lt; 안전재고"인 상품이 발주 대상이 되며, 발주수량은 안전재고 수량 그대로 산정됩니다.</p>
    </div>

    <div class="card">
      <h3 style="margin-bottom:8px;">새 안전재고 수동 등록</h3>
      <div style="display:flex;gap:8px;flex-wrap:wrap;">
        <select id="ssBranch" style="flex:1;min-width:120px;" onchange="ssBranchChanged()">{branch_select_options}</select>
        <div class="ss-autocomplete">
          <input type="text" id="ssItemSearch" placeholder="상품명 검색 (지점 먼저 선택)" autocomplete="off" oninput="ssSearchProduct()" disabled>
          <div class="ss-autocomplete-list" id="ssAutocompleteList"></div>
        </div>
        <input type="number" id="ssQty" placeholder="안전재고 수량" style="flex:1;min-width:120px;">
        <button class="btn" type="button" onclick="addSsItem()">추가</button>
      </div>
      <div id="ssAddResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card ss-upload-box">
      <h3 style="margin-bottom:8px;">엑셀 업로드 (안전재고 일괄 등록/갱신)</h3>
      <p style="font-size:12px;color:#1E40AF;margin-bottom:8px;">
        헤더 2행 형식(1행 제목, 2행 컬럼명, 3행부터 데이터). 컬럼 위치: <b>B=지점 / D=상품명 / E=품번 / J=지점설정재고(발주 기준)</b><br>
        (지점+상품명 기준으로 이미 있으면 덮어씁니다.)
      </p>
      <input type="file" id="ssExcelFile" accept=".xlsx,.xls">
      <button class="btn" type="button" onclick="uploadSsExcel()">업로드</button>
      <div id="ssUploadResult" style="margin-top:8px;font-size:13px;"></div>
    </div>

    <div class="card">
      <form method="get" action="/master/purchase-order/safety-stock" style="display:flex;gap:8px;flex-wrap:wrap;margin-bottom:16px;">
        <input name="search_item" value="{search_item}" placeholder="상품명/품번 검색" style="flex:2;min-width:180px;">
        <select name="filter_branch" style="flex:1;min-width:120px;">{branch_filter_options}</select>
        <button class="btn" type="submit">검색</button>
        <a href="/master/purchase-order/safety-stock" style="padding:10px 14px;background:#eee;border-radius:8px;font-size:13px;text-decoration:none;color:#555;">초기화</a>
      </form>
      <p style="font-size:13px;color:#888;margin-bottom:12px;">{len(rows)}건 (최대 300건까지 표시)</p>
      <table class="ss-table">
        <thead><tr>
          <th style="cursor:pointer;" onclick="sortSs('branch_name')">지점 {'▲' if sort_by=='branch_name' and sort_dir=='asc' else ('▼' if sort_by=='branch_name' else '')}</th>
          <th style="cursor:pointer;" onclick="sortSs('item_name')">상품명 {'▲' if sort_by=='item_name' and sort_dir=='asc' else ('▼' if sort_by=='item_name' else '')}</th>
          <th>품번</th>
          <th style="cursor:pointer;" onclick="sortSs('qty')">안전재고수량 {'▲' if sort_by=='qty' and sort_dir=='asc' else ('▼' if sort_by=='qty' else '')}</th>
          <th style="cursor:pointer;" onclick="sortSs('updated_at')">수정일시 {'▲' if sort_by=='updated_at' and sort_dir=='asc' else ('▼' if sort_by=='updated_at' else '')}</th>
          <th></th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
    </div>

    <script>
      const PRODUCT_LIST = {product_list_json};
      let ssSelectedItem = null;

      function sortSs(col) {{
        const params = new URLSearchParams(window.location.search);
        const curSort = params.get('sort_by') || 'item_name';
        const curDir = params.get('sort_dir') || 'asc';
        const newDir = (curSort === col && curDir === 'asc') ? 'desc' : 'asc';
        params.set('sort_by', col);
        params.set('sort_dir', newDir);
        window.location.search = params.toString();
      }}

      function ssBranchChanged() {{
        const branch = document.getElementById('ssBranch').value;
        const searchInput = document.getElementById('ssItemSearch');
        searchInput.disabled = !branch;
        searchInput.value = '';
        ssSelectedItem = null;
        document.getElementById('ssAutocompleteList').style.display = 'none';
      }}

      let ssCurrentMatches = [];

      function ssSearchProduct() {{
        const branch = document.getElementById('ssBranch').value;
        const kw = document.getElementById('ssItemSearch').value.trim().toLowerCase();
        const listEl = document.getElementById('ssAutocompleteList');
        if (!branch || !kw) {{ listEl.style.display = 'none'; return; }}
        ssCurrentMatches = PRODUCT_LIST.filter(p => p.branch === branch &&
          (p.name.toLowerCase().includes(kw) || p.code.toLowerCase().includes(kw))).slice(0, 30);
        if (ssCurrentMatches.length === 0) {{
          listEl.innerHTML = '<div class="ss-autocomplete-item" style="color:#888;">일치하는 상품이 없습니다 (상품마스터 미등록 상품일 수 있음)</div>';
          listEl.style.display = 'block';
          return;
        }}
        listEl.innerHTML = ssCurrentMatches.map((p, idx) =>
          '<div class="ss-autocomplete-item" data-idx="' + idx + '">' +
          p.name + (p.code ? ' <span style="color:#888;">(' + p.code + ')</span>' : '') + '</div>'
        ).join('');
        listEl.style.display = 'block';
      }}

      document.getElementById('ssAutocompleteList').addEventListener('click', function(e) {{
        const item = e.target.closest('.ss-autocomplete-item[data-idx]');
        if (!item) return;
        const idx = parseInt(item.dataset.idx);
        const p = ssCurrentMatches[idx];
        if (!p) return;
        ssSelectedItem = p;
        document.getElementById('ssItemSearch').value = p.name + (p.code ? ' (' + p.code + ')' : '');
        document.getElementById('ssAutocompleteList').style.display = 'none';
      }});

      document.addEventListener('click', function(e) {{
        if (!e.target.closest('.ss-autocomplete')) {{
          document.getElementById('ssAutocompleteList').style.display = 'none';
        }}
      }});

      async function addSsItem() {{
        const branch = document.getElementById('ssBranch').value;
        const qty = parseInt(document.getElementById('ssQty').value);
        if (!branch || !ssSelectedItem || isNaN(qty)) {{
          alert('지점 선택 후 상품을 검색해서 선택하고, 안전재고수량을 입력하세요.');
          return;
        }}
        const res = await fetch('/master/purchase-order/safety-stock/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ branch_name: branch, item_name: ssSelectedItem.name, item_code: ssSelectedItem.code, qty: qty }})
        }});
        if (res.ok) {{ location.reload(); }} else {{
          const err = await res.json();
          document.getElementById('ssAddResult').innerText = '오류: ' + (err.detail || '추가 실패');
        }}
      }}

      async function saveSsRow(id) {{
        const qtyEl = document.querySelector('.ss-qty[data-id="' + id + '"]');
        const res = await fetch('/master/purchase-order/safety-stock/save', {{
          method: 'POST', headers: {{ 'Content-Type': 'application/json' }},
          body: JSON.stringify({{ id: id, qty: parseInt(qtyEl.value) || 0 }})
        }});
        if (res.ok) {{ alert('저장되었습니다.'); }} else {{ alert('저장 실패'); }}
      }}

      async function uploadSsExcel() {{
        const fileEl = document.getElementById('ssExcelFile');
        if (!fileEl.files.length) {{ alert('파일을 선택하세요.'); return; }}
        const formData = new FormData();
        formData.append('file', fileEl.files[0]);
        document.getElementById('ssUploadResult').innerText = '업로드 중...';
        const res = await fetch('/master/purchase-order/safety-stock/upload', {{
          method: 'POST', body: formData
        }});
        const data = await res.json();
        if (res.ok) {{
          document.getElementById('ssUploadResult').innerText = '완료: ' + data.inserted + '건 등록/갱신 (건너뜀: ' + data.skipped + '건)';
          setTimeout(() => location.reload(), 1200);
        }} else {{
          document.getElementById('ssUploadResult').innerText = '오류: ' + (data.detail || '업로드 실패');
        }}
      }}
    </script>
    """
    return HTMLResponse(content=render_page(content, user, "master"))


@app.post("/master/purchase-order/safety-stock/save")
async def safety_stock_save(request: Request, session_token: str = Cookie(default=None)):
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    data = await request.json()
    row_id = data.get("id")
    branch_name = (data.get("branch_name") or "").strip()
    item_name = (data.get("item_name") or "").strip()
    item_code = (data.get("item_code") or "").strip()
    qty = data.get("qty", 0)

    now = datetime.now().isoformat()
    conn = get_conn()
    if row_id:
        conn.execute(
            "UPDATE safety_stock SET qty=?, updated_at=? WHERE id=?",
            (qty, now, row_id)
        )
    else:
        if not branch_name or not item_name:
            conn.close()
            return JSONResponse(status_code=400, content={"detail": "지점과 상품명을 입력하세요."})
        # 구 컬럼 safety_qty가 NOT NULL로 남아있을 수 있어 qty와 동일 값을 함께 채워 방어
        conn.execute("""
            INSERT INTO safety_stock (branch_name, item_name, item_code, qty, safety_qty, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT (branch_name, item_name) DO UPDATE SET
                item_code=excluded.item_code, qty=excluded.qty,
                safety_qty=excluded.safety_qty, updated_at=excluded.updated_at
        """, (branch_name, item_name, item_code, qty, qty, now))
    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok"})


@app.post("/master/purchase-order/safety-stock/upload")
async def safety_stock_upload(request: Request, session_token: str = Cookie(default=None)):
    """고정 컬럼 위치 기반 파싱: B=지점(idx1) D=상품명(idx3) E=품번(idx4) J=지점설정재고(idx9)
    1행=제목, 2행=헤더, 3행부터 데이터."""
    user = get_session(session_token)
    if not user or user["role"] != "master":
        return JSONResponse(status_code=403, content={"detail": "권한이 없습니다."})

    form = await request.form()
    file = form.get("file")
    if not file:
        return JSONResponse(status_code=400, content={"detail": "파일이 없습니다."})

    import openpyxl
    from io import BytesIO
    raw = await file.read()
    wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    ws = wb.active

    COL_BRANCH = 1   # B
    COL_ITEM_NAME = 3   # D
    COL_ITEM_CODE = 4   # E
    COL_QTY = 9   # J

    now = datetime.now().isoformat()
    conn = get_conn()
    inserted = 0
    skipped = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or len(row) <= COL_QTY:
            skipped += 1
            continue

        branch_name = str(row[COL_BRANCH] or "").strip()
        item_name = str(row[COL_ITEM_NAME] or "").strip()
        if not branch_name or not item_name:
            skipped += 1
            continue

        item_code = str(row[COL_ITEM_CODE] or "").strip()
        qty_raw = row[COL_QTY]
        try:
            qty = int(float(qty_raw)) if qty_raw not in (None, "") else 0
        except (ValueError, TypeError):
            skipped += 1
            continue

        conn.execute("""
            INSERT INTO safety_stock (branch_name, item_name, item_code, qty, safety_qty, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT (branch_name, item_name) DO UPDATE SET
                item_code=excluded.item_code, qty=excluded.qty,
                safety_qty=excluded.safety_qty, updated_at=excluded.updated_at
        """, (branch_name, item_name, item_code, qty, qty, now))
        inserted += 1

    conn.commit()
    conn.close()
    return JSONResponse(content={"status": "ok", "inserted": inserted, "skipped": skipped})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=SERVER_PORT)